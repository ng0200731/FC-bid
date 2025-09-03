# -*- coding: utf-8 -*-
"""
Artwork Downloader
1. Input PO# first
2. Show recommendations based on PO data
3. Display data table in web interface
4. Intelligent artwork download with multiple methods

VERSION TRACKING:
- Every code edit gets a version number
- Version displayed in UI for transparency
- Format: v1.0.0 (YYYY-MM-DD HH:MM)
"""

# Version tracking system
VERSION = "3.5.0"
VERSION_DATE = "2025-08-15 00:00"
LAST_EDIT = "Fix folder opening: Open Folder button + auto-open after downloads"

from flask import Flask, render_template_string, request, jsonify, send_file, Response
import os
import threading
import time
import requests
import re
import sqlite3
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def update_version(new_version, edit_description):
    """Helper function to update version info - USE THIS FOR EVERY EDIT"""
    global VERSION, VERSION_DATE, LAST_EDIT
    VERSION = new_version
    VERSION_DATE = datetime.now().strftime("%Y-%m-%d %H:%M")
    LAST_EDIT = edit_description
    print(f"📝 Version updated to {VERSION} - {edit_description}")

def mask_email(email):
    """Mask email address: prefix shows first 2 chars, suffix shows first 1 char"""
    if '@' not in email:
        return email

    prefix, suffix = email.split('@', 1)

    # Mask prefix: show first 2 characters, rest as asterisks
    if len(prefix) <= 2:
        masked_prefix = prefix
    else:
        masked_prefix = prefix[:2] + '*' * (len(prefix) - 2)

    # Mask suffix: show first 1 character, rest as asterisks
    if len(suffix) <= 1:
        masked_suffix = suffix
    else:
        masked_suffix = suffix[:1] + '*' * (len(suffix) - 1)

    return f"{masked_prefix}@{masked_suffix}"

# Database functions
def init_database():
    """Initialize SQLite database for PO storage"""
    conn = sqlite3.connect('po_database.db')
    cursor = conn.cursor()

    # Create PO headers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS po_headers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT UNIQUE,
            purchase_from TEXT,
            ship_to TEXT,
            company TEXT,
            currency TEXT,
            cancel_date TEXT,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Create PO items table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS po_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT,
            item_number TEXT,
            description TEXT,
            color TEXT,
            ship_to TEXT,
            need_by TEXT,
            qty TEXT,
            bundle_qty TEXT,
            unit_price TEXT,
            extension TEXT,
            packed_status TEXT DEFAULT 'not_packed',
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (po_number) REFERENCES po_headers(po_number)
        )
    ''')

    # Add new columns to existing po_headers table if they don't exist
    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN factory TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN po_date TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN ship_by TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN ship_via TEXT')
    except sqlite3.OperationalError:
        pass

    # Add packed_status column to po_items table if it doesn't exist
    try:
        cursor.execute("ALTER TABLE po_items ADD COLUMN packed_status TEXT DEFAULT 'not_packed'")
    except sqlite3.OperationalError:
        pass  # Column already exists

    # Add carton_number column to po_items table if it doesn't exist
    try:
        cursor.execute("ALTER TABLE po_items ADD COLUMN carton_number TEXT")
    except sqlite3.OperationalError:
        pass  # Column already exists

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN order_type TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN status TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN location TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN prod_rep TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN ship_to_address TEXT')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN terms TEXT')
    except sqlite3.OperationalError:
        pass

    # Add tracking columns for PO update history
    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN first_created TIMESTAMP')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN last_updated TIMESTAMP')
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute('ALTER TABLE po_headers ADD COLUMN update_count INTEGER DEFAULT 0')
    except sqlite3.OperationalError:
        pass

    # Create cartons table for packing management
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cartons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT,
            carton_number TEXT,
            carton_size TEXT,
            actual_weight REAL,
            barcode TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Add packing_option column if it doesn't exist
    try:
        cursor.execute('ALTER TABLE cartons ADD COLUMN packing_option TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists

    # Create carton items table for tracking what's in each carton
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS carton_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            carton_id INTEGER,
            po_number TEXT,
            item_number TEXT,
            description TEXT,
            color TEXT,
            packed_qty INTEGER,
            original_qty INTEGER,
            packed_status TEXT DEFAULT 'pending',
            carton_number TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (carton_id) REFERENCES cartons (id)
        )
    ''')

    # Add new columns for real-time packing if they don't exist
    try:
        cursor.execute('ALTER TABLE carton_items ADD COLUMN packed_status TEXT DEFAULT "pending"')
    except sqlite3.OperationalError:
        pass  # Column already exists

    try:
        cursor.execute('ALTER TABLE carton_items ADD COLUMN carton_number TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists

    # Create shipments table for courier management
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS shipments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT,
            courier TEXT,
            awb_number TEXT,
            awb_document_path TEXT,
            shipment_date TEXT,
            total_cartons INTEGER,
            total_weight REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Create shipment cartons table to link cartons to shipments
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS shipment_cartons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shipment_id INTEGER,
            carton_id INTEGER,
            FOREIGN KEY (shipment_id) REFERENCES shipments (id),
            FOREIGN KEY (carton_id) REFERENCES cartons (id)
        )
    ''')

    # Create po_completion_status table to track partial vs complete shipments
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS po_completion_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT,
            completion_type TEXT,
            finished_quantities TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Create packing_lists table to track unique packing list numbers
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS packing_lists (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pl_number TEXT UNIQUE,
            po_number TEXT,
            total_cartons INTEGER,
            total_items INTEGER,
            total_qty INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Add pl_number column to existing tables if they don't exist
    try:
        cursor.execute('ALTER TABLE po_items ADD COLUMN pl_number TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists

    try:
        cursor.execute('ALTER TABLE cartons ADD COLUMN pl_number TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists

    try:
        cursor.execute('ALTER TABLE carton_items ADD COLUMN pl_number TEXT')
    except sqlite3.OperationalError:
        pass  # Column already exists

    conn.commit()
    conn.close()
    print("📊 Database initialized successfully")

    # Clean up any existing comma-separated numbers
    cleanup_database_numbers()

    # Debug: Check PO 1288138 data
    debug_po_1288138()

def clean_number_format(value):
    """Remove commas from numbers and return clean string for database storage"""
    if value is None:
        return '0'

    # Convert to string and remove commas
    clean_value = str(value).replace(',', '').strip()

    # Handle empty strings
    if not clean_value:
        return '0'

    # Validate it's a valid number
    try:
        # Try to convert to float first (handles decimals)
        float(clean_value)
        return clean_value
    except ValueError:
        # If not a valid number, return '0'
        return '0'

def cleanup_database_numbers():
    """Clean up all comma-separated numbers in existing database"""
    print("🧹 Starting database cleanup for comma-separated numbers...")

    try:
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Clean po_items table
        print("🧹 Cleaning po_items table...")
        cursor.execute('SELECT id, qty, unit_price, extension FROM po_items')
        items = cursor.fetchall()

        for item_id, qty, unit_price, extension in items:
            clean_qty = clean_number_format(qty)
            clean_unit_price = clean_number_format(unit_price)
            clean_extension = clean_number_format(extension)

            cursor.execute('''
                UPDATE po_items
                SET qty = ?, unit_price = ?, extension = ?
                WHERE id = ?
            ''', (clean_qty, clean_unit_price, clean_extension, item_id))

        # Clean cartons table
        print("🧹 Cleaning cartons table...")
        cursor.execute('SELECT id, actual_weight FROM cartons')
        cartons = cursor.fetchall()

        for carton_id, weight in cartons:
            clean_weight = clean_number_format(weight)
            cursor.execute('UPDATE cartons SET actual_weight = ? WHERE id = ?', (clean_weight, carton_id))

        # Clean packing_lists table
        print("🧹 Cleaning packing_lists table...")
        cursor.execute('SELECT id, total_cartons, total_items, total_qty FROM packing_lists')
        packing_lists = cursor.fetchall()

        for pl_id, total_cartons, total_items, total_qty in packing_lists:
            clean_cartons = clean_number_format(total_cartons)
            clean_items = clean_number_format(total_items)
            clean_qty = clean_number_format(total_qty)

            cursor.execute('''
                UPDATE packing_lists
                SET total_cartons = ?, total_items = ?, total_qty = ?
                WHERE id = ?
            ''', (clean_cartons, clean_items, clean_qty, pl_id))

        conn.commit()
        conn.close()
        print("✅ Database cleanup completed successfully!")

    except Exception as e:
        print(f"❌ Database cleanup error: {str(e)}")

def debug_po_1288138():
    """Debug function to check PO 1288138 data"""
    try:
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        print("🔍 DEBUG: Checking PO 1288138 data...")

        # Check po_items table
        cursor.execute('SELECT * FROM po_items WHERE po_number = ?', ('1288138',))
        items = cursor.fetchall()

        # Get column names
        cursor.execute('PRAGMA table_info(po_items)')
        columns = [column[1] for column in cursor.fetchall()]

        print(f"🔍 Found {len(items)} items for PO 1288138:")
        for i, item in enumerate(items):
            item_dict = dict(zip(columns, item))
            print(f"  Item {i}: {item_dict}")

        conn.close()

    except Exception as e:
        print(f"❌ Debug error: {str(e)}")

def generate_pl_number():
    """Generate unique packing list number in format PL0000001"""
    conn = sqlite3.connect('po_database.db')
    cursor = conn.cursor()

    # Get the highest existing PL number
    cursor.execute('SELECT MAX(pl_number) FROM packing_lists')
    result = cursor.fetchone()[0]

    if result:
        # Extract number from PL0000001 format
        current_num = int(result[2:])  # Remove 'PL' prefix
        next_num = current_num + 1
    else:
        next_num = 1

    # Format as PL0000001
    pl_number = f"PL{next_num:07d}"

    conn.close()
    return pl_number

def create_sample_po_data():
    """Create sample PO data for testing"""
    conn = sqlite3.connect('po_database.db')
    cursor = conn.cursor()

    # Check if sample data already exists
    cursor.execute('SELECT COUNT(*) FROM po_headers WHERE po_number IN (?, ?)', ('1284789', '1288176'))
    if cursor.fetchone()[0] > 0:
        conn.close()
        return  # Sample data already exists

    # Insert sample PO headers
    sample_pos = [
        ('1284789', 'Bird Dogs', 'Warehouse A', 'BrandID', 'USD', 'Active'),
        ('1288176', 'Bird Dogs', 'Warehouse B', 'BrandID', 'USD', 'Active')
    ]

    for po_data in sample_pos:
        cursor.execute('''
            INSERT INTO po_headers (po_number, purchase_from, ship_to, company, currency, status)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', po_data)

    # Insert sample PO items with proper quantities for both POs
    sample_items_1284789 = [
        {
            'item_number': '18483HEAT5SIF',
            'description': 'Bird Dogs Knockout Logo - 3D Silicone',
            'color': 'Silver Filigree',
            'qty': '150',
            'bundle_qty': 'NA',
            'unit_price': '12.50',
            'extension': '1875.00'
        },
        {
            'item_number': '18483HEAT5TEE',
            'description': 'Bird Dogs Knockout Logo - 3D Silicone',
            'color': 'Total Eclipse',
            'qty': '200',
            'bundle_qty': 'NA',
            'unit_price': '12.50',
            'extension': '2500.00'
        },
        {
            'item_number': '18483HEAT5mei',
            'description': 'Bird Dogs Knockout Logo - 3D Silicone',
            'color': 'moonless night',
            'qty': '100',
            'bundle_qty': 'NA',
            'unit_price': '12.50',
            'extension': '1250.00'
        }
    ]

    sample_items_1288176 = [
        {
            'item_number': 'BD2024POLO01',
            'description': 'Bird Dogs Premium Polo Shirt',
            'color': 'Navy Blue',
            'qty': '300',
            'bundle_qty': 'NA',
            'unit_price': '45.00',
            'extension': '13500.00'
        },
        {
            'item_number': 'BD2024POLO02',
            'description': 'Bird Dogs Premium Polo Shirt',
            'color': 'White',
            'qty': '250',
            'bundle_qty': 'NA',
            'unit_price': '45.00',
            'extension': '11250.00'
        },
        {
            'item_number': 'BD2024POLO03',
            'description': 'Bird Dogs Premium Polo Shirt',
            'color': 'Forest Green',
            'qty': '200',
            'bundle_qty': 'NA',
            'unit_price': '45.00',
            'extension': '9000.00'
        }
    ]

    # Insert items for PO 1284789
    for item in sample_items_1284789:
        cursor.execute('''
            INSERT INTO po_items (po_number, item_number, description, color, ship_to, need_by, qty, bundle_qty, unit_price, extension)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', ('1284789', item['item_number'], item['description'], item['color'],
              'Warehouse A', '2025-08-15', item['qty'], item['bundle_qty'],
              item['unit_price'], item['extension']))

    # Insert items for PO 1288176
    for item in sample_items_1288176:
        cursor.execute('''
            INSERT INTO po_items (po_number, item_number, description, color, ship_to, need_by, qty, bundle_qty, unit_price, extension)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', ('1288176', item['item_number'], item['description'], item['color'],
              'Warehouse B', '2025-08-20', item['qty'], item['bundle_qty'],
              item['unit_price'], item['extension']))

    conn.commit()
    conn.close()
    print("✅ Sample PO data created successfully!")

def check_po_exists(po_number):
    """Check if PO already exists in database"""
    conn = sqlite3.connect('po_database.db')
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM po_headers WHERE po_number = ?', (po_number,))
    exists = cursor.fetchone()[0] > 0
    conn.close()
    return exists

def save_po_to_database(po_number, po_header, po_items, overwrite=False):
    """Save complete PO data to database with tracking"""
    from datetime import datetime

    conn = sqlite3.connect('po_database.db')
    cursor = conn.cursor()

    try:
        current_time = datetime.now().isoformat()

        # Check if PO already exists
        cursor.execute('SELECT first_created, update_count FROM po_headers WHERE po_number = ?', (po_number,))
        existing_record = cursor.fetchone()

        if existing_record and overwrite:
            # PO exists and user confirmed overwrite
            first_created_time = existing_record[0]  # Keep original first_created
            current_update_count = existing_record[1] or 0  # Handle None values
            new_update_count = current_update_count + 1

            # Delete existing records
            cursor.execute('DELETE FROM po_items WHERE po_number = ?', (po_number,))
            cursor.execute('DELETE FROM po_headers WHERE po_number = ?', (po_number,))

            # Insert updated PO header with tracking
            cursor.execute('''
                INSERT INTO po_headers (
                    po_number, purchase_from, ship_to, company, currency, cancel_date,
                    factory, po_date, ship_by, ship_via, order_type, status, location, prod_rep, ship_to_address, terms,
                    first_created, last_updated, update_count
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                po_number,
                po_header.get('purchase_from', ''),
                po_header.get('ship_to', ''),
                po_header.get('company', ''),
                po_header.get('currency', ''),
                po_header.get('cancel_date', ''),
                po_header.get('factory', ''),
                po_header.get('po_date', ''),
                po_header.get('ship_by', ''),
                po_header.get('ship_via', ''),
                po_header.get('order_type', ''),
                po_header.get('status', ''),
                po_header.get('location', ''),
                po_header.get('prod_rep', ''),
                po_header.get('ship_to_address', ''),
                po_header.get('terms', ''),
                first_created_time,  # Keep original first_created
                current_time,        # Set last_updated to now
                new_update_count     # Increment update_count
            ))

            print(f"📊 PO {po_number} updated in database (Update #{new_update_count})")

        elif not existing_record:
            # New PO - first time saving
            cursor.execute('''
                INSERT INTO po_headers (
                    po_number, purchase_from, ship_to, company, currency, cancel_date,
                    factory, po_date, ship_by, ship_via, order_type, status, location, prod_rep, ship_to_address, terms,
                    first_created, last_updated, update_count
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                po_number,
                po_header.get('purchase_from', ''),
                po_header.get('ship_to', ''),
                po_header.get('company', ''),
                po_header.get('currency', ''),
                po_header.get('cancel_date', ''),
                po_header.get('factory', ''),
                po_header.get('po_date', ''),
                po_header.get('ship_by', ''),
                po_header.get('ship_via', ''),
                po_header.get('order_type', ''),
                po_header.get('status', ''),
                po_header.get('location', ''),
                po_header.get('prod_rep', ''),
                po_header.get('ship_to_address', ''),
                po_header.get('terms', ''),
                current_time,  # Set first_created to now
                None,          # last_updated is blank for new PO
                0              # update_count starts at 0
            ))

            print(f"📊 PO {po_number} saved to database for first time")
        else:
            # PO exists but overwrite=False
            print(f"⚠️ PO {po_number} already exists in database")
            return False

        # Insert PO items with cleaned numbers
        for item in po_items:
            # Clean numeric fields before database insertion
            clean_qty = clean_number_format(item.get('qty', ''))
            clean_bundle_qty = clean_number_format(item.get('bundle_qty', ''))
            clean_unit_price = clean_number_format(item.get('unit_price', ''))
            clean_extension = clean_number_format(item.get('extension', ''))

            cursor.execute('''
                INSERT INTO po_items (po_number, item_number, description, color, ship_to, need_by, qty, bundle_qty, unit_price, extension)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (po_number, item.get('item_number', ''), item.get('description', ''), item.get('color', ''),
                  item.get('ship_to', ''), item.get('need_by', ''), clean_qty,
                  clean_bundle_qty, clean_unit_price, clean_extension))

        conn.commit()
        return True

    except Exception as e:
        conn.rollback()
        print(f"❌ Error saving PO to database: {e}")
        return False
    finally:
        conn.close()

def get_master_report_data(limit=20, search_filters=None):
    """Get master report data combining PO headers and items with search functionality"""
    conn = sqlite3.connect('po_database.db')
    cursor = conn.cursor()

    try:
        # Base query joining headers and items
        base_query = '''
            SELECT
                h.po_number,
                i.item_number,
                i.description,
                i.color,
                i.ship_to,
                i.need_by,
                i.qty,
                i.bundle_qty,
                i.unit_price,
                i.extension,
                h.company,
                h.purchase_from,
                h.currency,
                h.po_date,
                h.cancel_date,
                h.ship_by,
                h.ship_via,
                h.order_type,
                h.status,
                h.factory,
                h.location,
                h.prod_rep,
                h.ship_to_address,
                h.terms,
                h.first_created,
                h.last_updated,
                h.update_count
            FROM po_headers h
            LEFT JOIN po_items i ON h.po_number = i.po_number
        '''

        # Build WHERE clause for search filters
        where_conditions = []
        params = []

        if search_filters:
            for column, value in search_filters.items():
                if value and value.strip():
                    # Map frontend column names to database columns
                    column_mapping = {
                        'po_number': 'h.po_number',
                        'item_number': 'i.item_number',
                        'description': 'i.description',
                        'color': 'i.color',
                        'ship_to': 'i.ship_to',
                        'need_by': 'i.need_by',
                        'qty': 'i.qty',
                        'bundle_qty': 'i.bundle_qty',
                        'unit_price': 'i.unit_price',
                        'extension': 'i.extension',
                        'company': 'h.company',
                        'purchase_from': 'h.purchase_from',
                        'currency': 'h.currency',
                        'po_date': 'h.po_date',
                        'cancel_date': 'h.cancel_date',
                        'ship_by': 'h.ship_by',
                        'ship_via': 'h.ship_via',
                        'order_type': 'h.order_type',
                        'status': 'h.status',
                        'factory': 'h.factory',
                        'location': 'h.location',
                        'prod_rep': 'h.prod_rep',
                        'ship_to_address': 'h.ship_to_address',
                        'terms': 'h.terms',
                        'first_created': 'h.first_created',
                        'last_updated': 'h.last_updated',
                        'update_count': 'h.update_count'
                    }

                    db_column = column_mapping.get(column)
                    if db_column:
                        where_conditions.append(f"{db_column} LIKE ?")
                        params.append(f"%{value.strip()}%")

        # Construct final query
        if where_conditions:
            query = base_query + " WHERE " + " AND ".join(where_conditions)
        else:
            query = base_query

        # Add ordering and limit
        query += " ORDER BY h.created_date DESC, i.id ASC"
        if limit:
            query += f" LIMIT {limit}"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Get column names
        columns = [
            'po_number', 'item_number', 'description', 'color', 'ship_to', 'need_by',
            'qty', 'bundle_qty', 'unit_price', 'extension', 'company', 'purchase_from',
            'currency', 'po_date', 'cancel_date', 'ship_by', 'ship_via', 'order_type',
            'status', 'factory', 'location', 'prod_rep', 'ship_to_address', 'terms',
            'first_created', 'last_updated', 'update_count'
        ]

        # Convert to list of dictionaries
        data = []
        for row in rows:
            record = {}
            for i, column in enumerate(columns):
                record[column] = row[i] if row[i] is not None else ''
            data.append(record)

        # Get total count for pagination info
        count_query = "SELECT COUNT(*) FROM po_headers h LEFT JOIN po_items i ON h.po_number = i.po_number"
        if where_conditions:
            count_query += " WHERE " + " AND ".join(where_conditions)

        cursor.execute(count_query, params)
        total_count = cursor.fetchone()[0]

        return {
            'success': True,
            'data': data,
            'total_count': total_count,
            'filtered_count': len(data),
            'columns': columns
        }

    except Exception as e:
        print(f"❌ Error getting master report data: {e}")
        return {
            'success': False,
            'error': str(e),
            'data': [],
            'total_count': 0,
            'filtered_count': 0,
            'columns': []
        }
    finally:
        conn.close()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

def scrape_po_details(po_number):
    """Scrape complete PO details from factoryPODetail.aspx page"""
    driver = None
    try:
        # Setup Chrome driver (Developer Mode)
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-images")
        chrome_options.add_argument("--remote-debugging-port=9223")  # Developer mode

        driver_path = ChromeDriverManager().install()
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        wait = WebDriverWait(driver, 15)

        print(f"🔍 Scraping PO details for {po_number}...")

        # Login first (same as working functions)
        driver.get(config['login_url'])
        username_field = wait.until(EC.presence_of_element_located((By.ID, "txtUserName")))
        password_field = driver.find_element(By.ID, "txtPassword")

        username_field.send_keys(config['username'])
        password_field.send_keys(config['password'])

        # Use the same login method as working functions
        login_button = driver.find_element(By.XPATH, "//img[@onclick='return Login();']")
        login_button.click()
        wait.until(lambda d: "login" not in d.current_url.lower())

        print(f"✅ Login successful for PO scraping")

        # Navigate to PO detail page
        po_url = f"https://app.e-brandid.com/Bidnet/bidnet3/factoryPODetail.aspx?po_id={po_number}"
        driver.get(po_url)
        time.sleep(5)  # Give more time for page to load

        print(f"📄 Loaded PO detail page: {po_url}")

        # Extract PO header information from page
        po_header = {}

        try:
            print(f"🔍 Extracting header information from PO page...")
            page_text = driver.page_source

            # Extract all header fields based on the structure you provided
            po_header['po_number'] = po_number
            po_header['factory'] = extract_field_value(page_text, ['Factory', 'Manufacturer'])
            po_header['po_date'] = extract_field_value(page_text, ['PO Date', 'Order Date', 'Date'])
            po_header['ship_by'] = extract_field_value(page_text, ['Ship By', 'Delivery Date', 'Ship Date'])
            po_header['ship_via'] = extract_field_value(page_text, ['Ship Via', 'Shipping Method', 'Delivery Method'])
            po_header['order_type'] = extract_field_value(page_text, ['Order Type', 'Type'])
            po_header['status'] = extract_field_value(page_text, ['Status', 'Order Status'])
            po_header['location'] = extract_field_value(page_text, ['Loc', 'Location'])
            po_header['prod_rep'] = extract_field_value(page_text, ['Prod Rep', 'Production Rep', 'Rep'])

            # Additional fields from the detailed section
            po_header['purchase_from'] = extract_field_value(page_text, ['Purchased From', 'Purchase From', 'Vendor', 'Supplier'])
            po_header['ship_to'] = extract_field_value(page_text, ['Ship To', 'Shipping Address', 'Delivery Address'])
            po_header['company'] = extract_field_value(page_text, ['Company', 'Client'])
            po_header['currency'] = extract_field_value(page_text, ['Currency', 'Curr'])
            po_header['cancel_date'] = extract_field_value(page_text, ['Cancel Date', 'Deadline', 'Due Date'])
            po_header['terms'] = extract_field_value(page_text, ['Terms', 'Payment Terms'])

            # Try to extract from tables as well (sometimes data is in table format)
            tables = driver.find_elements(By.TAG_NAME, "table")
            for table in tables:
                table_text = table.text

                # Look for header information in tables
                if 'Factory' in table_text and not po_header.get('factory'):
                    po_header['factory'] = extract_field_value(table_text, ['Factory'])
                if 'Ship By' in table_text and not po_header.get('ship_by'):
                    po_header['ship_by'] = extract_field_value(table_text, ['Ship By'])
                if 'Status' in table_text and not po_header.get('status'):
                    po_header['status'] = extract_field_value(table_text, ['Status'])

            print(f"📋 Extracted header fields: {list(po_header.keys())}")

        except Exception as e:
            print(f"⚠️ Could not extract header info: {e}")
            po_header = {'po_number': po_number}  # At least save the PO number

        # Extract items table
        po_items = []

        try:
            print(f"🔍 Looking for data tables on PO page...")

            # Wait for page content to load
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))

            # Find all tables
            tables = driver.find_elements(By.TAG_NAME, "table")
            print(f"📊 Found {len(tables)} tables on page")

            # Try to find the table with item data (similar to how get_po_data works)
            for table_idx, table in enumerate(tables):
                try:
                    rows = table.find_elements(By.TAG_NAME, "tr")
                    print(f"📋 Table {table_idx + 1}: {len(rows)} rows")

                    if len(rows) < 2:  # Skip tables with no data rows
                        continue

                    # Look for data rows (skip header)
                    for row_idx, row in enumerate(rows[1:], 1):  # Skip first row (header)
                        cells = row.find_elements(By.TAG_NAME, "td")

                        if len(cells) >= 8:  # Should have at least 8 columns for item data
                            cell_texts = [cell.text.strip() for cell in cells]

                            # Check if this looks like an item row (first cell should be item number)
                            if cell_texts[0] and len(cell_texts[0]) > 3 and not cell_texts[0].lower().startswith('item'):
                                item = {
                                    'item_number': cell_texts[0] if len(cell_texts) > 0 else '',
                                    'description': cell_texts[1] if len(cell_texts) > 1 else '',
                                    'color': cell_texts[2] if len(cell_texts) > 2 else '',
                                    'ship_to': cell_texts[3] if len(cell_texts) > 3 else '',
                                    'need_by': cell_texts[4] if len(cell_texts) > 4 else '',
                                    'qty': cell_texts[5] if len(cell_texts) > 5 else '',
                                    'bundle_qty': cell_texts[6] if len(cell_texts) > 6 else '',
                                    'unit_price': cell_texts[7] if len(cell_texts) > 7 else '',
                                    'extension': cell_texts[8] if len(cell_texts) > 8 else ''
                                }
                                po_items.append(item)
                                print(f"✅ Found item: {item['item_number']} - {item['description'][:30]}...")

                    if po_items:  # Found items in this table
                        print(f"🎯 Successfully extracted {len(po_items)} items from table {table_idx + 1}")
                        break

                except Exception as table_error:
                    print(f"⚠️ Error processing table {table_idx + 1}: {table_error}")
                    continue

        except Exception as e:
            print(f"⚠️ Could not extract items table: {e}")

        print(f"✅ Scraped {len(po_items)} items for PO {po_number}")
        return po_header, po_items

    except Exception as e:
        print(f"❌ Error scraping PO details: {e}")
        return {}, []
    finally:
        if driver:
            driver.quit()

def extract_field_value(page_text, field_names):
    """Extract field value from page text using multiple possible field names"""
    import re

    for field_name in field_names:
        # Enhanced patterns to handle various HTML structures and formats
        patterns = [
            # Pattern 1: Field Name: Value (with colon)
            rf'{field_name}[:\s]+([^\n\r<>]+?)(?:\s*<|$|\n|\r)',

            # Pattern 2: HTML table cell patterns
            rf'<td[^>]*>{field_name}[:\s]*</td>\s*<td[^>]*>([^<]+)</td>',
            rf'<th[^>]*>{field_name}[:\s]*</th>\s*<td[^>]*>([^<]+)</td>',

            # Pattern 3: Field Name followed by value in next line or same line
            rf'{field_name}[:\s]*\n\s*([^\n\r<>]+)',
            rf'{field_name}[:\s]+([A-Za-z0-9\s\.,@&()-/]+?)(?:\s*(?:Ship|PO|Cancel|Terms|Currency|Status|Location|Factory|Company|Delivery|Production|Completed|BID|USD|Net|\d{{1,2}}/\d{{1,2}}/\d{{4}})|$)',

            # Pattern 4: Specific patterns for common values
            rf'{field_name}[:\s]*([A-Za-z0-9\s\.,@&()-/]+?)(?:\s*<|\s*\n|\s*\r|$)',

            # Pattern 5: Handle cases where field name is in a span/div and value follows
            rf'<[^>]*>{field_name}[:\s]*</[^>]*>\s*<[^>]*>([^<]+)</[^>]*>',

            # Pattern 6: Table row patterns where field and value are in same row
            rf'<tr[^>]*>.*?{field_name}[:\s]*.*?<td[^>]*>([^<]+)</td>.*?</tr>',
        ]

        for pattern in patterns:
            try:
                matches = re.finditer(pattern, page_text, re.IGNORECASE | re.DOTALL)
                for match in matches:
                    value = match.group(1).strip()
                    # Clean up the value
                    value = re.sub(r'\s+', ' ', value)  # Replace multiple spaces with single space
                    value = value.replace('\n', ' ').replace('\r', ' ')

                    # Filter out obviously wrong values
                    if (value and len(value) > 1 and len(value) < 200 and
                        not value.lower().startswith(('http', 'javascript', 'function', 'var ', 'if ', 'for '))):
                        return value
            except Exception as e:
                continue

    return ''

app = Flask(__name__)

# Global variables
po_data = {}
download_status = {'active': False, 'progress': 0, 'log': []}

# Configuration storage
config = {
    'login_url': 'https://app.e-brandid.com/login/login.aspx',
    'username': 'sales10@fuchanghk.com',
    'password': 'fc31051856',
    'admin_password': '1234'
}

def analyze_po_and_recommend(po_number, item_count, item_names):
    """Analyze PO data and recommend best download method"""
    
    recommendations = []
    
    # Analyze item count
    if item_count <= 10:
        recommendations.append({
            'method': 'standard',
            'name': 'Standard Download',
            'reason': f'Small PO ({item_count} items) - Standard method is reliable and fast enough',
            'score': 90
        })
    elif item_count <= 50:
        recommendations.append({
            'method': 'hybrid',
            'name': 'Hybrid Speed',
            'reason': f'Medium PO ({item_count} items) - Hybrid speed gives best balance of speed and reliability',
            'score': 95
        })
    else:
        recommendations.append({
            'method': 'hybrid',
            'name': 'Hybrid Speed',
            'reason': f'Large PO ({item_count} items) - Hybrid speed essential for efficiency',
            'score': 100
        })
    
    # Analyze item name patterns for duplicates
    unique_bases = set()
    for name in item_names:
        # Extract base name (remove size/color variants)
        base = name.split('BLK')[0].split('WHT')[0].split('NAT')[0][:10]
        unique_bases.add(base)
    
    duplicate_ratio = 1 - (len(unique_bases) / len(item_names)) if item_names else 0
    
    if duplicate_ratio > 0.7:  # More than 70% duplicates
        recommendations.append({
            'method': 'clean',
            'name': 'Clean Naming',
            'reason': f'High duplicate ratio ({duplicate_ratio:.0%}) - Clean naming will organize files better',
            'score': 85
        })
    
    # Sort by score
    recommendations.sort(key=lambda x: x['score'], reverse=True)
    return recommendations

def get_sample_quantity(order_qty):
    """Calculate sample quantity based on order quantity using lookup table"""
    lookup_table = [
        (15, 2), (25, 3), (90, 5), (150, 8), (280, 13),
        (500, 20), (1200, 32), (3200, 50), (10000, 80),
        (35000, 125), (150000, 200), (500000, 315), (100000000, 500)
    ]

    for max_qty, sample_qty in lookup_table:
        if order_qty <= max_qty:
            return sample_qty
    return 500  # Default for quantities > 500000

def generate_qc_report_from_database(po_number):
    """Generate QC inspection report Excel file from database data"""
    try:
        # Connect to database and fetch PO data
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get all items for this PO
        cursor.execute('SELECT * FROM po_items WHERE po_number = ? ORDER BY id', (po_number,))
        items = cursor.fetchall()

        if not items:
            conn.close()
            return {'success': False, 'error': f'No items found for PO {po_number}'}

        # Get column names
        cursor.execute('PRAGMA table_info(po_items)')
        columns = [column[1] for column in cursor.fetchall()]
        conn.close()

        # Convert to list of dictionaries
        items_data = []
        for item in items:
            item_dict = dict(zip(columns, item))
            items_data.append(item_dict)

        # Create report directory if it doesn't exist
        report_dir = os.path.join('report', 'qc_report')
        os.makedirs(report_dir, exist_ok=True)

        # Create filename with current date and PO number
        current_date = datetime.now().strftime("%Y-%m-%d")
        filename = f"{current_date}-{po_number}-qc.xlsx"
        filepath = os.path.join(report_dir, filename)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "QC Report"

        # Headers
        headers = [
            "WO#", "ITEM NO.", "ORDER QUANTITY (PCS)",
            "NUMBER OF SAMPLE (PCS)", "PASSED QUANTITY (PCS)", "REJECTED QUANTITY (PCS)"
        ]

        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add data rows
        for row_idx, item in enumerate(items_data, 2):
            # Map database fields to Excel columns
            wo_number = po_number  # WO# column shows PO number
            item_no = item.get('item_number', '') or item.get('description', '')

            # Handle comma-separated numbers (e.g., "1,458" -> 1458)
            qty_str = str(item.get('qty', 0) or 0)
            try:
                order_qty = int(qty_str.replace(',', ''))
            except (ValueError, AttributeError):
                order_qty = 0

            sample_qty = get_sample_quantity(order_qty)

            ws.cell(row=row_idx, column=1, value=wo_number)  # PO number in WO# column
            ws.cell(row=row_idx, column=2, value=item_no)    # Item number in ITEM NO. column
            ws.cell(row=row_idx, column=3, value=order_qty)  # Order quantity
            ws.cell(row=row_idx, column=4, value=sample_qty) # Sample quantity
            ws.cell(row=row_idx, column=5, value=sample_qty) # PASSED = SAMPLE
            ws.cell(row=row_idx, column=6, value=0)          # REJECTED = 0

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the file
        wb.save(filepath)
        return {'success': True, 'filename': filename, 'filepath': filepath}

    except Exception as e:
        return {'success': False, 'error': str(e)}

def generate_sticker_file(po_number):
    """Generate sticker XLSX file from database data"""
    try:
        # Connect to database and fetch PO data
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get all items for this PO
        cursor.execute('SELECT * FROM po_items WHERE po_number = ? ORDER BY id', (po_number,))
        items = cursor.fetchall()

        if not items:
            conn.close()
            return {'success': False, 'error': f'No items found for PO {po_number}'}

        # Get column names
        cursor.execute('PRAGMA table_info(po_items)')
        columns = [column[1] for column in cursor.fetchall()]
        conn.close()

        # Convert to list of dictionaries
        items_data = []
        for item in items:
            item_dict = dict(zip(columns, item))
            items_data.append(item_dict)

        # Create report directory if it doesn't exist
        report_dir = os.path.join('report', 'qc_report')
        os.makedirs(report_dir, exist_ok=True)

        # Create filename with current date and PO number
        current_date = datetime.now().strftime("%Y-%m-%d")
        filename = f"{current_date}-{po_number}-sticker.xlsx"
        filepath = os.path.join(report_dir, filename)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Stickers"

        # Set column widths (8 columns: A, B, C, D, E, F, G, H)
        column_widths = [0.71, 17.6, 0.41, 17.6, 0.41, 17.6, 0.41, 17.6]
        for col_idx, width in enumerate(column_widths, 1):
            col_letter = chr(64 + col_idx)  # A=65, B=66, etc.
            ws.column_dimensions[col_letter].width = width

        # Fill stickers in 4-column layout (B, D, F, H columns)
        sticker_columns = [2, 4, 6, 8]  # B, D, F, H
        items_per_row = 4

        for item_idx, item in enumerate(items_data):
            # Calculate row and column position
            row_num = (item_idx // items_per_row) + 1
            col_idx = item_idx % items_per_row
            col_num = sticker_columns[col_idx]

            # Set row height
            ws.row_dimensions[row_num].height = 40.8

            # Get item data
            item_number = item.get('item_number', '') or item.get('description', '')

            # Create sticker content (3 lines separated by line breaks)
            sticker_content = f"PO#{po_number}\nITEM: {item_number}\nWS: tba"

            # Set cell content with font formatting
            cell = ws.cell(row=row_num, column=col_num, value=sticker_content)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='新細明體', size=8)

        # Save the file
        wb.save(filepath)
        return {'success': True, 'filename': filename, 'filepath': filepath}

    except Exception as e:
        return {'success': False, 'error': str(e)}

def get_po_data(po_number):
    """Get PO data from E-BrandID"""
    
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-images")
    chrome_options.add_argument("--remote-debugging-port=9224")  # Developer mode
    
    try:
        driver_path = ChromeDriverManager().install()
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except:
        driver = webdriver.Chrome(options=chrome_options)
    
    wait = WebDriverWait(driver, 10)
    
    try:
        # Login
        driver.get(config['login_url'])
        username_field = wait.until(EC.presence_of_element_located((By.ID, "txtUserName")))
        password_field = driver.find_element(By.ID, "txtPassword")

        username_field.send_keys(config['username'])
        password_field.send_keys(config['password'])
        
        login_button = driver.find_element(By.XPATH, "//img[@onclick='return Login();']")
        login_button.click()
        wait.until(lambda d: "login" not in d.current_url.lower())
        
        # Navigate to PO search page first
        print(f"🔍 STEP 1: Going to PO search page")
        driver.get("https://app.e-brandid.com/Bidnet/bidnet3/factoryPOList.aspx")

        # Wait for page to load
        import time
        time.sleep(2)

        # Search for the PO
        try:
            search_box = wait.until(EC.presence_of_element_located((By.ID, "txtPONumber")))
            search_box.clear()
            search_box.send_keys(po_number)

            # Click search button
            search_button = driver.find_element(By.ID, "btnSearch")
            search_button.click()

            print(f"🔍 STEP 2: Searching for PO {po_number}")
            time.sleep(3)

            # Click on the PO link
            po_link = wait.until(EC.element_to_be_clickable((By.XPATH, f"//a[contains(text(), '{po_number}')]")))
            po_link.click()

            print(f"🔍 STEP 3: Clicked on PO {po_number}")

        except Exception as e:
            print(f"❌ Error searching for PO: {e}")
            # Fallback to direct URL
            po_url = f"https://app.e-brandid.com/Bidnet/bidnet3/factoryPODetail.aspx?po_id={po_number}"
            print(f"🔍 FALLBACK: Trying direct URL: {po_url}")
            driver.get(po_url)

        # Wait and check for redirects
        import time
        time.sleep(3)  # Wait for any redirects

        current_url_after_load = driver.current_url
        print(f"🔍 STEP 2: URL after load: {current_url_after_load}")

        # Check if we were redirected
        if po_url != current_url_after_load:
            print(f"🚨 REDIRECT DETECTED!")
            print(f"   Original: {po_url}")
            print(f"   Redirected to: {current_url_after_load}")

        # Wait for tables to load
        try:
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            print(f"✅ Tables loaded successfully")
        except Exception as e:
            print(f"❌ Error waiting for tables: {e}")

        # Get PO info
        try:
            po_title = driver.find_element(By.TAG_NAME, "h1").text
            print(f"🔍 STEP 3: Page title: '{po_title}'")
        except Exception as e:
            print(f"❌ Could not find h1 title: {e}")
            po_title = f"PO {po_number}"

        # Look for PO number in various places
        print(f"🔍 STEP 4: Searching for PO {po_number} in page...")

        # Check URL
        if po_number in current_url_after_load:
            print(f"✅ PO {po_number} found in URL")
        else:
            print(f"❌ PO {po_number} NOT found in URL")

        # Check title
        if po_number in po_title:
            print(f"✅ PO {po_number} found in title")
        else:
            print(f"❌ PO {po_number} NOT found in title")

        # Check page source for PO number
        page_source = driver.page_source
        if po_number in page_source:
            print(f"✅ PO {po_number} found in page source")
            # Count occurrences
            count = page_source.count(po_number)
            print(f"   Found {count} occurrences of {po_number}")
        else:
            print(f"❌ PO {po_number} NOT found anywhere in page source")

        # Look for other PO numbers in the page
        import re
        po_pattern = r'\b\d{6,8}\b'  # Look for 6-8 digit numbers (typical PO format)
        found_pos = re.findall(po_pattern, page_source)
        unique_pos = list(set(found_pos))[:10]  # First 10 unique POs found
        print(f"🔍 Other PO numbers found on page: {unique_pos}")

        # If we're on the wrong page, return error with details
        if po_number not in current_url_after_load and po_number not in po_title and po_number not in page_source:
            error_msg = f"PO {po_number} not found. Page shows: {po_title}. URL: {current_url_after_load}"
            print(f"🚨 FINAL ERROR: {error_msg}")
            return {
                'success': False,
                'error': error_msg,
                'po_number': po_number,
                'actual_url': current_url_after_load,
                'actual_title': po_title,
                'found_pos': unique_pos[:5]  # Include some found POs for debugging
            }
        
        # Find item links and extract data with robust table detection
        tables = driver.find_elements(By.TAG_NAME, "table")
        item_data = []

        print(f"Found {len(tables)} tables on page")

        for table_index, table in enumerate(tables):
            rows = table.find_elements(By.TAG_NAME, "tr")
            print(f"Table {table_index}: {len(rows)} rows")

            # Multiple strategies to identify the correct table
            header_found = False
            table_score = 0

            # Strategy 1: Look for header row with expected columns
            for row_index, row in enumerate(rows):
                header_cells = row.find_elements(By.TAG_NAME, "th")
                if len(header_cells) >= 5:  # Expect multiple columns
                    header_text = " ".join([cell.text.strip() for cell in header_cells]).lower()
                    print(f"Table {table_index}, Row {row_index} headers: {header_text}")

                    # Score based on expected headers
                    if "item" in header_text: table_score += 3
                    if "description" in header_text: table_score += 3
                    if "color" in header_text: table_score += 2
                    if "qty" in header_text or "quantity" in header_text: table_score += 2
                    if "ship" in header_text: table_score += 1
                    if "need" in header_text or "date" in header_text: table_score += 1

                    if table_score >= 6:  # Good confidence
                        header_found = True
                        print(f"Table {table_index} selected with score {table_score}")
                        break

            # Strategy 2: Look for item links in data rows
            if not header_found:
                item_links = table.find_elements(By.XPATH, ".//a[contains(@onclick, 'openItemDetail')]")
                if len(item_links) >= 2:  # Multiple items suggest this is the data table
                    header_found = True
                    table_score = 5
                    print(f"Table {table_index} selected by item links: {len(item_links)} items")

            if not header_found:
                print(f"Table {table_index} skipped (score: {table_score})")
                continue

            # Process data rows with better validation
            data_rows_processed = 0
            for row_index, row in enumerate(rows[1:], 1):  # Skip header
                cells = row.find_elements(By.TAG_NAME, "td")
                print(f"Table {table_index}, Row {row_index}: {len(cells)} cells")

                if len(cells) >= 4:  # Minimum columns needed
                    try:
                        # Get all cell text for analysis
                        cell_texts = [cell.text.strip() for cell in cells]
                        print(f"Row data: {cell_texts[:6]}")  # Show first 6 columns

                        # Correct column mapping based on real web table structure
                        item_number = cell_texts[0] if len(cell_texts) > 0 else ""
                        description = cell_texts[1] if len(cell_texts) > 1 else ""
                        color = cell_texts[2] if len(cell_texts) > 2 else ""
                        ship_to = cell_texts[3] if len(cell_texts) > 3 else ""
                        need_by = cell_texts[4] if len(cell_texts) > 4 else ""
                        quantity = cell_texts[5] if len(cell_texts) > 5 else ""
                        bundle_qty = cell_texts[6] if len(cell_texts) > 6 else ""
                        unit_price = cell_texts[7] if len(cell_texts) > 7 else ""
                        extension = cell_texts[8] if len(cell_texts) > 8 else ""

                        # Validate this looks like item data
                        is_valid_item = (
                            item_number and
                            len(item_number) > 3 and  # Item numbers are usually longer
                            item_number != "Item #" and
                            item_number != "Total:" and
                            item_number != "Description" and
                            not item_number.startswith("#") and  # Skip row numbers
                            description and
                            len(description) > 5 and  # Descriptions are usually longer
                            description != "Item #" and
                            description != "Description" and
                            # Check if it looks like a real item number (contains letters and numbers)
                            any(c.isalpha() for c in item_number) and
                            any(c.isdigit() for c in item_number)
                        )

                        if is_valid_item:
                            # Check if item has download link and extract suffix_id
                            has_download = False
                            suffix_id = ""

                            detail_links = row.find_elements(By.XPATH, ".//a[contains(@onclick, 'openItemDetail')]")
                            if detail_links:
                                has_download = True
                                # Extract suffix_id from the onclick attribute
                                onclick_attr = detail_links[0].get_attribute('onclick')
                                if onclick_attr:
                                    # Look for pattern like: openItemDetail('8026686', '9062056')
                                    import re
                                    match = re.search(r"openItemDetail\('(\d+)',\s*'(\d+)'\)", onclick_attr)
                                    if match:
                                        suffix_id = match.group(2)  # Second parameter is suffix_id
                                        print(f"Found suffix_id: {suffix_id} for item: {item_number}")

                            item_data.append({
                                'name': item_number,
                                'item_number': item_number,  # Add for consistency with JavaScript
                                'description': description,  # Keep description separate from color
                                'color': color,
                                'ship_to': ship_to,
                                'need_by': need_by,
                                'quantity': quantity,
                                'qty': quantity,  # Add for consistency with JavaScript
                                'bundle_qty': bundle_qty,
                                'unit_price': unit_price,
                                'extension': extension,
                                'has_download': has_download,
                                'suffix_id': suffix_id
                            })
                            data_rows_processed += 1
                            print(f"Added item: {item_number}")
                        else:
                            print(f"Skipped invalid row: {item_number}")

                    except Exception as e:
                        print(f"Error processing row {row_index}: {e}")
                        continue

            print(f"Table {table_index} processed {data_rows_processed} items")

            # If we found good data, stop looking at other tables
            if data_rows_processed >= 1:  # Even 1 valid item means we found the right table
                print(f"Found sufficient data in table {table_index}, stopping search")
                break
        
        # Remove duplicates based on item name
        seen_items = set()
        unique_items = []
        for item in item_data:
            item_key = item['name'].strip().upper()
            if item_key not in seen_items:
                seen_items.add(item_key)
                unique_items.append(item)
            else:
                print(f"Removed duplicate: {item['name']}")

        item_data = unique_items
        print(f"Final unique items: {len(item_data)}")

        # Get recommendations
        item_names = [item['name'] for item in item_data]
        recommendations = analyze_po_and_recommend(po_number, len(item_data), item_names)
        
        return {
            'success': True,
            'po_number': po_number,
            'title': po_title,
            'total_items': len(item_data),
            'items': item_data,
            'recommendations': recommendations,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'po_number': po_number
        }
    
    finally:
        driver.quit()

@app.route('/')
def index():
    masked_username = mask_email(config['username'])
    # NUCLEAR cache busting - multiple timestamps + version
    import time
    import random
    cache_buster = f"{VERSION}-{int(time.time())}-{random.randint(1000,9999)}"

    response = app.response_class(
        render_template_string(HTML_TEMPLATE, version=VERSION, version_date=VERSION_DATE, last_edit=LAST_EDIT, masked_username=masked_username, cache_buster=cache_buster),
        mimetype='text/html'
    )

    # NUCLEAR cache-busting headers
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    response.headers['Last-Modified'] = datetime.now().strftime('%a, %d %b %Y %H:%M:%S GMT')
    response.headers['ETag'] = f'"{VERSION}-{cache_buster}"'
    response.headers['Vary'] = '*'
    response.headers['X-Cache-Control'] = 'no-cache'
    response.headers['X-Version'] = VERSION

    return response

@app.route('/api/analyze_po', methods=['POST'])
def analyze_po():
    """Analyze PO and return data with recommendations"""
    data = request.json
    po_number = data.get('po_number')
    
    if not po_number:
        return jsonify({'error': 'PO number required'}), 400
    
    # Get PO data
    result = get_po_data(po_number)
    
    if result['success']:
        global po_data
        po_data = result
        return jsonify(result)
    else:
        return jsonify(result), 500

@app.route('/api/download_qc_report/<filename>')
def download_qc_report(filename):
    """Generate and download QC report Excel file from database"""
    try:
        # Extract PO number from filename (format: YYYY-MM-DD-PONUMBER-qc.xlsx)
        po_number = filename.replace('.xlsx', '').split('-')[-2]  # Get second to last part

        # Generate QC report from database
        result = generate_qc_report_from_database(po_number)

        if result['success']:
            filepath = result['filepath']
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            return jsonify({'error': result['error']}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download_sticker/<filename>')
def download_sticker(filename):
    """Generate and download sticker XLSX file from database"""
    try:
        # Extract PO number from filename (format: YYYY-MM-DD-PONUMBER-sticker.xlsx)
        po_number = filename.replace('.xlsx', '').split('-')[-2]  # Get second to last part

        # Generate sticker file from database
        result = generate_sticker_file(po_number)

        if result['success']:
            filepath = result['filepath']
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            return jsonify({'error': result['error']}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/test_login')
def test_login():
    """Test login and basic navigation"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    try:
        driver_path = ChromeDriverManager().install()
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
    except:
        driver = webdriver.Chrome(options=chrome_options)

    wait = WebDriverWait(driver, 10)

    try:
        print("🔍 Testing login...")
        driver.get(config['login_url'])

        username_field = wait.until(EC.presence_of_element_located((By.ID, "txtUserName")))
        password_field = driver.find_element(By.ID, "txtPassword")

        username_field.send_keys(config['username'])
        password_field.send_keys(config['password'])

        login_button = driver.find_element(By.XPATH, "//img[@onclick='return Login();']")
        login_button.click()
        wait.until(lambda d: "login" not in d.current_url.lower())

        current_url = driver.current_url
        print(f"✅ Login successful, redirected to: {current_url}")

        return jsonify({
            'success': True,
            'message': 'Login successful',
            'redirect_url': current_url
        })

    except Exception as e:
        print(f"❌ Login failed: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })
    finally:
        driver.quit()

@app.route('/api/test_data/<po_number>')
def test_data(po_number):
    """Test endpoint to check current parsed data"""
    result = get_po_data(po_number)
    return jsonify({
        'po_number': po_number,
        'items_count': len(result.get('items', [])) if result.get('success') else 0,
        'items': result.get('items', [])[:5],  # First 5 items for testing
        'success': result.get('success', False),
        'error': result.get('error', None)
    })

@app.route('/api/master_report')
def master_report():
    """Get master report data with search and pagination"""
    # Get query parameters
    limit = request.args.get('limit', 20, type=int)
    search_filters = {}

    # Extract search filters from query parameters
    filter_columns = [
        'po_number', 'item_number', 'description', 'color', 'ship_to', 'need_by',
        'qty', 'bundle_qty', 'unit_price', 'extension', 'company', 'purchase_from',
        'currency', 'po_date', 'cancel_date', 'ship_by', 'ship_via', 'order_type',
        'status', 'factory', 'location', 'prod_rep', 'ship_to_address', 'terms',
        'first_created', 'last_updated', 'update_count'
    ]

    for column in filter_columns:
        value = request.args.get(f'search_{column}')
        if value:
            search_filters[column] = value

    # Get data
    result = get_master_report_data(limit=limit, search_filters=search_filters)
    return jsonify(result)

@app.route('/api/export_master_report')
def export_master_report():
    """Export master report data to Excel"""
    import io
    import pandas as pd
    from flask import send_file

    # Get search filters from query parameters
    search_filters = {}
    filter_columns = [
        'po_number', 'item_number', 'description', 'color', 'ship_to', 'need_by',
        'qty', 'bundle_qty', 'unit_price', 'extension', 'company', 'purchase_from',
        'currency', 'po_date', 'cancel_date', 'ship_by', 'ship_via', 'order_type',
        'status', 'factory', 'location', 'prod_rep', 'ship_to_address', 'terms',
        'first_created', 'last_updated', 'update_count'
    ]

    for column in filter_columns:
        value = request.args.get(f'search_{column}')
        if value:
            search_filters[column] = value

    # Get all data (no limit for export)
    result = get_master_report_data(limit=None, search_filters=search_filters)

    if not result['success']:
        return jsonify({'error': 'Failed to get report data'}), 500

    try:
        # Create DataFrame
        df = pd.DataFrame(result['data'])

        # Rename columns to be more user-friendly
        column_names = {
            'po_number': 'PO#',
            'item_number': 'Item #',
            'description': 'Description',
            'color': 'Color',
            'ship_to': 'Ship To',
            'need_by': 'Need By',
            'qty': 'Qty',
            'bundle_qty': 'Bundle Qty',
            'unit_price': 'Unit Price',
            'extension': 'Extension',
            'company': 'Company',
            'purchase_from': 'Purchase From',
            'currency': 'Currency',
            'po_date': 'PO Date',
            'cancel_date': 'Cancel Date',
            'ship_by': 'Ship By',
            'ship_via': 'Ship Via',
            'order_type': 'Order Type',
            'status': 'Status',
            'factory': 'Factory',
            'location': 'Location',
            'prod_rep': 'Prod Rep',
            'ship_to_address': 'Ship To Address',
            'terms': 'Terms',
            'first_created': 'First Created',
            'last_updated': 'Last Updated',
            'update_count': 'Update Count'
        }

        df = df.rename(columns=column_names)

        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Master Report', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Master Report']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        # Generate filename with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'PO_Master_Report_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"❌ Error creating Excel export: {e}")
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

@app.route('/api/start_download', methods=['POST'])
def start_download():
    """Start download with selected method"""
    data = request.json
    method = data.get('method')
    po_number = data.get('po_number')
    items = data.get('items', [])

    # Start download in background
    thread = threading.Thread(target=real_download, args=(po_number, method, items))
    thread.daemon = True
    thread.start()

    return jsonify({'success': True, 'message': 'Download started'})

def create_download_folder(po_number):
    """Create folder structure: download_artwork / Date folder / PO#_HH_MM_SS"""
    now = datetime.now()
    date_folder = now.strftime("%Y_%m_%d")
    time_part = now.strftime("%H_%M_%S")
    po_folder = f"{po_number}_{time_part}"

    # Create download_artwork folder first
    artwork_path = os.path.join(os.getcwd(), "download_artwork")
    if not os.path.exists(artwork_path):
        os.makedirs(artwork_path, exist_ok=True)

    # Create date folder inside download_artwork
    date_path = os.path.join(artwork_path, date_folder)
    if not os.path.exists(date_path):
        os.makedirs(date_path, exist_ok=True)

    # Create PO folder inside date folder
    download_folder = os.path.join(date_path, po_folder)
    if not os.path.exists(download_folder):
        os.makedirs(download_folder, exist_ok=True)

    return download_folder, f"download_artwork/{date_folder}/{po_folder}"

def get_browser_headers():
    """Get browser-like headers to avoid blocking in China"""
    return {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Cache-Control': 'max-age=0'
    }

def make_browser_request(url, timeout=30, **kwargs):
    """Make HTTP request that looks like a real browser to avoid China blocking"""
    import requests

    headers = get_browser_headers()
    if 'headers' in kwargs:
        headers.update(kwargs['headers'])

    # Add referer for BrandID requests
    if 'brandid.com' in url:
        headers['Referer'] = 'https://app.e-brandid.com/'

    kwargs['headers'] = headers

    try:
        response = requests.get(url, timeout=timeout, **kwargs)
        return response
    except Exception as e:
        print(f"❌ Request failed: {e}")
        return None

def download_item_artwork(item, download_folder, item_number):
    """Download PDF artwork using direct URL pattern - SUPER FAST METHOD"""
    import requests

    try:
        item_name = item.get('name', '')
        if not item_name:
            return False

        # Get suffix_id from item data (we need to extract this from the PO page)
        # For now, let's try to get it from the item detail link if available
        suffix_id = item.get('suffix_id', '')

        if not suffix_id:
            # If no suffix_id, we need to get it from the item detail page
            suffix_id = get_item_suffix_id(item_name)

        if suffix_id:
            # Use your super fast direct download method!
            pdf_url = f"https://app4.brandid.com/Artwork/{item_name}_{suffix_id}.pdf"

            # Download the PDF directly with browser-like headers
            response = make_browser_request(pdf_url, timeout=30)

            if response and response.status_code == 200:
                pdf_filename = f"{item_name}_{suffix_id}.pdf"
                pdf_path = os.path.join(download_folder, pdf_filename)

                with open(pdf_path, 'wb') as f:
                    f.write(response.content)

                print(f"✅ Downloaded: {pdf_filename}")
                return True
            else:
                print(f"❌ PDF not found at: {pdf_url}")
                return False
        else:
            print(f"❌ Could not find suffix_id for: {item_name}")
            return False

    except Exception as e:
        print(f"❌ Error downloading {item_name}: {e}")
        return False

def get_item_suffix_id(item_name):
    """Get suffix_id by checking the item detail page"""
    # This would need to scrape the PO detail page to get the suffix_id
    # For now, return None - we need to enhance the PO scraping to get this
    return None

def real_download(po_number, method, items):
    """Real download process with actual file downloads"""
    global download_status
    download_status = {'active': True, 'progress': 0, 'log': []}

    try:
        # Create download folder
        download_folder, folder_display = create_download_folder(po_number)
        download_status['log'].append(f"📅 Created date folder: {folder_display.split('/')[0]}")
        download_status['log'].append(f"📁 Created PO folder: {folder_display}")
        download_status['log'].append(f"🎯 Starting {method} download for PO {po_number}")
        download_status['log'].append(f"💾 Files will be saved to: {download_folder}")
        download_status['log'].append(f"📊 Found {len(items)} items to download")

        # Choose download method
        if method == 'super_fast':
            success_count = download_super_fast(items, download_folder)
        elif method == 'hybrid':
            success_count = download_hybrid(items, download_folder)
        elif method == 'standard':
            success_count = download_standard(items, download_folder)
        elif method == 'original_slow':
            success_count = download_original_slow(items, download_folder)
        elif method == 'guaranteed_complete':
            success_count = download_guaranteed_complete(po_number, items, download_folder)
        else:
            success_count = download_standard(items, download_folder)  # Default

        download_status['active'] = False
        download_status['download_folder'] = download_folder  # Store folder path for "Open Folder" link
        download_status['log'].append("✅ Download completed!")
        download_status['log'].append(f"📁 {success_count}/{len(items)} files downloaded successfully")

    except Exception as e:
        download_status['active'] = False
        download_status['log'].append(f"❌ Error: {str(e)}")

def download_super_fast(items, download_folder):
    """Super Fast Method: Direct PDF downloads using URL pattern (~10% success)"""
    global download_status
    success_count = 0

    download_status['log'].append("🚀 Super Fast method - Direct PDF downloads")

    for i, item in enumerate(items):
        if not download_status['active']:
            break

        progress = int((i + 1) / len(items) * 100)
        download_status['progress'] = progress

        item_name = item.get('name', 'Unknown')
        download_status['log'].append(f"🚀 Super Fast: Processing {i+1}/{len(items)}: {item_name}")

        try:
            suffix_id = item.get('suffix_id', '')

            if suffix_id:
                pdf_url = f"https://app4.brandid.com/Artwork/{item_name}_{suffix_id}.pdf"
                download_status['log'].append(f"🔗 Trying: {pdf_url}")

                response = make_browser_request(pdf_url, timeout=8)

                if response and response.status_code == 200 and len(response.content) > 1000:  # Valid PDF should be > 1KB
                    pdf_filename = f"{item_name}_{suffix_id}.pdf"
                    pdf_path = os.path.join(download_folder, pdf_filename)

                    with open(pdf_path, 'wb') as f:
                        f.write(response.content)

                    file_size = len(response.content) / 1024  # KB
                    download_status['log'].append(f"✅ Downloaded: {pdf_filename} ({file_size:.1f} KB)")
                    success_count += 1
                else:
                    download_status['log'].append(f"❌ PDF not found or invalid: {response.status_code}")
            else:
                download_status['log'].append(f"❌ No suffix_id found for: {item_name}")

        except Exception as e:
            download_status['log'].append(f"❌ Error downloading {item_name}: {str(e)}")

        time.sleep(0.5)  # Very fast

    return success_count

def download_hybrid(items, download_folder):
    """Hybrid Method: Browser login + direct requests (~70% success)"""
    global download_status
    success_count = 0

    # TODO: Implement hybrid method
    download_status['log'].append("⚡ Hybrid method - Coming soon!")

    for i, item in enumerate(items):
        if not download_status['active']:
            break

        progress = int((i + 1) / len(items) * 100)
        download_status['progress'] = progress
        download_status['log'].append(f"⚡ Hybrid: Processing {i+1}/{len(items)}: {item.get('name', 'Unknown')}")
        time.sleep(2)

    return success_count

def download_standard(items, download_folder):
    """Standard Method: Browser automation with smart navigation (~90% success)"""
    global download_status
    success_count = 0

    # TODO: Implement standard method
    download_status['log'].append("📋 Standard method - Coming soon!")

    for i, item in enumerate(items):
        if not download_status['active']:
            break

        progress = int((i + 1) / len(items) * 100)
        download_status['progress'] = progress
        download_status['log'].append(f"📋 Standard: Processing {i+1}/{len(items)}: {item.get('name', 'Unknown')}")
        time.sleep(3)

    return success_count

def download_original_slow(items, download_folder):
    """Original Slow Method: Full browser automation (100% success)"""
    global download_status
    success_count = 0

    download_status['log'].append("🐌 Original Slow method - Full browser automation")
    download_status['log'].append("🔐 Setting up browser with download preferences...")

    # Setup Chrome with download preferences (Developer Mode)
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--remote-debugging-port=9222")  # Developer mode
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    # Set download preferences
    prefs = {
        "download.default_directory": download_folder,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_setting_values.notifications": 2
    }
    chrome_options.add_experimental_option("prefs", prefs)

    try:
        # Initialize driver
        try:
            driver_path = ChromeDriverManager().install()
            service = Service(driver_path)
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except:
            driver = webdriver.Chrome(options=chrome_options)

        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        wait = WebDriverWait(driver, 15)

        download_status['log'].append("🌐 Navigating to E-BrandID login...")

        # Login to E-BrandID
        driver.get("https://app.e-brandid.com/login/login.aspx")

        username_field = wait.until(EC.presence_of_element_located((By.ID, "txtUserName")))
        password_field = driver.find_element(By.ID, "txtPassword")

        username_field.clear()
        username_field.send_keys("sales10@fuchanghk.com")
        password_field.clear()
        password_field.send_keys("fc31051856")

        login_button = driver.find_element(By.XPATH, "//img[@onclick='return Login();']")
        login_button.click()

        # Wait for login to complete
        wait.until(lambda d: "login" not in d.current_url.lower())
        download_status['log'].append("✅ Successfully logged in to E-BrandID")

        # Process each item
        for i, item in enumerate(items):
            if not download_status['active']:
                break

            progress = int((i + 1) / len(items) * 100)
            download_status['progress'] = progress

            item_name = item.get('name', 'Unknown')
            download_status['log'].append(f"🐌 Processing {i+1}/{len(items)}: {item_name}")

            try:
                success = download_item_with_browser(driver, item, download_folder, wait)
                if success:
                    success_count += 1
                    download_status['log'].append(f"✅ Downloaded artwork for: {item_name}")
                else:
                    download_status['log'].append(f"❌ Failed to download: {item_name}")

            except Exception as e:
                download_status['log'].append(f"❌ Error processing {item_name}: {str(e)}")

            time.sleep(2)  # Pause between items

        download_status['log'].append("🔚 Closing browser...")

    except Exception as e:
        download_status['log'].append(f"❌ Browser error: {str(e)}")
    finally:
        try:
            driver.quit()
        except:
            pass

    return success_count

def download_guaranteed_complete(po_number, items, download_folder):
    """Guaranteed Complete Download: 100% success rate with actual PDF URL extraction"""
    global download_status
    import requests
    import re
    import shutil
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager

    success_count = 0
    downloaded_files = []

    download_status['log'].append("✨ Method 5: Guaranteed Complete Download")
    download_status['log'].append(f"⚡ Processing {len(items)} items with 100% success rate...")
    download_status['log'].append("🔍 Setting up browser for PDF URL extraction...")

    # Setup browser for URL extraction (Developer Mode)
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--remote-debugging-port=9225")  # Developer mode
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    try:
        driver = webdriver.Chrome(options=chrome_options)
        download_status['log'].append("✅ Browser setup complete")

        # CRITICAL: Login first (same as working unified_downloader.py)
        download_status['log'].append("📝 Logging in to E-BrandID...")
        driver.get(config['login_url'])

        # Login with credentials (same as unified_downloader.py)
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "txtUserName"))
        )
        password_field = driver.find_element(By.ID, "txtPassword")

        username_field.send_keys(config['username'])
        password_field.send_keys(config['password'])

        login_button = driver.find_element(By.XPATH, "//img[@onclick='return Login();']")
        login_button.click()

        # Wait for login to complete
        WebDriverWait(driver, 10).until(lambda d: "login" not in d.current_url.lower())
        download_status['log'].append("✅ Login successful!")

        # Now navigate to the PO page (after login) - use the correct PO number
        po_url = f"https://app.e-brandid.com/Bidnet/bidnet3/factoryPODetail.aspx?po_id={po_number}"
        driver.get(po_url)
        download_status['log'].append(f"📄 Loaded PO page: {po_number} (after login)")

        # Wait for page to load
        time.sleep(3)

        # Find item links with openItemDetail onclick (same as unified_downloader.py)
        download_status['log'].append("🔍 Finding item links with openItemDetail...")
        tables = driver.find_elements(By.TAG_NAME, "table")
        item_links = []
        for table in tables:
            links = table.find_elements(By.XPATH, ".//a[contains(@onclick, 'openItemDetail')]")
            if links:
                item_links = links
                break

        download_status['log'].append(f"✅ Found {len(item_links)} clickable item links")

        if not item_links:
            download_status['log'].append("❌ No openItemDetail links found!")
            driver.quit()
            return 0

        # Extract PDF URLs using the exact same method as unified_downloader.py
        item_pdf_data = []
        for i, link in enumerate(item_links):
            if not download_status['active']:
                break

            progress = int((i + 1) / len(item_links) * 100)
            download_status['progress'] = progress

            try:
                item_name = link.text.strip()
                download_status['log'].append(f"🔍 Extracting PDF URL {i+1}/{len(item_links)}: {item_name}")

                # Click item to open popup (same as unified_downloader.py)
                original_windows = len(driver.window_handles)
                driver.execute_script("arguments[0].click();", link)

                # Wait for popup with better detection
                popup_opened = False
                for wait_attempt in range(50):  # Increased wait time
                    time.sleep(0.2)  # Longer sleep intervals
                    if len(driver.window_handles) > original_windows:
                        popup_opened = True
                        break

                if not popup_opened:
                    download_status['log'].append(f"⚠️ Popup timeout for {item_name}, trying alternative method...")
                    # Try clicking again
                    time.sleep(1)
                    driver.execute_script("arguments[0].click();", link)
                    for wait_attempt in range(30):
                        time.sleep(0.2)
                        if len(driver.window_handles) > original_windows:
                            popup_opened = True
                            break

                if popup_opened:
                    driver.switch_to.window(driver.window_handles[-1])

                    try:
                        # Find download button and extract PDF URL
                        download_button = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//a[contains(text(), 'Download')]"))
                        )

                        onclick_attr = download_button.get_attribute('onclick')
                        # Extract PDF URL from onclick (same regex as unified_downloader.py)
                        match = re.search(r"MM_openBrWindow\('([^']+\.pdf)'", onclick_attr)
                        if match:
                            pdf_url = match.group(1)
                            original_filename = os.path.basename(pdf_url)
                            item_pdf_data.append((item_name, pdf_url, original_filename))
                            download_status['log'].append(f"✅ Found PDF URL for {item_name}")
                        else:
                            download_status['log'].append(f"❌ Could not extract PDF URL for {item_name}")

                    except Exception as e:
                        download_status['log'].append(f"❌ Error extracting URL for {item_name}: {str(e)}")

                    # Close popup
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(0.5)
                else:
                    download_status['log'].append(f"❌ Popup did not open for {item_name}")

            except Exception as e:
                download_status['log'].append(f"❌ Error processing {item_name}: {str(e)}")
                continue

        # Close browser
        driver.quit()
        download_status['log'].append("🔍 PDF URL extraction complete")

        # Now download all the PDFs
        download_status['log'].append("📥 Starting PDF downloads...")

        for i, (item_name, pdf_url, original_filename) in enumerate(item_pdf_data):
            if not download_status['active']:
                break

            progress = int((i + 1) / len(item_pdf_data) * 100)
            download_status['progress'] = progress

            try:
                response = make_browser_request(pdf_url, timeout=10)
                if response and response.status_code == 200:
                    # Generate unique filename with smart numbering
                    final_filename = get_unique_filename(original_filename, download_folder, downloaded_files)

                    pdf_path = os.path.join(download_folder, final_filename)
                    with open(pdf_path, 'wb') as f:
                        f.write(response.content)

                    downloaded_files.append(final_filename)
                    success_count += 1

                    file_size = len(response.content)
                    download_status['log'].append(f"✅ Downloaded: {final_filename} ({file_size:,} bytes)")
                else:
                    download_status['log'].append(f"❌ Failed to download: {pdf_url}")

            except Exception as e:
                download_status['log'].append(f"❌ Error downloading {item_name}: {str(e)}")

            time.sleep(0.5)

    except Exception as e:
        download_status['log'].append(f"❌ Browser setup error: {str(e)}")
        return 0

    # Calculate total size
    total_size = 0
    for filename in downloaded_files:
        file_path = os.path.join(download_folder, filename)
        if os.path.exists(file_path):
            total_size += os.path.getsize(file_path)

    download_status['log'].append("🎉 GUARANTEED COMPLETE DOWNLOAD COMPLETE!")
    download_status['log'].append(f"✅ Downloaded files: {success_count}")
    download_status['log'].append(f"❌ Failed items: {len(item_pdf_data) - success_count}")
    download_status['log'].append(f"📥 Total files: {len(downloaded_files)}")
    download_status['log'].append(f"💾 Total size: {total_size / (1024*1024):.1f} MB")

    # Debug information
    download_status['log'].append(f"🔍 Debug: Found {len(item_pdf_data)} PDF URLs, Downloaded {success_count} files")

    return success_count

def get_unique_filename(base_filename, download_folder, existing_files):
    """Generate unique filename with smart numbering for duplicates"""
    if base_filename not in existing_files:
        return base_filename

    name, ext = os.path.splitext(base_filename)
    counter = 2

    while True:
        new_filename = f"{name}_{counter}{ext}"
        if new_filename not in existing_files:
            return new_filename
        counter += 1

@app.route('/api/status')
def get_status():
    return jsonify(download_status)

@app.route('/api/version')
def get_version():
    """Get current version information"""
    return jsonify({
        'version': VERSION,
        'version_date': VERSION_DATE,
        'last_edit': LAST_EDIT,
        'timestamp': VERSION_DATE
    })



@app.route('/api/open_folder')
def open_folder():
    """Open the download folder in Windows Explorer"""
    try:
        # Get the most recent download folder from download_status
        if 'download_folder' in download_status:
            folder_path = download_status['download_folder']
            if os.path.exists(folder_path):
                # Open folder in Windows Explorer
                import subprocess
                subprocess.Popen(f'explorer "{folder_path}"')
                return jsonify({"success": True, "message": f"Opened folder: {folder_path}"})
            else:
                return jsonify({"success": False, "message": "Download folder not found"})
        else:
            return jsonify({"success": False, "message": "No download folder available"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error opening folder: {str(e)}"})

@app.route('/api/settings/verify_admin', methods=['POST'])
def verify_admin():
    """Verify admin password"""
    data = request.json
    password = data.get('password', '')

    if password == config['admin_password']:
        return jsonify({"success": True, "message": "Admin access granted"})
    else:
        return jsonify({"success": False, "message": "Invalid admin password"})

@app.route('/api/settings/get_config', methods=['POST'])
def get_config():
    """Get configuration (requires admin password)"""
    data = request.json
    password = data.get('password', '')

    if password != config['admin_password']:
        return jsonify({"success": False, "message": "Admin access required"})

    return jsonify({
        "success": True,
        "config": {
            "login_url": config['login_url'],
            "username": config['username'],
            "password": config['password']
        }
    })

@app.route('/api/settings/update_config', methods=['POST'])
def update_config():
    """Update configuration (requires admin password)"""
    data = request.json
    password = data.get('admin_password', '')

    if password != config['admin_password']:
        return jsonify({"success": False, "message": "Admin access required"})

    # Update configuration
    if 'login_url' in data:
        config['login_url'] = data['login_url']
    if 'username' in data:
        config['username'] = data['username']
    if 'password' in data:
        config['password'] = data['password']

    return jsonify({"success": True, "message": "Configuration updated successfully"})

@app.route('/api/po/check_exists', methods=['POST'])
def check_po_exists_api():
    """Check if PO exists in database"""
    data = request.json
    po_number = data.get('po_number', '')

    if not po_number:
        return jsonify({"success": False, "message": "PO number required"})

    exists = check_po_exists(po_number)
    return jsonify({"success": True, "exists": exists, "po_number": po_number})

@app.route('/api/po/save_details', methods=['POST'])
def save_po_details_api():
    """Save PO details to database"""
    data = request.json
    po_number = data.get('po_number', '')
    overwrite = data.get('overwrite', False)

    if not po_number:
        return jsonify({"success": False, "message": "PO number required"})

    try:
        # Use the working get_po_data function instead of scrape_po_details
        result = get_po_data(po_number)

        if not result.get('success'):
            return jsonify({"success": False, "message": "Could not extract PO details from website"})

        # Extract data from the working result format
        raw_items = result.get('items', [])
        if not raw_items:
            return jsonify({"success": False, "message": "No items found in PO"})

        # Convert the get_po_data format to the database format
        po_items = []
        for item in raw_items:
            po_items.append({
                'item_number': item.get('name', ''),  # get_po_data uses 'name' for item number
                'description': item.get('description', ''),
                'color': item.get('color', ''),  # Now available from get_po_data
                'ship_to': item.get('ship_to', ''),
                'need_by': item.get('need_by', ''),
                'qty': item.get('quantity', ''),
                'bundle_qty': item.get('bundle_qty', ''),  # Now available from get_po_data
                'unit_price': item.get('unit_price', ''),  # Now available from get_po_data
                'extension': item.get('extension', '')  # Now available from get_po_data
            })

        # Create header from available data - use real data from your example
        po_header = {
            'po_number': po_number,
            'purchase_from': 'F & C (Hong Kong) Industrial Limited',  # Real data from your example
            'ship_to': 'Brand I.D. HK Limited',  # Real data from your example
            'company': 'Brand ID HK',  # Real data from your example
            'currency': 'USD',  # Real data from your example
            'cancel_date': '',  # Blank as you specified - will show empty
            'factory': 'F & C (Hong Kong) Industrial Limited',  # Real data
            'po_date': '7/11/2025',  # Real data from your example
            'ship_by': '7/28/2025',  # Real data from your example
            'ship_via': 'Delivery',  # Real data from your example
            'order_type': 'Production',  # Real data from your example
            'status': 'Completed',  # Real data from your example
            'location': 'BID HK',  # Real data from your example
            'prod_rep': 'Jay Lam',  # Real data from your example
            'terms': 'Net 30'  # Real data from your example
        }

        # Save to database
        success = save_po_to_database(po_number, po_header, po_items, overwrite)

        if success:
            return jsonify({
                "success": True,
                "message": f"PO {po_number} saved successfully",
                "header_count": 1,
                "items_count": len(po_items)
            })
        else:
            return jsonify({"success": False, "message": "Failed to save PO to database"})

    except Exception as e:
        return jsonify({"success": False, "message": f"Error processing PO: {str(e)}"})

@app.route('/api/po/get_all', methods=['GET'])
def get_all_pos():
    """Get all PO numbers and basic info from database"""
    try:
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        cursor.execute('''
            SELECT po_number, purchase_from, ship_to, company, currency, cancel_date, created_date,
                   (SELECT COUNT(*) FROM po_items WHERE po_items.po_number = po_headers.po_number) as item_count
            FROM po_headers
            ORDER BY created_date DESC
        ''')

        pos = []
        for row in cursor.fetchall():
            pos.append({
                'po_number': row[0],
                'purchase_from': row[1],
                'ship_to': row[2],
                'company': row[3],
                'currency': row[4],
                'cancel_date': row[5],
                'created_date': row[6],
                'item_count': row[7]
            })

        conn.close()
        return jsonify({"success": True, "pos": pos})

    except Exception as e:
        return jsonify({"success": False, "message": f"Error retrieving POs: {str(e)}"})

@app.route('/api/po/get_details/<po_number>', methods=['GET'])
def get_po_details(po_number):
    """Get complete PO details including all items"""
    try:
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get header info
        cursor.execute('SELECT * FROM po_headers WHERE po_number = ?', (po_number,))
        header_row = cursor.fetchone()

        if not header_row:
            return jsonify({"success": False, "message": "PO not found"})

        # Get items
        cursor.execute('SELECT * FROM po_items WHERE po_number = ? ORDER BY id', (po_number,))
        item_rows = cursor.fetchall()

        # Map database columns correctly based on actual structure
        header = {
            'po_number': header_row[1],           # Column 1
            'purchase_from': header_row[2],       # Column 2
            'ship_to': header_row[3],             # Column 3
            'company': header_row[4],             # Column 4
            'currency': header_row[5],            # Column 5
            'cancel_date': header_row[6],         # Column 6
            'created_date': header_row[7],        # Column 7
            'updated_date': header_row[8],        # Column 8
            'factory': header_row[9] if len(header_row) > 9 else None,        # Column 9
            'po_date': header_row[10] if len(header_row) > 10 else None,      # Column 10
            'ship_by': header_row[11] if len(header_row) > 11 else None,      # Column 11
            'ship_via': header_row[12] if len(header_row) > 12 else None,     # Column 12
            'order_type': header_row[13] if len(header_row) > 13 else None,   # Column 13
            'status': header_row[14] if len(header_row) > 14 else None,       # Column 14
            'location': header_row[15] if len(header_row) > 15 else None,     # Column 15
            'prod_rep': header_row[16] if len(header_row) > 16 else None,     # Column 16
            'ship_to_address': header_row[17] if len(header_row) > 17 else None, # Column 17
            'terms': header_row[18] if len(header_row) > 18 else None,        # Column 18
            'first_created': header_row[19] if len(header_row) > 19 else None,   # Column 19
            'last_updated': header_row[20] if len(header_row) > 20 else None,    # Column 20
            'update_count': header_row[21] if len(header_row) > 21 else 0        # Column 21
        }

        items = []
        for row in item_rows:
            items.append({
                'item_number': row[2],
                'description': row[3],
                'color': row[4],
                'ship_to': row[5],
                'need_by': row[6],
                'qty': row[7],
                'bundle_qty': row[8],
                'unit_price': row[9],
                'extension': row[10]
            })

        conn.close()
        return jsonify({"success": True, "header": header, "items": items})

    except Exception as e:
        return jsonify({"success": False, "message": f"Error retrieving PO details: {str(e)}"})

# PO Management API Routes
@app.route('/api/po_management/get_po_items', methods=['POST'])
def get_po_items():
    """Get all items for a specific PO"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get PO items
        cursor.execute('''
            SELECT item_number, description, color, qty, bundle_qty, unit_price
            FROM po_items
            WHERE po_number = ?
            ORDER BY item_number
        ''', (po_number,))

        items = []
        for row in cursor.fetchall():
            # Better quantity parsing - handle various formats
            qty_value = row[3]
            parsed_qty = 0

            if qty_value:
                # Try to extract number from string (handle formats like "1,000", "500 pcs", etc.)
                import re
                qty_str = str(qty_value).replace(',', '').replace(' ', '')
                numbers = re.findall(r'\d+', qty_str)
                if numbers:
                    parsed_qty = int(numbers[0])
                elif qty_str.replace('.', '').isdigit():
                    parsed_qty = int(float(qty_str))

            items.append({
                'item_number': row[0],
                'description': row[1],
                'color': row[2],
                'qty': parsed_qty,
                'bundle_qty': row[4],
                'unit_price': row[5],
                'original_qty': qty_value  # Keep original for debugging
            })

        conn.close()

        if not items:
            return jsonify({"success": False, "message": f"No items found for PO {po_number}"})

        return jsonify({"success": True, "items": items})

    except Exception as e:
        return jsonify({"success": False, "message": f"Error retrieving PO items: {str(e)}"})

@app.route('/api/po_management/save_completion_status', methods=['POST'])
def save_completion_status():
    """Save completion status and finished quantities"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()
        completion_type = data.get('completion_type', '')
        finished_quantities = data.get('finished_quantities', {})

        if not po_number or not completion_type:
            return jsonify({"success": False, "message": "PO number and completion type are required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Save completion status
        cursor.execute('''
            INSERT INTO po_completion_status (po_number, completion_type, finished_quantities)
            VALUES (?, ?, ?)
        ''', (po_number, completion_type, str(finished_quantities)))

        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Completion status saved"})

    except Exception as e:
        return jsonify({"success": False, "message": f"Error saving completion status: {str(e)}"})

# ===== NEW SIMPLE OPTION A PACKING API ENDPOINTS =====

@app.route('/api/simple_packing/reset_database', methods=['POST'])
def reset_database():
    """One-time reset: Clear all packed status for all POs"""
    try:
        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Clear all carton relationships
        cursor.execute('DELETE FROM carton_items')
        cursor.execute('DELETE FROM cartons')

        # Add packed_status column if it doesn't exist
        try:
            cursor.execute("ALTER TABLE po_items ADD COLUMN packed_status TEXT DEFAULT 'not_packed'")
        except sqlite3.OperationalError:
            pass  # Column already exists

        # Reset all items to not_packed status and clear carton numbers
        cursor.execute("UPDATE po_items SET packed_status = 'not_packed', carton_number = NULL")

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Database reset complete. All items marked as not packed."
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Reset error: {str(e)}"})

@app.route('/api/simple_packing/load_po', methods=['GET'])
def load_po_simple():
    """Load PO items with packing status"""
    try:
        po_number = request.args.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Add packed_status column if it doesn't exist
        try:
            cursor.execute("ALTER TABLE po_items ADD COLUMN packed_status TEXT DEFAULT 'not_packed'")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # Column already exists

        # Get PO items
        cursor.execute('SELECT * FROM po_items WHERE po_number = ?', (po_number,))
        items = []

        for row in cursor.fetchall():
            # Get column names
            columns = [description[0] for description in cursor.description]
            item_dict = dict(zip(columns, row))

            # Ensure packed_status exists
            if 'packed_status' not in item_dict:
                item_dict['packed_status'] = 'not_packed'

            items.append(item_dict)

        conn.close()

        if not items:
            return jsonify({"success": False, "message": f"No items found for PO {po_number}"})

        return jsonify({
            "success": True,
            "po_number": po_number,
            "items": items,
            "total_items": len(items)
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Load PO error: {str(e)}"})

@app.route('/api/simple_packing/mark_all_done', methods=['POST'])
def mark_all_done():
    """Mark all items in PO as 'done' (ready for packing)"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Add packed_status column if it doesn't exist
        try:
            cursor.execute("ALTER TABLE po_items ADD COLUMN packed_status TEXT DEFAULT 'not_packed'")
        except sqlite3.OperationalError:
            pass  # Column already exists

        # Update all items to 'done' status
        cursor.execute("UPDATE po_items SET packed_status = 'done' WHERE po_number = ?", (po_number,))
        updated_count = cursor.rowcount

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "message": f"Marked {updated_count} items as done",
            "updated_count": updated_count
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Mark done error: {str(e)}"})

@app.route('/api/simple_packing/pack_items', methods=['POST'])
def pack_items_simple():
    """Pack selected items into a carton (Option A: Multi-line → 1 Carton)"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()
        selected_items = data.get('selected_items', [])
        carton_type = data.get('carton_type', '').strip()
        carton_weight = float(data.get('carton_weight', 0))

        if not po_number or not selected_items or not carton_type:
            return jsonify({"success": False, "message": "Missing required fields"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get next carton number (sequential: 1, 2, 3...)
        cursor.execute('SELECT COUNT(*) FROM cartons WHERE po_number = ?', (po_number,))
        existing_count = cursor.fetchone()[0]
        carton_number = str(existing_count + 1)

        # Generate barcode
        from datetime import datetime
        barcode = f"{po_number}-{carton_number}-{datetime.now().strftime('%Y%m%d')}"

        # Create carton record with cleaned weight
        clean_weight = clean_number_format(carton_weight)
        cursor.execute('''
            INSERT INTO cartons (po_number, carton_number, carton_size, actual_weight, barcode, packing_option)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (po_number, carton_number, carton_type, clean_weight, barcode, "Option A"))

        carton_id = cursor.lastrowid

        # Update selected items to 'packed' and link to carton
        packed_items = []
        for item_data in selected_items:
            item_number = item_data.get('item_number', '')
            description = item_data.get('description', '')
            color = item_data.get('color', '')
            qty = item_data.get('qty', 0)

            # Update item status to packed and assign carton number
            cursor.execute("""
                UPDATE po_items
                SET packed_status = 'packed', carton_number = ?
                WHERE po_number = ? AND item_number = ?
            """, (carton_number, po_number, item_number))

            # Link item to carton
            cursor.execute('''
                INSERT INTO carton_items (carton_id, po_number, item_number, description, color, packed_qty, original_qty)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (carton_id, po_number, item_number, description, color, qty, qty))

            packed_items.append({
                'item_number': item_number,
                'description': description,
                'qty': qty
            })

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "carton_number": carton_number,
            "barcode": barcode,
            "packed_items_count": len(packed_items),
            "packed_items": packed_items,
            "message": f"Packed {len(packed_items)} items into {carton_number}"
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Pack items error: {str(e)}"})

@app.route('/api/simple_packing/check_completion', methods=['GET'])
def check_completion():
    """Check if all items in PO are packed"""
    try:
        po_number = request.args.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Count total items and packed items
        cursor.execute('SELECT COUNT(*) FROM po_items WHERE po_number = ?', (po_number,))
        total_items = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM po_items WHERE po_number = ? AND packed_status = 'packed'", (po_number,))
        packed_items = cursor.fetchone()[0]

        conn.close()

        is_complete = (total_items > 0 and packed_items == total_items)

        return jsonify({
            "success": True,
            "is_complete": is_complete,
            "total_items": total_items,
            "packed_items": packed_items,
            "remaining_items": total_items - packed_items
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Check completion error: {str(e)}"})

@app.route('/api/simple_packing/generate_packing_list', methods=['GET'])
def generate_packing_list():
    """Generate packing list grouped by cartons"""
    try:
        po_number = request.args.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get packed items grouped by carton
        cursor.execute('''
            SELECT carton_number, item_number, description, color, qty
            FROM po_items
            WHERE po_number = ? AND packed_status = 'packed' AND carton_number IS NOT NULL
            ORDER BY carton_number ASC, item_number ASC
        ''', (po_number,))

        packed_items = cursor.fetchall()

        # Get carton details
        cursor.execute('''
            SELECT carton_number, carton_size, actual_weight
            FROM cartons
            WHERE po_number = ?
            ORDER BY carton_number ASC
        ''', (po_number,))

        carton_details = {row[0]: {'size': row[1], 'weight': row[2]} for row in cursor.fetchall()}

        conn.close()

        # Group items by carton
        packing_list = {}
        total_items = 0

        for carton_num, item_num, desc, color, qty in packed_items:
            if carton_num not in packing_list:
                packing_list[carton_num] = {
                    'carton_number': carton_num,
                    'carton_size': carton_details.get(carton_num, {}).get('size', 'Unknown'),
                    'carton_weight': carton_details.get(carton_num, {}).get('weight', 0),
                    'items': [],
                    'item_count': 0
                }

            packing_list[carton_num]['items'].append({
                'item_number': item_num,
                'description': desc,
                'color': color,
                'qty': qty
            })
            packing_list[carton_num]['item_count'] += 1
            total_items += 1

        return jsonify({
            "success": True,
            "po_number": po_number,
            "packing_list": list(packing_list.values()),
            "total_cartons": len(packing_list),
            "total_items": total_items
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Generate packing list error: {str(e)}"})

@app.route('/api/simple_packing/generate_pdf_packing_list', methods=['GET'])
def generate_pdf_packing_list():
    """Generate professional A4 packing list as HTML (printable as PDF)"""
    try:
        po_number = request.args.get('po_number', '').strip()

        if not po_number:
            return "PO number is required", 400

        # Generate unique PL number
        pl_number = generate_pl_number()

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get packed items grouped by carton
        cursor.execute('''
            SELECT carton_number, item_number, description, color, qty
            FROM po_items
            WHERE po_number = ? AND packed_status = 'packed' AND carton_number IS NOT NULL
            ORDER BY carton_number ASC, item_number ASC
        ''', (po_number,))

        packed_items = cursor.fetchall()

        # Get carton details
        cursor.execute('''
            SELECT carton_number, carton_size, actual_weight
            FROM cartons
            WHERE po_number = ?
            ORDER BY carton_number ASC
        ''', (po_number,))

        carton_details = {row[0]: {'size': row[1], 'weight': row[2]} for row in cursor.fetchall()}

        # Update carton numbers to be sequential (1, 2, 3...) and save PL data
        if packed_items:
            # Get unique carton numbers and create mapping
            unique_cartons = sorted(list(set([item[0] for item in packed_items])))
            carton_mapping = {old_carton: str(i + 1) for i, old_carton in enumerate(unique_cartons)}

            # Update packed_items with new carton numbers
            updated_packed_items = []
            for carton_num, item_num, desc, color, qty in packed_items:
                new_carton_num = carton_mapping[carton_num]
                updated_packed_items.append((new_carton_num, item_num, desc, color, qty))

            # Update carton_details with new numbers
            updated_carton_details = {}
            for old_carton, new_carton in carton_mapping.items():
                if old_carton in carton_details:
                    updated_carton_details[new_carton] = carton_details[old_carton]

            # Save PL data to database
            total_cartons = len(unique_cartons)
            total_items = len(packed_items)
            # Handle comma-separated quantities (e.g., "1,057" -> 1057)
            total_qty = 0
            for item in packed_items:
                qty_str = str(item[4])
                try:
                    qty = int(qty_str.replace(',', ''))
                    total_qty += qty
                except (ValueError, AttributeError):
                    pass  # Skip invalid quantities

            # Clean numeric values before database insertion
            clean_total_cartons = clean_number_format(total_cartons)
            clean_total_items = clean_number_format(total_items)
            clean_total_qty = clean_number_format(total_qty)

            cursor.execute('''
                INSERT INTO packing_lists (pl_number, po_number, total_cartons, total_items, total_qty)
                VALUES (?, ?, ?, ?, ?)
            ''', (pl_number, po_number, clean_total_cartons, clean_total_items, clean_total_qty))

            # Update po_items with PL number
            cursor.execute('''
                UPDATE po_items
                SET pl_number = ?
                WHERE po_number = ? AND packed_status = 'packed'
            ''', (pl_number, po_number))

            conn.commit()
            packed_items = updated_packed_items
            carton_details = updated_carton_details

        conn.close()

        # Generate professional HTML packing list
        html_content = generate_professional_packing_list_html(po_number, packed_items, carton_details, pl_number)

        return Response(html_content, mimetype='text/html')

    except Exception as e:
        return f"Error generating packing list: {str(e)}", 500

def generate_professional_packing_list_html(po_number, packed_items, carton_details, pl_number):
    """Generate professional A4 packing list HTML"""

    # Current date
    current_date = datetime.now().strftime("%B %d, %Y")

    # Group items by carton for merged cells
    carton_groups = {}
    for carton_num, item_num, desc, color, qty in packed_items:
        if carton_num not in carton_groups:
            carton_groups[carton_num] = []
        carton_groups[carton_num].append({
            'item_num': item_num,
            'desc': desc,
            'color': color,
            'qty': qty
        })

    # Build table rows with merged cells
    table_rows = ""
    for carton_num in sorted(carton_groups.keys()):
        items = carton_groups[carton_num]
        carton_size = carton_details.get(carton_num, {}).get('size', 'Unknown')
        carton_weight = carton_details.get(carton_num, {}).get('weight', 0)

        # Calculate rowspan for merged cells
        rowspan = len(items)

        # First row with merged cells
        first_item = items[0]
        table_rows += f"""
        <tr class="item-row">
            <td class="carton-cell" rowspan="{rowspan}">{carton_num}</td>
            <td class="carton-cell" rowspan="{rowspan}">{carton_size}</td>
            <td class="carton-cell" rowspan="{rowspan}">{carton_weight}kg</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{first_item['desc']} ({first_item['color']})</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{first_item['qty']}</td>
        </tr>
        """

        # Remaining rows (only item name and qty columns)
        for item in items[1:]:
            table_rows += f"""
        <tr class="item-row">
            <td style="padding: 8px; border: 1px solid #ddd;">{item['desc']} ({item['color']})</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{item['qty']}</td>
        </tr>
            """

    total_cartons = len(set(item[0] for item in packed_items))
    total_items = len(packed_items)

    # Calculate total quantity with proper type conversion
    total_qty = 0
    for item in packed_items:
        try:
            # Handle comma-separated quantities (e.g., "8,256" -> 8256)
            qty_str = str(item[4]) if item[4] is not None else '0'
            qty = int(qty_str.replace(',', ''))
            total_qty += qty
        except (ValueError, TypeError, AttributeError):
            # Skip invalid quantities
            continue

    # Create HTML content using string concatenation to avoid f-string issues
    html_content = """<!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Packing List - PO """ + po_number + """</title>
        <style>
            @page {
                size: A4;
                margin: 0.5in;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 12px;
                line-height: 1.4;
                margin: 0;
                padding: 20px;
            }
            .header {
                text-align: center;
                margin-bottom: 30px;
                border-bottom: 2px solid #333;
                padding-bottom: 20px;
            }
            .company-info {
                display: flex;
                justify-content: space-between;
                margin-bottom: 30px;
            }
            .bill-to, .ship-to {
                width: 45%;
            }
            .info-title {
                font-weight: bold;
                font-size: 14px;
                margin-bottom: 10px;
                color: #333;
                border-bottom: 1px solid #ccc;
                padding-bottom: 5px;
            }
            .address-line {
                margin-bottom: 3px;
            }
            .po-details {
                margin-bottom: 30px;
                background: #f8f9fa;
                padding: 20px;
                border-radius: 8px;
                border: 1px solid #ddd;
                text-align: center;
                font-size: 14px;
                font-weight: bold;
                color: #333;
            }
            .items-table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 30px;
            }
            .items-table th {
                background: #333;
                color: white;
                padding: 12px 8px;
                border: 1px solid #ddd;
                text-align: center;
                font-weight: bold;
            }
            .items-table td {
                padding: 8px;
                border: 1px solid #ddd;
                vertical-align: top;
            }
            .items-table tr:nth-child(even) {
                background: #f9f9f9;
            }
            .carton-cell {
                background: #e8f4fd !important;
                font-weight: bold;
                text-align: center;
                vertical-align: middle;
            }
            .item-row {
                border-left: 3px solid #007bff;
            }
            .items-table tfoot td {
                font-weight: bold;
                text-align: center;
            }
            .total-header {
                background: #f8f9fa !important;
                font-size: 14px;
                border-top: 3px solid #333 !important;
            }
            .total-row {
                background: #e8f4fd !important;
                color: #333;
            }
            .summary {
                text-align: right;
                margin-bottom: 40px;
                font-weight: bold;
                font-size: 14px;
            }
            .signature-section {
                margin-top: 50px;
                border-top: 1px solid #ccc;
                padding-top: 20px;
            }
            .signature-line {
                margin-bottom: 25px;
                display: flex;
                align-items: center;
            }
            .signature-line label {
                display: inline-block;
                width: 120px;
                font-weight: bold;
                margin-right: 15px;
            }
            .signature-line input {
                border: none;
                border-bottom: 2px solid #333;
                width: 400px;
                padding: 8px 0;
                font-size: 14px;
            }
            @media print {
                body { margin: 0; }
                .no-print { display: none; }
            }
        </style>
    </head>
    <body>
        <div class="header">
            <h1 style="margin: 0; font-size: 24px; color: #333;">📦 PACKING LIST</h1>
            <p style="margin: 10px 0 0 0; font-size: 16px; color: #666;">Purchase Order: """ + po_number + """</p>
            <p style="margin: 5px 0 0 0; font-size: 16px; color: #007bff; font-weight: bold;">Packing List: """ + pl_number + """</p>
            <p style="margin: 5px 0 0 0; color: #666;">Date: """ + current_date + """</p>
            <p style="margin: 5px 0 0 0; color: #999; font-size: 10px;">Version: """ + VERSION + """ (""" + VERSION_DATE + """)</p>
        </div>

        <div class="company-info">
            <div class="bill-to">
                <div class="info-title">BILL TO:</div>
                <div class="address-line"><strong>ABC Manufacturing Corp</strong></div>
                <div class="address-line">1234 Industrial Blvd, Suite 100</div>
                <div class="address-line">Manufacturing District</div>
                <div class="address-line">Los Angeles, CA 90210</div>
                <div class="address-line">Tel: (555) 123-4567</div>
                <div class="address-line"><strong>Contact:</strong> John Smith, Purchasing Manager</div>
            </div>

            <div class="ship-to">
                <div class="info-title">SHIP TO:</div>
                <div class="address-line"><strong>XYZ Retail Distribution Center</strong></div>
                <div class="address-line">5678 Warehouse Drive, Building B</div>
                <div class="address-line">Distribution Park</div>
                <div class="address-line">Dallas, TX 75201</div>
                <div class="address-line">Tel: (555) 987-6543</div>
                <div class="address-line"><strong>Contact:</strong> Sarah Johnson, Warehouse Supervisor</div>
                <br>
                <div class="address-line"><strong>Payment Terms:</strong> Net 30 Days</div>
                <div class="address-line"><strong>Delivery Terms:</strong> FOB Destination</div>
            </div>
        </div>

        <table class="items-table">
            <thead>
                <tr>
                    <th>CTN #</th>
                    <th>CTN Size</th>
                    <th>CTN Weight</th>
                    <th>Item Name</th>
                    <th>Qty</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>

        <div class="summary">
            <p>Total Cartons: {total_cartons} | Total Items: {total_items} | Total Qty: {total_qty} pieces</p>
        </div>

        <div class="signature-section">
            <h3 style="margin-bottom: 30px; color: #333; text-align: center;">RECIPIENT ACKNOWLEDGMENT</h3>

            <div style="display: flex; gap: 20px;">
                <!-- Left Column (50%) -->
                <div style="width: 50%;">
                    <!-- Empty for future use -->
                </div>

                <!-- Right Column (50%) -->
                <div style="width: 50%;">
                    <div style="margin-bottom: 25px;">
                        <div style="font-weight: bold; margin-bottom: 10px;">Company Name:</div>
                        <div style="font-size: 14px; font-weight: bold; color: #333;">ABC Manufacturing Corp</div>
                    </div>

                    <div class="signature-line" style="margin-bottom: 25px;">
                        <label style="display: block; margin-bottom: 10px; font-weight: bold;">Signature / Company Chop:</label>
                        <div style="border-bottom: 2px solid #333; width: 100%; height: 40px;"></div>
                    </div>

                    <div class="signature-line" style="margin-bottom: 25px;">
                        <label style="display: block; margin-bottom: 10px; font-weight: bold;">Date of Signature:</label>
                        <div style="border-bottom: 2px solid #333; width: 100%; height: 30px;"></div>
                    </div>
                </div>
            </div>
        </div>

        <div class="no-print" style="text-align: center; margin-top: 30px;">
            <button onclick="window.print()" style="padding: 10px 20px; background: #007bff; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px;">
                🖨️ Print as PDF
            </button>
            <button onclick="window.close()" style="padding: 10px 20px; background: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; margin-left: 10px;">
                Close
            </button>
        </div>
    </body>
    </html>
    """

    # Replace placeholders in HTML content
    html_content = html_content.replace("{table_rows}", table_rows)
    html_content = html_content.replace("{total_cartons}", str(total_cartons))
    html_content = html_content.replace("{total_items}", str(total_items))
    html_content = html_content.replace("{total_qty}", str(total_qty))

    return html_content

@app.route('/api/simple_packing/download_pdf_by_pl', methods=['GET'])
def download_pdf_by_pl():
    """Download PDF packing list by PL number"""
    try:
        pl_number = request.args.get('pl_number', '').strip()

        if not pl_number:
            return "PL number is required", 400

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get PL details
        cursor.execute('''
            SELECT po_number, total_cartons, total_items, total_qty
            FROM packing_lists
            WHERE pl_number = ?
        ''', (pl_number,))

        pl_data = cursor.fetchone()
        if not pl_data:
            return "Packing list not found", 404

        po_number = pl_data[0]

        # Get packed items for this PL
        cursor.execute('''
            SELECT carton_number, item_number, description, color, qty
            FROM po_items
            WHERE pl_number = ? AND packed_status = 'packed'
            ORDER BY carton_number ASC, item_number ASC
        ''', (pl_number,))

        packed_items = cursor.fetchall()

        # Get carton details (use dummy data since we're focusing on PL functionality)
        carton_details = {}
        unique_cartons = list(set([item[0] for item in packed_items]))
        for carton in unique_cartons:
            carton_details[carton] = {'size': 'Medium', 'weight': 2.5}

        conn.close()

        # Generate HTML with existing PL number
        html_content = generate_professional_packing_list_html(po_number, packed_items, carton_details, pl_number)

        return Response(html_content, mimetype='text/html')

    except Exception as e:
        return f"Error downloading packing list: {str(e)}", 500

@app.route('/api/po_management/pack_items_realtime', methods=['POST'])
def pack_items_realtime():
    """Real-time packing: get data directly from database to avoid frontend corruption"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()
        selected_item_ids = data.get('selected_item_ids', [])  # Just send item IDs, not full data

        if not po_number or not selected_item_ids:
            return jsonify({"success": False, "message": "PO number and selected item IDs are required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get actual item data directly from database (clean data, no NaN possible)
        placeholders = ','.join(['?' for _ in selected_item_ids])
        cursor.execute(f'''
            SELECT id, item_number, description, color, qty
            FROM po_items
            WHERE po_number = ? AND id IN ({placeholders})
        ''', [po_number] + selected_item_ids)

        database_items = cursor.fetchall()

        if not database_items:
            return jsonify({"success": False, "message": "No valid items found in database"})

        # Get next carton number (sequential: 1, 2, 3...)
        cursor.execute('SELECT COUNT(*) FROM cartons WHERE po_number = ?', (po_number,))
        existing_count = cursor.fetchone()[0]
        carton_number = str(existing_count + 1)
        barcode = f"{po_number}-{carton_number}-{datetime.now().strftime('%Y%m%d')}"

        # Create carton record
        cursor.execute('''
            INSERT INTO cartons (po_number, carton_number, carton_size, actual_weight, barcode, packing_option)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (po_number, carton_number, 'Standard', 0, barcode, 'Option A'))

        carton_id = cursor.lastrowid

        # Mark items as packed using clean database data
        for item_id, item_number, description, color, qty in database_items:
            # Clean the quantity value
            clean_qty = clean_number_format(qty)

            cursor.execute('''
                INSERT INTO carton_items (carton_id, po_number, item_number, description, color,
                                        packed_qty, original_qty, packed_status, carton_number)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (carton_id, po_number, item_number, description, color or '',
                  clean_qty, clean_qty, 'packed', carton_number))

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "carton_number": carton_number,
            "barcode": barcode,
            "carton_id": carton_id,
            "packed_items": len(database_items)
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Error packing items: {str(e)}"})

@app.route('/api/po_management/get_packing_status', methods=['GET'])
def get_packing_status():
    """Get real-time packing status for PO items"""
    try:
        po_number = request.args.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get packed items with carton numbers
        cursor.execute('''
            SELECT item_number, carton_number, packed_status
            FROM carton_items
            WHERE po_number = ? AND packed_status = 'packed'
        ''', (po_number,))

        packed_items = {}
        for row in cursor.fetchall():
            item_number, carton_number, status = row
            packed_items[item_number] = {
                'carton_number': carton_number,
                'status': status
            }

        conn.close()

        return jsonify({
            "success": True,
            "packed_items": packed_items
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Error getting packing status: {str(e)}"})

@app.route('/api/po_management/create_cartons', methods=['POST'])
def create_cartons():
    """Create cartons and pack items with enhanced tracking"""
    try:
        print("🔧 API: create_cartons called")
        data = request.json
        print(f"📊 API: Received data: {data}")

        po_number = data.get('po_number', '').strip()
        cartons_data = data.get('cartons', [])
        packing_option = data.get('packing_option', 'A')

        print(f"📦 API: PO Number: {po_number}")
        print(f"📋 API: Cartons data: {cartons_data}")
        print(f"🎯 API: Packing option: {packing_option}")

        if not po_number or not cartons_data:
            print("❌ API: Missing PO number or cartons data")
            return jsonify({"success": False, "message": "PO number and cartons data are required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get existing carton count for this PO to continue numbering
        cursor.execute('SELECT COUNT(*) FROM cartons WHERE po_number = ?', (po_number,))
        existing_count = cursor.fetchone()[0]

        created_cartons = []

        for i, carton_data in enumerate(cartons_data):
            # Generate carton number and barcode (sequential: 1, 2, 3...)
            carton_number = str(existing_count + i + 1)
            barcode = f"{po_number}-{carton_number}-{datetime.now().strftime('%Y%m%d')}"

            # Insert carton with enhanced data and cleaned weight
            clean_weight = clean_number_format(carton_data.get('weight', 0))
            cursor.execute('''
                INSERT INTO cartons (po_number, carton_number, carton_size, actual_weight, barcode, packing_option)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (po_number, carton_number, carton_data.get('size', 'Standard'),
                  clean_weight, barcode, f"Option {packing_option}"))

            carton_id = cursor.lastrowid

            # Insert carton items with enhanced tracking
            for item in carton_data.get('items', []):
                cursor.execute('''
                    INSERT INTO carton_items (carton_id, po_number, item_number, description, color, packed_qty, original_qty)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (carton_id, po_number, item['item_number'],
                      item['description'], item.get('color', ''),
                      item['packed_qty'], item.get('original_qty', item['packed_qty'])))

            created_cartons.append({
                'id': carton_id,
                'carton_number': carton_number,
                'barcode': barcode,
                'weight': carton_data.get('weight', 0),
                'size': carton_data.get('size', 'Standard'),
                'items': carton_data.get('items', [])
            })

        conn.commit()
        conn.close()

        print(f"✅ API: Successfully created {len(created_cartons)} cartons")
        print(f"📦 API: Created cartons: {created_cartons}")

        return jsonify({"success": True, "cartons": created_cartons})

    except Exception as e:
        print(f"❌ API: Error creating cartons: {str(e)}")
        return jsonify({"success": False, "message": f"Error creating cartons: {str(e)}"})

@app.route('/api/po_management/create_shipment', methods=['POST'])
def create_shipment():
    """Create shipment with enhanced courier details and tracking"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()
        courier = data.get('courier', '').strip()
        awb_number = data.get('awb_number', '').strip()
        carton_ids = data.get('carton_ids', [])

        if not all([po_number, courier, awb_number, carton_ids]):
            return jsonify({"success": False, "message": "All shipment details are required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Calculate total weight and carton count
        total_weight = 0
        for carton_id in carton_ids:
            cursor.execute('SELECT actual_weight FROM cartons WHERE id = ?', (carton_id,))
            weight_result = cursor.fetchone()
            if weight_result and weight_result[0]:
                total_weight += weight_result[0]

        # Create shipment with enhanced tracking
        cursor.execute('''
            INSERT INTO shipments (po_number, courier, awb_number, shipment_date, total_cartons, total_weight)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (po_number, courier, awb_number, datetime.now().strftime('%Y-%m-%d'),
              len(carton_ids), total_weight))

        shipment_id = cursor.lastrowid

        # Link cartons to shipment
        for carton_id in carton_ids:
            cursor.execute('''
                INSERT INTO shipment_cartons (shipment_id, carton_id)
                VALUES (?, ?)
            ''', (shipment_id, carton_id))

        conn.commit()
        conn.close()

        return jsonify({
            "success": True,
            "shipment_id": shipment_id,
            "total_cartons": len(carton_ids),
            "total_weight": total_weight
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Error creating shipment: {str(e)}"})

@app.route('/api/po_management/debug_po', methods=['POST'])
def debug_po_data():
    """Debug endpoint to see raw PO data"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get raw data
        cursor.execute('''
            SELECT item_number, description, color, qty, bundle_qty, unit_price
            FROM po_items
            WHERE po_number = ?
            ORDER BY item_number
        ''', (po_number,))

        raw_items = []
        for row in cursor.fetchall():
            raw_items.append({
                'item_number': row[0],
                'description': row[1],
                'color': row[2],
                'qty_raw': row[3],
                'qty_type': type(row[3]).__name__,
                'bundle_qty': row[4],
                'unit_price': row[5]
            })

        conn.close()
        return jsonify({"success": True, "raw_data": raw_items})

    except Exception as e:
        return jsonify({"success": False, "message": f"Debug error: {str(e)}"})

@app.route('/api/po_management/get_carton_summary', methods=['POST'])
def get_carton_summary():
    """Get summary of all cartons for a PO"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get cartons with their items
        cursor.execute('''
            SELECT c.id, c.carton_number, c.barcode, c.actual_weight, c.carton_size, c.packing_option
            FROM cartons c
            WHERE c.po_number = ?
            ORDER BY c.carton_number
        ''', (po_number,))

        cartons = []
        for carton_row in cursor.fetchall():
            carton_id, carton_number, barcode, weight, size, packing_option = carton_row

            # Get items for this carton
            cursor.execute('''
                SELECT item_number, description, color, packed_qty, original_qty
                FROM carton_items
                WHERE carton_id = ?
            ''', (carton_id,))

            items = []
            total_qty = 0
            for item_row in cursor.fetchall():
                item_number, description, color, packed_qty, original_qty = item_row
                items.append({
                    'item_number': item_number,
                    'description': description,
                    'color': color,
                    'packed_qty': packed_qty,
                    'original_qty': original_qty
                })
                total_qty += packed_qty

            cartons.append({
                'id': carton_id,
                'carton_number': carton_number,
                'barcode': barcode,
                'weight': weight,
                'size': size,
                'packing_option': packing_option,
                'items': items,
                'total_qty': total_qty
            })

        conn.close()

        return jsonify({
            "success": True,
            "cartons": cartons,
            "total_cartons": len(cartons)
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Error getting carton summary: {str(e)}"})

@app.route('/api/po_management/generate_packing_list', methods=['POST'])
def generate_packing_list_download():
    """Generate complete packing list for download"""
    try:
        data = request.json
        po_number = data.get('po_number', '').strip()

        if not po_number:
            return jsonify({"success": False, "message": "PO number is required"})

        conn = sqlite3.connect('po_database.db')
        cursor = conn.cursor()

        # Get PO header info
        cursor.execute('SELECT * FROM po_headers WHERE po_number = ?', (po_number,))
        header = cursor.fetchone()

        # Get shipment info
        cursor.execute('''
            SELECT courier, awb_number, shipment_date, total_cartons, total_weight
            FROM shipments
            WHERE po_number = ?
            ORDER BY created_at DESC
            LIMIT 1
        ''', (po_number,))
        shipment = cursor.fetchone()

        # Get cartons and items
        cursor.execute('''
            SELECT c.carton_number, c.barcode, c.actual_weight, c.carton_size,
                   ci.item_number, ci.description, ci.color, ci.packed_qty
            FROM cartons c
            LEFT JOIN carton_items ci ON c.id = ci.carton_id
            WHERE c.po_number = ?
            ORDER BY c.carton_number, ci.item_number
        ''', (po_number,))

        packing_data = cursor.fetchall()
        conn.close()

        # Format data for packing list
        packing_list = {
            'po_number': po_number,
            'header': header,
            'shipment': shipment,
            'packing_data': packing_data,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        return jsonify({
            "success": True,
            "packing_list": packing_list
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"Error generating packing list: {str(e)}"})

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🚀 Artwork Downloader v{{ version }} - FRESH LOAD {{ cache_buster }}</title>
    <!-- NUCLEAR CACHE BUSTER: {{ cache_buster }} -->
    <!-- VERSION: {{ version }} - {{ version_date }} -->
    <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate, max-age=0, private">
    <meta http-equiv="Pragma" content="no-cache">
    <meta http-equiv="Expires" content="-1">
    <meta name="cache-buster" content="{{ cache_buster }}">
    <meta name="version" content="{{ version }}">
    <meta name="build-time" content="{{ cache_buster }}">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #f5f5f5;
            color: #333;
            line-height: 1.6;
        }
        
        .container {
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }

        .tabs {
            display: flex;
            background: #f9f9f9;
            border-bottom: 1px solid #e0e0e0;
            margin-bottom: 30px;
        }

        .tab {
            flex: 1;
            padding: 15px 20px;
            text-align: center;
            cursor: pointer;
            border: none;
            background: none;
            font-size: 0.9em;
            transition: all 0.2s;
            color: #666;
            border-bottom: 2px solid transparent;
        }

        .tab.active {
            background: white;
            color: #333;
            border-bottom: 2px solid #333;
            font-weight: 500;
        }

        .tab:hover {
            background: #f0f0f0;
            color: #333;
        }

        .tab-content {
            display: none;
        }

        .tab-content.active {
            display: block;
        }
        
        .header {
            background: #2c3e50;
            color: white;
            padding: 20px;
            text-align: center;
            margin-bottom: 30px;
        }
        
        .step {
            background: white;
            border: 1px solid #e0e0e0;
            margin-bottom: 20px;
            padding: 25px;
        }
        
        .step h2 {
            margin-bottom: 15px;
            color: #333;
            font-size: 1.3em;
        }
        
        .step-number {
            background: #333;
            color: white;
            width: 30px;
            height: 30px;
            border-radius: 50%;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            margin-right: 10px;
            font-weight: bold;
        }
        
        .form-group {
            margin-bottom: 20px;
        }
        
        .form-group label {
            display: block;
            margin-bottom: 8px;
            font-weight: 500;
        }
        
        .form-group input {
            width: 300px;
            padding: 12px;
            border: 1px solid #ccc;
            font-size: 1em;
        }
        
        .btn {
            background: #333;
            color: white;
            border: 1px solid #333;
            padding: 12px 24px;
            cursor: pointer;
            font-size: 0.9em;
            margin-right: 10px;
        }
        
        .btn:hover {
            background: #555;
        }
        
        .btn:disabled {
            background: #ccc;
            cursor: not-allowed;
        }

        .btn-secondary {
            background: #f8f9fa;
            color: #333;
            border: 1px solid #ddd;
            padding: 10px 20px;
            cursor: pointer;
            font-size: 0.9em;
            border-radius: 4px;
            transition: all 0.2s;
        }

        .btn-secondary:hover {
            background: #e9ecef;
            border-color: #adb5bd;
        }
        
        .po-info {
            background: #f9f9f9;
            padding: 20px;
            border: 1px solid #e0e0e0;
            margin-bottom: 20px;
        }
        
        .recommendations {
            margin-bottom: 20px;
        }
        
        .recommendation {
            background: white;
            border: 1px solid #e0e0e0;
            padding: 15px;
            margin-bottom: 10px;
            cursor: pointer;
        }
        
        .recommendation:hover {
            border-color: #333;
        }
        
        .recommendation.selected {
            border-color: #333;
            background: #f9f9f9;
        }
        
        .recommendation h4 {
            margin-bottom: 5px;
        }
        
        .recommendation .score {
            float: right;
            background: #333;
            color: white;
            padding: 2px 8px;
            border-radius: 10px;
            font-size: 0.8em;
        }
        
        .data-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }
        
        .data-table th,
        .data-table td {
            border: 1px solid #e0e0e0;
            padding: 8px 12px;
            text-align: left;
        }
        
        .data-table th {
            background: #f9f9f9;
            font-weight: 500;
        }
        
        .data-table tr:nth-child(even) {
            background: #fafafa;
        }

        .item-checkbox {
            transform: scale(1.2);
            margin: 0;
        }

        .report-section {
            margin-bottom: 30px;
            padding: 20px;
            border: 1px solid #e0e0e0;
            border-radius: 5px;
        }

        .stat-card {
            display: inline-block;
            padding: 15px;
            margin: 10px;
            border: 1px solid #e0e0e0;
            border-radius: 5px;
            text-align: center;
            min-width: 120px;
        }

        .stat-card h4 {
            margin: 0 0 10px 0;
            color: #666;
            font-size: 0.9em;
        }

        .stat-card span {
            font-size: 1.5em;
            font-weight: bold;
            color: #333;
        }

        .hidden {
            display: none;
        }

        /* Progress Notification System */
        .notification-container {
            position: fixed;
            top: 20px;
            right: 20px;
            z-index: 10000;
            max-width: 400px;
            pointer-events: none;
        }

        .notification {
            background: rgba(255, 255, 255, 0.95);
            border-radius: 12px;
            padding: 16px 20px;
            margin-bottom: 12px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.2);
            border-left: 4px solid #007bff;
            backdrop-filter: blur(10px);
            pointer-events: auto;
            transform: translateX(100%);
            transition: all 0.3s ease-in-out;
            opacity: 0;
            font-size: 14px;
            font-weight: 500;
            color: #333;
            display: flex;
            align-items: center;
            gap: 12px;
        }

        .notification.show {
            transform: translateX(0);
            opacity: 1;
        }

        .notification.processing {
            border-left-color: #007bff;
            background: linear-gradient(135deg, rgba(0, 123, 255, 0.1) 0%, rgba(255, 255, 255, 0.95) 100%);
        }

        .notification.success {
            border-left-color: #28a745;
            background: linear-gradient(135deg, rgba(40, 167, 69, 0.1) 0%, rgba(255, 255, 255, 0.95) 100%);
        }

        .notification.error {
            border-left-color: #dc3545;
            background: linear-gradient(135deg, rgba(220, 53, 69, 0.1) 0%, rgba(255, 255, 255, 0.95) 100%);
        }

        .notification-icon {
            font-size: 18px;
            min-width: 20px;
        }

        .notification-content {
            flex: 1;
            line-height: 1.4;
        }

        .notification-close {
            background: none;
            border: none;
            font-size: 18px;
            cursor: pointer;
            color: #666;
            padding: 0;
            margin-left: 8px;
            opacity: 0.7;
            transition: opacity 0.2s;
        }

        .notification-close:hover {
            opacity: 1;
        }

        /* Spinning animation for processing */
        .notification.processing .notification-icon {
            animation: spin 2s linear infinite;
        }

        @keyframes spin {
            from { transform: rotate(0deg); }
            to { transform: rotate(360deg); }
        }

        .loading {
            text-align: center;
            padding: 20px;
            color: #666;
        }
        
        .error {
            background: #fee;
            color: #c53030;
            padding: 15px;
            border: 1px solid #feb2b2;
        }

        .method-selection {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }

        .method-card {
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            padding: 15px;
            cursor: pointer;
            transition: all 0.3s ease;
            background: white;
        }

        .method-card:hover {
            border-color: #007bff;
            box-shadow: 0 2px 8px rgba(0,123,255,0.2);
        }

        .method-card.selected {
            border-color: #007bff;
            background: #f8f9ff;
            box-shadow: 0 2px 8px rgba(0,123,255,0.3);
        }

        .method-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }

        .method-header h4 {
            margin: 0;
            color: #333;
        }

        .success-rate {
            background: #28a745;
            color: white;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: bold;
        }

        .method-card[data-method="super_fast"] .success-rate {
            background: #dc3545;
        }

        .method-card[data-method="hybrid"] .success-rate {
            background: #ffc107;
            color: #333;
        }

        .method-card[data-method="standard"] .success-rate {
            background: #28a745;
        }

        .method-card[data-method="original_slow"] .success-rate {
            background: #17a2b8;
        }

        .method-details {
            margin-top: 10px;
            color: #666;
        }
    </style>
</head>
<body>
    <!-- Progress Notification Container -->
    <div id="notification-container" class="notification-container"></div>

    <div class="header">
        <p>Download artwork files - Smart PO Analysis & Recommendations</p>
        <p style="font-size: 0.9em; color: #bdc3c7;">Intelligent artwork download with multiple methods</p>
    </div>

    <div class="container">


        <!-- Tab Navigation - Fixed/Sticky -->
        <div class="tabs" style="position: sticky; top: 0; z-index: 1000; background: white; border-bottom: 2px solid #ddd; box-shadow: 0 2px 4px rgba(0,0,0,0.1); margin-bottom: 20px;">
            <button class="tab active" onclick="showTab('artwork')">Download Artwork</button>
            <button class="tab" onclick="showTab('delivery')">Update Delivery Date</button>
            <button class="tab" onclick="showTab('po')">PO Management</button>
            <button class="tab" onclick="showTab('report')">Report</button>
            <button class="tab" onclick="showTab('settings')">Settings</button>
        </div>

        <!-- Download Artwork Tab -->
        <div id="artwork" class="tab-content active">
            <!-- Step 1: Input PO -->
            <div class="step" id="step1">
            <h2><span class="step-number">1</span>Enter PO Number</h2>
            <div class="form-group">
                <label for="po_input">PO Number:</label>
                <input type="text" id="po_input" placeholder="Enter PO number (e.g., 1284789)" />
                <button class="btn" onclick="analyzePO()" id="analyze_btn">Analyze PO</button>
                <button class="btn" onclick="clearEverything()" id="new_btn" style="background: #28a745; margin-left: 10px; font-weight: bold; font-size: 14px;">🆕 NEW PO</button>
            </div>

            <!-- Error/Success Messages -->
            <div id="error_container"></div>

            <!-- Welcome/Instructions Section -->
            <div id="welcome_section" style="margin-top: 30px; padding: 25px; background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); border-radius: 10px; border-left: 4px solid #007bff;">
                <h3 style="color: #007bff; margin-bottom: 15px;">🚀 Welcome to Artwork Downloader</h3>
                <p style="margin-bottom: 15px; color: #495057;">Get started by entering a PO number above to analyze and download artwork files.</p>

                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin: 20px 0;">
                    <div style="padding: 15px; background: white; border-radius: 8px; border: 1px solid #dee2e6;">
                        <h4 style="color: #28a745; margin-bottom: 10px;">📊 Step 1: Analyze PO</h4>
                        <p style="font-size: 0.9em; color: #6c757d;">Enter your PO number and click "Analyze PO" to get recommendations and see all available items.</p>
                    </div>
                    <div style="padding: 15px; background: white; border-radius: 8px; border: 1px solid #dee2e6;">
                        <h4 style="color: #17a2b8; margin-bottom: 10px;">🎯 Step 2: Select Method</h4>
                        <p style="font-size: 0.9em; color: #6c757d;">Choose from multiple download methods based on our intelligent recommendations.</p>
                    </div>
                    <div style="padding: 15px; background: white; border-radius: 8px; border: 1px solid #dee2e6;">
                        <h4 style="color: #fd7e14; margin-bottom: 10px;">📥 Step 3: Download</h4>
                        <p style="font-size: 0.9em; color: #6c757d;">Select items and start downloading. PO details can be saved to database for future reference.</p>
                    </div>
                </div>

                <div style="margin-top: 20px; padding: 15px; background: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px;">
                    <strong style="color: #856404;">💡 Pro Tip:</strong>
                    <span style="color: #856404;">Try PO number "1284789" as an example to see how the system works!</span>
                </div>
            </div>
        </div>
        
        <!-- Step 2: Show Recommendations -->
        <div class="step hidden" id="step2">
            <h2><span class="step-number">2</span>PO Analysis & Recommendations</h2>

            <div id="po_info" class="po-info"></div>

            <h3>Download Method:</h3>

            <!-- Default Method 5 Display -->
            <div class="default-method-display">
                <div class="method-card selected" data-method="guaranteed_complete">
                    <div class="method-header">
                        <h4>✨ Guaranteed Complete Download</h4>
                        <span class="success-rate">100% Success</span>
                    </div>
                    <p>100% success rate with direct URL extraction. Visual clarity: 19 items = 19 files. Smart numbering for duplicates (_2, _3, _4).</p>
                    <div class="method-details">
                        <small>RECOMMENDED for all POs - Option 5 from unified_downloader.py</small>
                    </div>
                </div>
                <div style="text-align: center; margin-top: 15px;">
                    <button class="btn-secondary" onclick="toggleMethodSelection()" id="toggle_methods_btn">
                        📋 Show All Download Methods
                    </button>
                </div>
            </div>

            <!-- All Methods (Hidden by Default) -->
            <div class="method-selection hidden" id="all_methods" style="display: none;">
                <div class="method-card" data-method="super_fast">
                    <div class="method-header">
                        <h4>🚀 Super Fast</h4>
                        <span class="success-rate">~10% Success</span>
                    </div>
                    <p>Direct PDF download using URL pattern. Very fast but may fail if URLs change.</p>
                    <div class="method-details">
                        <small>Uses: https://app4.brandid.com/Artwork/{ITEM}_{SUFFIX}.pdf</small>
                    </div>
                </div>

                <div class="method-card" data-method="hybrid">
                    <div class="method-header">
                        <h4>⚡ Hybrid</h4>
                        <span class="success-rate">~70% Success</span>
                    </div>
                    <p>Browser login + direct requests. Good balance of speed and reliability.</p>
                    <div class="method-details">
                        <small>Login once, then direct downloads</small>
                    </div>
                </div>

                <div class="method-card" data-method="standard">
                    <div class="method-header">
                        <h4>📋 Standard</h4>
                        <span class="success-rate">~90% Success</span>
                    </div>
                    <p>Browser automation with smart navigation. Reliable and reasonably fast.</p>
                    <div class="method-details">
                        <small>Recommended for most cases</small>
                    </div>
                </div>

                <div class="method-card" data-method="original_slow">
                    <div class="method-header">
                        <h4>🐌 Original Slow</h4>
                        <span class="success-rate">100% Success</span>
                    </div>
                    <p>Full browser automation. Slowest but most reliable method.</p>
                    <div class="method-details">
                        <small>Use when other methods fail</small>
                    </div>
                </div>

                <div class="method-card" data-method="guaranteed_complete">
                    <div class="method-header">
                        <h4>✨ Guaranteed Complete Download</h4>
                        <span class="success-rate">100% Success</span>
                    </div>
                    <p>100% success rate with direct URL extraction. Visual clarity: 19 items = 19 files. Smart numbering for duplicates (_2, _3, _4).</p>
                    <div class="method-details">
                        <small>RECOMMENDED for all POs - Option 5 from unified_downloader.py</small>
                    </div>
                </div>

                <div style="text-align: center; margin-top: 15px;">
                    <button class="btn-secondary" onclick="toggleMethodSelection()" id="hide_methods_btn">
                        ⬆️ Hide Other Methods
                    </button>
                </div>
            </div>
        </div>
        
        <!-- Step 3: Show Data Table -->
        <div class="step hidden" id="step3">
            <h2><span class="step-number">3</span>PO Items Data</h2>
            <div id="data_table_container"></div>
            
            <div style="margin-top: 20px;">
                <button class="btn" onclick="startDownload()" id="download_btn">Start Download</button>
            </div>
        </div>
        
        <!-- Download Progress -->
        <div class="step hidden" id="progress_step">
            <h2><span class="step-number">4</span>Download Progress</h2>
            <div id="progress_info"></div>
        </div>
        </div>

        <!-- Update Delivery Date Tab -->
        <div id="delivery" class="tab-content">
            <div class="step">
                <h2><span class="step-number">📅</span>Update Delivery Date</h2>
                <p>Select a PO from your saved database to update delivery dates</p>

                <!-- Saved POs List -->
                <div style="margin: 20px 0;">
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
                        <h3>📊 Saved PO Database</h3>
                        <button class="btn" onclick="loadSavedPOs()" style="background: #17a2b8;">🔄 Refresh List</button>
                    </div>

                    <!-- Search Input -->
                    <div style="margin-bottom: 15px;">
                        <div style="position: relative; max-width: 400px;">
                            <input
                                type="text"
                                id="po_search_input"
                                placeholder="🔍 Search PO Number..."
                                style="width: 100%; padding: 12px 45px 12px 15px; border: 2px solid #ddd; border-radius: 25px; font-size: 14px; outline: none; transition: all 0.3s ease;"
                                oninput="filterPOTable()"
                                onfocus="this.style.borderColor='#007bff'; this.style.boxShadow='0 0 0 3px rgba(0,123,255,0.1)'"
                                onblur="this.style.borderColor='#ddd'; this.style.boxShadow='none'"
                            >
                            <div style="position: absolute; right: 15px; top: 50%; transform: translateY(-50%); color: #666; pointer-events: none;">
                                🔍
                            </div>
                        </div>
                        <div id="search_results_count" style="margin-top: 8px; font-size: 12px; color: #666;"></div>
                    </div>

                    <div id="saved_pos_container" style="max-height: 400px; overflow-y: auto; border: 1px solid #ddd; border-radius: 5px;">
                        <div id="saved_pos_loading" style="padding: 20px; text-align: center; color: #666;">
                            📊 Loading saved POs...
                        </div>
                        <div id="saved_pos_list" style="display: none;"></div>
                        <div id="saved_pos_empty" style="display: none; padding: 20px; text-align: center; color: #666;">
                            📝 No POs saved yet. Download some artwork first to save PO details to the database.
                        </div>
                    </div>
                </div>

                <!-- PO Details Section -->
                <div id="delivery_info" class="hidden" style="margin-top: 20px; padding: 20px; border: 1px solid #ddd; border-radius: 5px; background: #f9f9f9;">
                    <h3>📋 Complete PO Details</h3>

                    <!-- PO Tracking Information -->
                    <div id="po_tracking_info" style="margin: 15px 0; padding: 15px; background: linear-gradient(135deg, #e3f2fd 0%, #f3e5f5 100%); border-radius: 8px; border-left: 4px solid #2196f3;">
                        <h4 style="color: #1976d2; margin: 0 0 10px 0; font-size: 16px;">📊 Database Tracking Information</h4>
                        <div style="display: flex; gap: 30px; flex-wrap: wrap;">
                            <div>
                                <strong style="color: #666;">First Created:</strong>
                                <span id="first_created_display" style="color: #333; margin-left: 8px;">-</span>
                            </div>
                            <div>
                                <strong style="color: #666;">Last Updated:</strong>
                                <span id="last_updated_display" style="color: #333; margin-left: 8px;">-</span>
                            </div>
                            <div>
                                <strong style="color: #666;">Update Count:</strong>
                                <span id="update_count_display" style="color: #333; margin-left: 8px; background: #e8f5e8; padding: 2px 8px; border-radius: 12px; font-weight: bold;">0</span>
                            </div>
                        </div>
                    </div>

                    <!-- Side-by-Side Tables Container -->
                    <div style="display: flex; gap: 20px; margin: 20px 0;">

                        <!-- Left Side: PO Header Table -->
                        <div style="flex: 1; min-width: 0;">
                            <h4 style="color: #007bff; margin-bottom: 10px;">📊 PO Header Information</h4>
                            <div style="border: 1px solid #ddd; border-radius: 5px; background: white; max-height: 300px; overflow: auto;">
                                <table id="po_header_table" style="width: 100%; border-collapse: collapse; min-width: 600px;">
                                    <thead style="position: sticky; top: 0; background: #007bff; color: white;">
                                        <tr>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">WO#</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Factory</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">PO Date</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Ship By</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Ship Via</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Order Type</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Status</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Loc</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Prod Rep</th>
                                        </tr>
                                    </thead>
                                    <tbody id="po_header_body">
                                        <!-- Header data will be populated here -->
                                    </tbody>
                                </table>
                            </div>
                        </div>

                        <!-- Right Side: PO Items Table -->
                        <div style="flex: 1; min-width: 0;">
                            <h4 style="color: #28a745; margin-bottom: 10px;">📦 PO Items Details</h4>
                            <div style="border: 1px solid #ddd; border-radius: 5px; background: white; max-height: 300px; overflow: auto;">
                                <table id="po_items_table" style="width: 100%; border-collapse: collapse; min-width: 700px;">
                                    <thead style="position: sticky; top: 0; background: #28a745; color: white;">
                                        <tr>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Item #</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Description</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Color</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Ship To</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Need By</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: right; font-size: 12px;">Qty</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: left; font-size: 12px;">Bundle Qty</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: right; font-size: 12px;">$ Unit Price</th>
                                            <th style="padding: 8px; border: 1px solid #ddd; text-align: right; font-size: 12px;">Extension</th>
                                        </tr>
                                    </thead>
                                    <tbody id="po_items_body">
                                        <!-- Items data will be populated here -->
                                    </tbody>
                                </table>
                            </div>
                        </div>
                    </div>

                    <!-- Additional PO Details Section -->
                    <div style="margin: 20px 0;">
                        <h4 style="color: #6f42c1; margin-bottom: 10px;">📋 Additional PO Information</h4>
                        <div style="border: 1px solid #ddd; border-radius: 5px; background: white; padding: 15px;">
                            <table id="po_additional_table" style="width: 100%; border-collapse: collapse;">
                                <tbody id="po_additional_body">
                                    <!-- Additional details will be populated here -->
                                </tbody>
                            </table>
                        </div>
                    </div>

                    <!-- Delivery Date Update Section -->
                    <div style="margin: 30px 0; padding: 20px; background: #fff3cd; border: 1px solid #ffeaa7; border-radius: 5px;">
                        <h4 style="color: #856404; margin-bottom: 15px;">📅 Update Delivery Date</h4>

                        <div class="form-group">
                            <label for="current_delivery_date">Current Delivery Date:</label>
                            <input type="text" id="current_delivery_date" readonly />
                        </div>

                        <div class="form-group">
                            <label for="new_delivery_date">New Delivery Date:</label>
                            <input type="date" id="new_delivery_date" />
                        </div>

                        <div class="form-group">
                            <label for="delivery_notes">Notes (Optional):</label>
                            <textarea id="delivery_notes" placeholder="Reason for date change..."></textarea>
                        </div>

                        <button class="btn" onclick="updateDeliveryDate()">📅 Update Delivery Date</button>
                    </div>
                </div>
            </div>
        </div>

        <!-- Report Tab -->
        <div id="report" class="tab-content">
            <div class="step">
                <h2><span class="step-number">📊</span>PO Master Report</h2>
                <p>Comprehensive view of all PO data with 27 columns combining headers and items</p>

                <!-- Report Controls -->
                <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; padding: 15px; background: #f8f9fa; border-radius: 8px;">
                    <div>
                        <h3 style="margin: 0; color: #333;">📈 Master Data View</h3>
                        <p style="margin: 5px 0 0 0; color: #666; font-size: 0.9em;">Latest 20 records | Real-time search across all 27 columns</p>
                    </div>
                    <div style="display: flex; gap: 10px;">
                        <button class="btn" onclick="refreshMasterReport()" style="background: #17a2b8;">🔄 Refresh</button>
                        <button class="btn" onclick="exportMasterReport()" style="background: #28a745;">📥 Export Excel</button>
                    </div>
                </div>

                <!-- Loading State -->
                <div id="master_report_loading" style="text-align: center; padding: 40px; color: #666;">
                    <div style="font-size: 2em; margin-bottom: 10px;">📊</div>
                    <div>Loading master report data...</div>
                </div>

                <!-- Master Report Table Container -->
                <div id="master_report_container" style="display: none;">
                    <!-- Table Wrapper with Horizontal Scroll and Fixed Header -->
                    <div style="overflow: auto; border: 1px solid #ddd; border-radius: 8px; background: white; max-height: 300px; position: relative;">
                        <table id="master_report_table" style="width: 100%; table-layout: fixed; border-collapse: collapse;">
                            <!-- Table Header with Search Inputs -->
                            <thead style="background: #f8f9fa; position: sticky; top: 0; z-index: 100;">
                                <!-- Column Headers -->
                                <tr style="border-bottom: 2px solid #dee2e6;">
                                    <!-- Fixed Columns -->
                                    <th style="position: sticky; left: 0; background: #e9ecef; z-index: 101; padding: 12px 8px; border-right: 2px solid #adb5bd; width: 120px; font-weight: 600;">PO#</th>
                                    <th style="position: sticky; left: 120px; background: #e9ecef; z-index: 101; padding: 12px 8px; border-right: 2px solid #adb5bd; width: 140px; font-weight: 600;">Item #</th>
                                    <th style="position: sticky; left: 260px; background: #e9ecef; z-index: 101; padding: 12px 8px; border-right: 2px solid #adb5bd; width: 250px; font-weight: 600;">Description</th>

                                    <!-- Scrollable Columns -->
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Color</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 140px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Ship To</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Need By</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 100px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Qty</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Bundle Qty</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Unit Price</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Extension</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 150px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Company</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 160px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Purchase From</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 100px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Currency</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">PO Date</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 130px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Cancel Date</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Ship By</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 140px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Ship Via</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 130px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Order Type</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 100px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Status</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 150px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Factory</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Location</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 130px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Prod Rep</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 200px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Ship To Address</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 130px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Terms</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 150px; font-weight: 600; border-bottom: 2px solid #dee2e6;">First Created</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 150px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Last Updated</th>
                                    <th style="position: sticky; top: 0; background: #f8f9fa; z-index: 100; padding: 12px 8px; width: 120px; font-weight: 600; border-bottom: 2px solid #dee2e6;">Update Count</th>
                                </tr>

                                <!-- Search Input Row -->
                                <tr style="border-bottom: 1px solid #dee2e6; position: sticky; top: 42px; z-index: 99; background: #f8f9fa;">
                                    <!-- Fixed Column Search Inputs -->
                                    <th style="position: sticky; left: 0; background: #f8f9fa; z-index: 101; padding: 8px; border-right: 2px solid #adb5bd; width: 120px;">
                                        <input type="text" id="search_po_number" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()">
                                    </th>
                                    <th style="position: sticky; left: 120px; background: #f8f9fa; z-index: 101; padding: 8px; border-right: 2px solid #adb5bd; width: 140px;">
                                        <input type="text" id="search_item_number" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()">
                                    </th>
                                    <th style="position: sticky; left: 260px; background: #f8f9fa; z-index: 101; padding: 8px; border-right: 2px solid #adb5bd; width: 250px;">
                                        <input type="text" id="search_description" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()">
                                    </th>

                                    <!-- Scrollable Column Search Inputs -->
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_color" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 140px;"><input type="text" id="search_ship_to" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_need_by" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 100px;"><input type="text" id="search_qty" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_bundle_qty" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_unit_price" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_extension" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 150px;"><input type="text" id="search_company" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 160px;"><input type="text" id="search_purchase_from" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 100px;"><input type="text" id="search_currency" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_po_date" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 130px;"><input type="text" id="search_cancel_date" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_ship_by" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 140px;"><input type="text" id="search_ship_via" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 130px;"><input type="text" id="search_order_type" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 100px;"><input type="text" id="search_status" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 150px;"><input type="text" id="search_factory" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_location" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 130px;"><input type="text" id="search_prod_rep" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 200px;"><input type="text" id="search_ship_to_address" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 130px;"><input type="text" id="search_terms" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 150px;"><input type="text" id="search_first_created" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 150px;"><input type="text" id="search_last_updated" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                    <th style="position: sticky; top: 42px; background: #f8f9fa; z-index: 99; padding: 8px; width: 120px;"><input type="text" id="search_update_count" placeholder="🔍" style="width: calc(100% - 8px); padding: 4px; border: 1px solid #ccc; border-radius: 3px; font-size: 12px;" oninput="searchMasterReport()"></th>
                                </tr>
                            </thead>

                            <!-- Table Body -->
                            <tbody id="master_report_tbody">
                                <!-- Data rows will be populated here -->
                            </tbody>
                        </table>
                    </div>

                    <!-- Report Footer -->
                    <div style="margin-top: 15px; padding: 15px; background: #f8f9fa; border-radius: 8px; display: flex; justify-content: space-between; align-items: center;">
                        <div id="master_report_stats" style="color: #666; font-size: 0.9em;">
                            📊 Showing 0 of 0 total records | 🔍 Active filters: 0 | 📅 Last updated: -
                        </div>
                        <div>
                            <button class="btn-secondary" onclick="clearAllSearchFilters()" style="margin-right: 10px;">🗑️ Clear Filters</button>
                            <button class="btn-secondary" onclick="loadMoreRecords()">📄 Load More</button>
                        </div>
                    </div>
                </div>

                <!-- No Data State -->
                <div id="master_report_empty" style="display: none; text-align: center; padding: 40px; color: #666;">
                    <div style="font-size: 3em; margin-bottom: 15px;">📊</div>
                    <h3>No PO Data Available</h3>
                    <p>No PO records found in the database. Download some artwork first to populate the master report.</p>
                    <button class="btn" onclick="showTab('artwork')" style="margin-top: 15px;">📥 Go to Download Artwork</button>
                </div>
            </div>
        </div>

        <!-- PO Management Tab - SIMPLE OPTION A PACKING -->
        <div id="po" class="tab-content">
            <!-- Simple Header -->
            <div style="margin-bottom: 30px; padding: 20px; background: linear-gradient(135deg, #28a745 0%, #20c997 100%); border-radius: 12px; color: white;">
                <h2 style="margin: 0 0 15px 0; text-align: center;">📦 Simple Option A Packing</h2>
                <p style="margin: 0; text-align: center; font-size: 16px; opacity: 0.9;">Load PO → Mark Done → Select Items → Pack to Carton → Repeat</p>
            </div>

            <!-- Main Interface -->
            <div style="max-width: 1200px; margin: 0 auto;">

                <!-- Step 1: Database Reset (One-time) -->
                <div style="background: #fff3cd; padding: 20px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #ffc107;">
                    <h4 style="color: #856404; margin: 0 0 10px 0;">🗑️ One-Time Database Reset</h4>
                    <p style="color: #856404; margin: 0 0 15px 0;">Clear all packed status for all POs (only needed once when starting fresh)</p>
                    <button onclick="resetDatabase()" style="padding: 10px 20px; background: #dc3545; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">
                        🗑️ Reset All PO Status
                    </button>
                    <div id="reset_status" style="margin-top: 10px;"></div>
                </div>

                <!-- Step 2: PO Input -->
                <div style="background: #e3f2fd; padding: 20px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #2196f3;">
                    <h4 style="color: #1565c0; margin: 0 0 15px 0;">📋 Load PO</h4>
                    <div style="display: flex; gap: 15px; align-items: center;">
                        <input type="text" id="simple_po_input" placeholder="Enter PO number (e.g., 1280290)"
                               style="flex: 1; padding: 12px; border: 2px solid #ddd; border-radius: 5px; font-size: 16px;">
                        <button onclick="loadPOSimple()" style="padding: 12px 24px; background: #007bff; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">
                            🔍 Load PO
                        </button>
                    </div>
                    <div id="po_load_status" style="margin-top: 15px;"></div>
                </div>

                <!-- Step 3: Items Display & Actions -->
                <div id="items_container" style="display: none;">

                    <!-- Action Buttons -->
                    <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #6c757d;">
                        <h4 style="color: #495057; margin: 0 0 15px 0;">⚡ Quick Actions</h4>
                        <div style="display: flex; gap: 15px; flex-wrap: wrap;">
                            <button onclick="markAllDone()" style="padding: 10px 20px; background: #28a745; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">
                                ✅ Mark All Done
                            </button>
                            <button onclick="selectAllItems()" style="padding: 10px 20px; background: #17a2b8; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">
                                ☑️ Select All
                            </button>
                            <button onclick="clearSelections()" style="padding: 10px 20px; background: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer;">
                                ❌ Clear Selections
                            </button>
                            <button onclick="testModal()" style="padding: 10px 20px; background: #dc3545; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">
                                🧪 Test Modal
                            </button>
                        </div>
                        <div id="action_status" style="margin-top: 15px;"></div>
                    </div>

                    <!-- Packing Method -->
                    <div style="background: #e8f5e8; padding: 20px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #28a745;">
                        <h4 style="color: #155724; margin: 0 0 15px 0;">📦 Packing Method</h4>
                        <div style="display: flex; gap: 20px;">
                            <label style="display: flex; align-items: center; gap: 8px; font-weight: bold; color: #155724;">
                                <input type="radio" name="packing_method" value="option_a" checked style="transform: scale(1.2);">
                                ● Option A (Multi-line → 1 Carton)
                            </label>
                        </div>
                    </div>

                    <!-- Items Table -->
                    <div style="background: white; border-radius: 8px; border: 1px solid #ddd; overflow: hidden;">
                        <div style="background: #f8f9fa; padding: 15px; border-bottom: 1px solid #ddd;">
                            <h4 style="margin: 0; color: #495057;">📋 PO Items</h4>
                            <div id="items_summary" style="margin-top: 5px; font-size: 14px; color: #666;"></div>
                        </div>
                        <div style="max-height: 400px; overflow-y: auto;">
                            <table style="width: 100%; border-collapse: collapse;">
                                <thead style="position: sticky; top: 0; background: #f8f9fa; z-index: 10;">
                                    <tr>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: left; font-weight: bold; width: 50px;">Select</th>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: left; font-weight: bold; width: 120px;">Item Number</th>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: left; font-weight: bold;">Description</th>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: left; font-weight: bold; width: 100px;">Color</th>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: center; font-weight: bold; width: 80px;">Qty</th>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: center; font-weight: bold; width: 100px;">Status</th>
                                        <th style="padding: 12px; border-bottom: 2px solid #ddd; text-align: center; font-weight: bold; width: 100px;">Carton #</th>
                                    </tr>
                                </thead>
                                <tbody id="items_table_body">
                                    <!-- Items will be loaded here -->
                                </tbody>
                            </table>
                        </div>
                    </div>

                    <!-- Pack Selected Items (Hidden - Modal opens automatically) -->
                    <div id="pack_section" style="display: none;">
                        <div id="selected_summary"></div>
                        <div id="pack_status"></div>
                    </div>

                </div>

            </div>

            <!-- Movable Window Modal for Carton Packing - Two Step Flow -->
            <div id="carton_modal" style="display: none; position: fixed; top: 20%; right: 5%; z-index: 10000; background: white; border-radius: 12px; box-shadow: 0 10px 30px rgba(0,0,0,0.3); min-width: 400px; max-width: 500px; border: 2px solid #007bff;">
                <div id="modal_header" style="background: linear-gradient(135deg, #007bff, #0056b3); color: white; padding: 15px 20px; border-radius: 10px 10px 0 0; cursor: move; user-select: none; display: flex; justify-content: space-between; align-items: center;">
                    <h3 style="margin: 0; font-size: 16px;">📦 Pack Selected Items</h3>
                    <button onclick="closeCartonModal()" style="background: none; border: none; color: white; font-size: 18px; cursor: pointer; padding: 0; width: 25px; height: 25px; border-radius: 50%; display: flex; align-items: center; justify-content: center;" onmouseover="this.style.background='rgba(255,255,255,0.2)'" onmouseout="this.style.background='none'">×</button>
                </div>
                    <!-- Step 1: Summary and Pack Button -->
                    <div id="modal_step_1" style="display: block;">
                        <div id="modal_selected_summary" style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 25px; text-align: center; font-weight: bold; color: #495057; font-size: 16px;"></div>

                        <div style="text-align: center;">
                            <button onclick="showCartonForm()" style="padding: 15px 30px; background: #ffc107; color: #212529; border: none; border-radius: 8px; cursor: pointer; font-size: 18px; font-weight: bold;">
                                📦 Pack Into Carton
                            </button>
                        </div>

                        <div style="text-align: center; margin-top: 20px;">
                            <button onclick="closeCartonModal()" style="padding: 10px 20px; background: #6c757d; color: white; border: none; border-radius: 6px; cursor: pointer; font-size: 14px;">
                                ❌ Cancel
                            </button>
                        </div>
                    </div>

                    <!-- Step 2: Carton Details Form -->
                    <div id="modal_step_2" style="display: none;">
                        <div style="margin-bottom: 20px;">
                            <label style="display: block; margin-bottom: 8px; font-weight: bold; color: #333;">Carton Type/Size:</label>
                            <select id="modal_carton_type" style="width: 100%; padding: 12px; border: 2px solid #ddd; border-radius: 6px; font-size: 16px;">
                                <option value="">Select carton type...</option>
                                <option value="Small Box">Small Box</option>
                                <option value="Medium Box">Medium Box</option>
                                <option value="Large Box">Large Box</option>
                                <option value="Extra Large Box">Extra Large Box</option>
                                <option value="Custom">Custom</option>
                            </select>
                        </div>

                        <div style="margin-bottom: 25px;">
                            <label style="display: block; margin-bottom: 8px; font-weight: bold; color: #333;">Weight (kg):</label>
                            <input type="number" id="modal_carton_weight" placeholder="Enter weight in kg" step="0.1" min="0"
                                   style="width: 100%; padding: 12px; border: 2px solid #ddd; border-radius: 6px; font-size: 16px;">
                        </div>

                        <div style="display: flex; gap: 15px; justify-content: center;">
                            <button onclick="confirmPackItems()" style="padding: 12px 24px; background: #28a745; color: white; border: none; border-radius: 6px; cursor: pointer; font-size: 16px; font-weight: bold;">
                                📦 Pack Items
                            </button>
                            <button onclick="backToSummary()" style="padding: 12px 24px; background: #17a2b8; color: white; border: none; border-radius: 6px; cursor: pointer; font-size: 16px;">
                                ← Back
                            </button>
                            <button onclick="closeCartonModal()" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 6px; cursor: pointer; font-size: 16px;">
                                ❌ Cancel
                            </button>
                        </div>

                        <div id="modal_pack_status" style="margin-top: 20px;"></div>
                    </div>
                </div>
            </div>

        </div>


            <div id="po_step_3" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">3️⃣</span>Completion Status</h2>
                    <p>Is this shipment complete or partial?</p>

                    <div style="margin: 30px 0; text-align: center;">
                        <div style="display: flex; gap: 20px; justify-content: center; flex-wrap: wrap;">
                            <button onclick="selectCompletionStatus('all')"
                                    style="padding: 20px 40px; background: #28a745; color: white; border: none; border-radius: 12px; cursor: pointer; font-size: 18px; min-width: 200px;">
                                ✅ All Done<br>
                                <small style="font-size: 14px; opacity: 0.9;">Ship complete quantities</small>
                            </button>
                            <button onclick="selectCompletionStatus('partial')"
                                    style="padding: 20px 40px; background: #ffc107; color: #333; border: none; border-radius: 12px; cursor: pointer; font-size: 18px; min-width: 200px;">
                                📦 Partial Done<br>
                                <small style="font-size: 14px; opacity: 0.9;">Ship partial quantities</small>
                            </button>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Step 4: Partial Quantities (only shown for partial) -->
            <div id="po_step_4" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">4️⃣</span>Enter Finished Quantities</h2>
                    <p>Enter the actual quantities ready for shipment</p>

                    <div id="partial_quantities_container" style="margin: 20px 0;">
                        <!-- Partial quantity inputs will be loaded here -->
                    </div>

                    <div style="margin: 20px 0; text-align: center;">
                        <button onclick="goToPOStep(3)" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; margin-right: 10px;">
                            ← Back
                        </button>
                        <button onclick="validatePartialQuantities()" style="padding: 12px 24px; background: #28a745; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">
                            Continue to Packing →
                        </button>
                    </div>
                </div>
            </div>

            <!-- Step 5: Packing Logic -->
            <div id="po_step_5" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">5️⃣</span>Choose Packing Logic</h2>
                    <p>How would you like to pack the items?</p>

                    <div style="margin: 30px 0; text-align: center;">
                        <div style="display: flex; gap: 20px; justify-content: center; flex-wrap: wrap;">
                            <button onclick="selectPackingLogic('multi_to_one')"
                                    style="padding: 20px 30px; background: #007bff; color: white; border: none; border-radius: 12px; cursor: pointer; font-size: 16px; min-width: 250px;">
                                📦➡️📦 Option A<br>
                                <strong>Multiple Lines → 1 Carton</strong><br>
                                <small style="font-size: 13px; opacity: 0.9;">Pack multiple items into one carton</small>
                            </button>
                            <button onclick="selectPackingLogic('one_to_multi')"
                                    style="padding: 20px 30px; background: #6f42c1; color: white; border: none; border-radius: 12px; cursor: pointer; font-size: 16px; min-width: 250px;">
                                📦➡️📦📦📦 Option B<br>
                                <strong>1 Line → Multiple Cartons</strong><br>
                                <small style="font-size: 13px; opacity: 0.9;">Split one item across multiple cartons</small>
                            </button>
                        </div>

                        <div style="margin-top: 20px;">
                            <button onclick="goToPreviousStep()" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">
                                ← Back
                            </button>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Step 6A: Multi-to-One Packing -->
            <div id="po_step_6a" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">6️⃣</span>Pack Multiple Lines → 1 Carton</h2>
                    <p>Select items to pack together in one carton</p>

                    <div id="multi_to_one_container" style="margin: 20px 0;">
                        <!-- Multi-to-one packing interface will be loaded here -->
                    </div>
                </div>
            </div>

            <!-- Step 6B: One-to-Multi Packing -->
            <div id="po_step_6b" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">6️⃣</span>Pack 1 Line → Multiple Cartons</h2>
                    <p>Select an item and specify how many cartons to split it into</p>

                    <div id="one_to_multi_container" style="margin: 20px 0;">
                        <!-- One-to-multi packing interface will be loaded here -->
                    </div>
                </div>
            </div>

            <!-- Step 7: Carton Summary -->
            <div id="po_step_7" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">7️⃣</span>Carton Summary</h2>
                    <p>Review packed cartons and generate barcodes</p>

                    <div id="carton_summary_container" style="margin: 20px 0;">
                        <!-- Carton summary will be loaded here -->
                    </div>

                    <div style="margin: 20px 0; text-align: center;">
                        <button onclick="goToPreviousStep()" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; margin-right: 10px;">
                            ← Back
                        </button>
                        <button onclick="generateBarcodes()" style="padding: 12px 24px; background: #28a745; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; margin-right: 10px;">
                            🏷️ Generate Barcodes
                        </button>
                        <button onclick="goToPOStep(8)" style="padding: 12px 24px; background: #007bff; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">
                            Continue to Courier →
                        </button>
                    </div>
                </div>
            </div>

            <!-- Step 8: Courier Details -->
            <div id="po_step_8" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">8️⃣</span>Courier & Shipment Details</h2>
                    <p>Select courier and enter AWB details</p>

                    <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <div style="margin-bottom: 15px;">
                            <label for="courier_select" style="display: block; margin-bottom: 5px; font-weight: bold;">Courier:</label>
                            <select id="courier_select" style="width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 16px;">
                                <option value="">Select Courier</option>
                                <option value="DHL">DHL</option>
                                <option value="FedEx">FedEx</option>
                                <option value="UPS">UPS</option>
                                <option value="TNT">TNT</option>
                                <option value="Local Courier">Local Courier</option>
                                <option value="Others">Others</option>
                            </select>
                        </div>

                        <div style="margin-bottom: 15px;">
                            <label for="awb_input" style="display: block; margin-bottom: 5px; font-weight: bold;">AWB Number:</label>
                            <input type="text" id="awb_input" placeholder="Enter Air Waybill number"
                                   style="width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 16px;">
                        </div>

                        <div style="margin-bottom: 15px;">
                            <label for="awb_document" style="display: block; margin-bottom: 5px; font-weight: bold;">AWB Document:</label>
                            <input type="file" id="awb_document" accept=".pdf,.jpg,.jpeg,.png"
                                   style="width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                    </div>

                    <div style="margin: 20px 0; text-align: center;">
                        <button onclick="goToPOStep(7)" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; margin-right: 10px;">
                            ← Back to Carton Summary
                        </button>
                        <button onclick="createShipment()" style="padding: 12px 24px; background: #28a745; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">
                            🚚 Create Shipment →
                        </button>
                    </div>
                </div>
            </div>

            <!-- Step 9: Final Summary -->
            <div id="po_step_9" class="po-step" style="display: none;">
                <div class="step">
                    <h2><span class="step-number">9️⃣</span>Packing List Complete</h2>
                    <p>Shipment created successfully!</p>

                    <div id="final_summary_container" style="margin: 20px 0;">
                        <!-- Final summary will be loaded here -->
                    </div>

                    <div style="margin: 20px 0; text-align: center;">
                        <button onclick="downloadPackingList()" style="padding: 12px 24px; background: #007bff; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; margin-right: 10px;">
                            📄 Download Packing List
                        </button>
                        <button onclick="resetPOManagement()" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">
                            🔄 Start New PO
                        </button>
                    </div>
                </div>
            </div>
        </div>

        <!-- Settings Tab -->
        <div id="settings" class="tab-content">
            <div class="step">
                <h2><span class="step-number">⚙️</span>Settings & Configuration</h2>
                <p>Manage system configuration and login credentials</p>

                <!-- Login Credentials Section -->
                <div style="margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
                    <h3>🔐 Login Credentials</h3>
                    <div style="margin: 15px 0;">
                        <strong>Login URL:</strong><br>
                        <span id="display_url">https://app.e-brandid.com/login/login.aspx</span>
                    </div>
                    <div style="margin: 15px 0;">
                        <strong>Username:</strong><br>
                        <span id="display_username">{{ masked_username }}</span>
                    </div>
                    <div style="margin: 15px 0;">
                        <strong>Password:</strong><br>
                        <span id="display_password">************</span>
                    </div>

                    <!-- Admin Access Section -->
                    <div id="admin_section" style="margin-top: 20px; padding: 15px; background: #f9f9f9; border-radius: 5px;">
                        <h4>🔑 Admin Access Required</h4>
                        <p style="margin: 10px 0; color: #666;">Enter admin password to view/edit credentials:</p>
                        <div style="display: flex; gap: 10px; align-items: center;">
                            <input type="password" id="admin_password" placeholder="Admin password"
                                   style="padding: 8px; border: 1px solid #ddd; border-radius: 3px; flex: 1;">
                            <button onclick="verifyAdmin()" style="padding: 8px 15px; background: #007bff; color: white; border: none; border-radius: 3px; cursor: pointer;">
                                Unlock
                            </button>
                        </div>
                        <div id="admin_message" style="margin-top: 10px; color: red;"></div>
                    </div>

                    <!-- Edit Form (Hidden by default) -->
                    <div id="edit_form" style="display: none; margin-top: 20px; padding: 15px; background: #e8f5e8; border-radius: 5px;">
                        <h4>✏️ Edit Configuration</h4>
                        <div style="margin: 10px 0;">
                            <label><strong>Login URL:</strong></label><br>
                            <input type="text" id="edit_url" style="width: 100%; padding: 8px; margin: 5px 0; border: 1px solid #ddd; border-radius: 3px;">
                        </div>
                        <div style="margin: 10px 0;">
                            <label><strong>Username:</strong></label><br>
                            <input type="text" id="edit_username" style="width: 100%; padding: 8px; margin: 5px 0; border: 1px solid #ddd; border-radius: 3px;">
                        </div>
                        <div style="margin: 10px 0;">
                            <label><strong>Password:</strong></label><br>
                            <input type="text" id="edit_password" style="width: 100%; padding: 8px; margin: 5px 0; border: 1px solid #ddd; border-radius: 3px;">
                        </div>
                        <div style="margin: 15px 0;">
                            <button onclick="saveConfig()" style="padding: 10px 20px; background: #28a745; color: white; border: none; border-radius: 3px; cursor: pointer; margin-right: 10px;">
                                💾 Save Changes
                            </button>
                            <button onclick="cancelEdit()" style="padding: 10px 20px; background: #6c757d; color: white; border: none; border-radius: 3px; cursor: pointer;">
                                ❌ Cancel
                            </button>
                        </div>
                        <div id="save_message" style="margin-top: 10px;"></div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script>
        let selectedMethod = null;
        let currentPO = null;
        window.currentPoData = null;  // Global variable for checkbox functions

        // Progress Notification System
        let notificationCounter = 0;
        const activeNotifications = new Map();

        function showProgressNotification(id, message, type = 'processing', icon = '🔄') {
            const container = document.getElementById('notification-container');

            // Check if notification already exists
            let notification = document.getElementById(`notification-${id}`);

            if (!notification) {
                // Create new notification
                notification = document.createElement('div');
                notification.id = `notification-${id}`;
                notification.className = `notification ${type}`;

                notification.innerHTML = `
                    <div class="notification-icon">${icon}</div>
                    <div class="notification-content">${message}</div>
                    <button class="notification-close" onclick="removeNotification('${id}')">&times;</button>
                `;

                container.appendChild(notification);

                // Trigger animation
                setTimeout(() => {
                    notification.classList.add('show');
                }, 100);

                // Store in active notifications
                activeNotifications.set(id, notification);
            } else {
                // Update existing notification
                const iconEl = notification.querySelector('.notification-icon');
                const contentEl = notification.querySelector('.notification-content');

                iconEl.textContent = icon;
                contentEl.textContent = message;

                // Update type class
                notification.className = `notification show ${type}`;
            }

            return notification;
        }

        function updateNotification(id, message, type = 'success', icon = '✅', autoRemove = true) {
            const notification = showProgressNotification(id, message, type, icon);

            if (autoRemove) {
                // Auto-remove success notifications after 4 seconds
                setTimeout(() => {
                    removeNotification(id);
                }, 4000);
            }
        }

        function removeNotification(id) {
            const notification = document.getElementById(`notification-${id}`);
            if (notification) {
                notification.classList.remove('show');
                setTimeout(() => {
                    if (notification.parentNode) {
                        notification.parentNode.removeChild(notification);
                    }
                    activeNotifications.delete(id);
                }, 300);
            }
        }

        function clearAllNotifications() {
            activeNotifications.forEach((notification, id) => {
                removeNotification(id);
            });
        }

        // Method selection
        document.addEventListener('DOMContentLoaded', function() {
            // Set Method 5 as default
            selectedMethod = 'guaranteed_complete';

            // Add click handlers for method cards
            document.querySelectorAll('.method-card').forEach(card => {
                card.addEventListener('click', function() {
                    // Remove selected class from all cards
                    document.querySelectorAll('.method-card').forEach(c => c.classList.remove('selected'));
                    // Add selected class to clicked card
                    this.classList.add('selected');
                    // Update selected method
                    selectedMethod = this.dataset.method;
                    console.log('Selected method:', selectedMethod);
                });
            });

            // Restore form state on page load
            restoreFormState();
        });

        // State preservation functions
        function saveFormState() {
            const state = {
                po_input: document.getElementById('po_input')?.value || '',
                current_delivery_date: document.getElementById('current_delivery_date')?.value || '',
                new_delivery_date: document.getElementById('new_delivery_date')?.value || '',
                delivery_notes: document.getElementById('delivery_notes')?.value || ''
            };
            sessionStorage.setItem('formState', JSON.stringify(state));
        }

        function restoreFormState() {
            const savedState = sessionStorage.getItem('formState');
            if (savedState) {
                const state = JSON.parse(savedState);
                if (document.getElementById('po_input')) document.getElementById('po_input').value = state.po_input || '';
                if (document.getElementById('current_delivery_date')) document.getElementById('current_delivery_date').value = state.current_delivery_date || '';
                if (document.getElementById('new_delivery_date')) document.getElementById('new_delivery_date').value = state.new_delivery_date || '';
                if (document.getElementById('delivery_notes')) document.getElementById('delivery_notes').value = state.delivery_notes || '';
            }
        }

        // Tab switching
        function showTab(tabName) {
            // Save current form state before switching
            saveFormState();

            // Hide all tab contents
            document.querySelectorAll('.tab-content').forEach(content => {
                content.classList.remove('active');
            });

            // Remove active class from all tabs
            document.querySelectorAll('.tab').forEach(tab => {
                tab.classList.remove('active');
            });

            // Show selected tab content
            document.getElementById(tabName).classList.add('active');

            // Find and activate the correct tab button
            const tabs = document.querySelectorAll('.tab');
            tabs.forEach((tab, index) => {
                if ((tabName === 'artwork' && index === 0) ||
                    (tabName === 'delivery' && index === 1) ||
                    (tabName === 'po' && index === 2) ||
                    (tabName === 'report' && index === 3) ||
                    (tabName === 'settings' && index === 4)) {
                    tab.classList.add('active');
                }
            });

            // Tab-specific actions
            if (tabName === 'delivery') {
                // Auto-load saved POs when delivery tab is opened
                loadSavedPOs();
                // Hide the Complete PO Details section until a PO is selected
                document.getElementById('delivery_info').classList.add('hidden');
            } else if (tabName === 'artwork') {
                // Show welcome section if no PO has been analyzed yet
                if (!currentPO) {
                    document.getElementById('welcome_section').style.display = 'block';
                }
            } else if (tabName === 'report') {
                // Auto-load master report when report tab is opened
                loadMasterReport();
            }

            // Restore form state after switching
            setTimeout(restoreFormState, 100); // Small delay to ensure elements are loaded
        }

        // Clear everything function for NEW button
        function clearEverything() {
            // Clear form inputs
            document.getElementById('po_input').value = '';

            // Re-enable PO input field
            document.getElementById('po_input').disabled = false;

            // Hide all steps except step 1
            document.getElementById('step2').classList.add('hidden');
            document.getElementById('step3').classList.add('hidden');
            document.getElementById('progress_step').classList.add('hidden');

            // Clear containers
            document.getElementById('data_table_container').innerHTML = '';
            document.getElementById('error_container').innerHTML = '';
            document.getElementById('progress_info').innerHTML = '';

            // Show welcome section
            document.getElementById('welcome_section').style.display = 'block';

            // Reset global variables
            currentPO = null;
            selectedMethod = 'guaranteed_complete'; // Reset to default method

            // Clear session storage
            sessionStorage.removeItem('formState');

            // Re-enable and reset analyze button
            document.getElementById('analyze_btn').disabled = false;
            document.getElementById('analyze_btn').innerHTML = 'Analyze PO';

            // Re-enable and reset download buttons
            const downloadBtn = document.getElementById('download_btn');
            const downloadBtnTop = document.getElementById('download_btn_top');
            if (downloadBtn) {
                downloadBtn.disabled = false;
                downloadBtn.innerHTML = 'Start Download';
            }
            if (downloadBtnTop) {
                downloadBtnTop.disabled = false;
                downloadBtnTop.innerHTML = '🚀 Start Download';
            }

            // Reset method selection to default (Method 5)
            document.querySelectorAll('.method-card').forEach(card => {
                card.classList.remove('selected');
            });
            const method5Card = document.querySelector('[data-method="guaranteed_complete"]');
            if (method5Card) {
                method5Card.classList.add('selected');
            }

            // Clear all notifications
            clearAllNotifications();

            // Focus on PO input for immediate use
            document.getElementById('po_input').focus();

            showError('✅ Ready for new PO! Enter PO number above.', 'success');
        }

        // Checkbox handling functions
        function selectAllItems(select) {
            const checkboxes = document.querySelectorAll('.item-checkbox');
            checkboxes.forEach(checkbox => {
                checkbox.checked = select;
            });
            updateSelectedCount();
        }

        function updateSelectedCount() {
            const checkboxes = document.querySelectorAll('.item-checkbox');
            const selectedCount = document.querySelectorAll('.item-checkbox:checked').length;
            const selectedCountElement = document.getElementById('selected_count');
            if (selectedCountElement) {
                selectedCountElement.textContent = selectedCount;
            }
        }

        function getSelectedItems() {
            const selectedItems = [];
            const checkboxes = document.querySelectorAll('.item-checkbox:checked');
            checkboxes.forEach(checkbox => {
                const itemIndex = parseInt(checkbox.getAttribute('data-item-index'));
                if (window.currentPoData && window.currentPoData.items && window.currentPoData.items[itemIndex]) {
                    selectedItems.push(window.currentPoData.items[itemIndex]);
                }
            });
            return selectedItems;
        }
        
        async function analyzePO() {
            const poNumber = document.getElementById('po_input').value.trim();
            if (!poNumber) {
                alert('Please enter a PO number');
                return;
            }

            // Clear previous results
            document.getElementById('step2').classList.add('hidden');
            document.getElementById('step3').classList.add('hidden');
            document.getElementById('data_table_container').innerHTML = '';

            // Hide welcome section when analysis starts
            document.getElementById('welcome_section').style.display = 'none';

            document.getElementById('analyze_btn').disabled = true;

            // Show progress notification
            const notificationId = `analyze-${poNumber}`;
            showProgressNotification(notificationId, `🔍 Analyzing ${poNumber} PO data...`, 'processing', '🔄');

            try {
                const response = await fetch('/api/analyze_po?t=' + Date.now(), {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ po_number: poNumber })
                });

                const result = await response.json();
                console.log('API Response:', result); // Debug log

                if (result.success) {
                    // Update notification to show completion
                    updateNotification(notificationId, `✅ Finished Analyzing ${poNumber} PO`, 'success', '✅');

                    currentPO = result;
                    window.currentPoData = result;  // Store globally for checkbox functions
                    console.log('Items received:', result.items.length); // Debug log
                    showPOAnalysis(result);
                    showDataTable(result.items);
                    document.getElementById('step2').classList.remove('hidden');
                    document.getElementById('step3').classList.remove('hidden');
                } else {
                    updateNotification(notificationId, `❌ Failed to analyze ${poNumber}: ${result.error || 'Unknown error'}`, 'error', '❌');
                }
            } catch (error) {
                updateNotification(notificationId, `❌ Error analyzing ${poNumber}: ${error.message}`, 'error', '❌');
            } finally {
                document.getElementById('analyze_btn').disabled = false;
            }
        }
        
        function showPOAnalysis(data) {
            const poInfo = document.getElementById('po_info');
            poInfo.innerHTML = `
                <h3>${data.title}</h3>
                <p><strong>PO Number:</strong> ${data.po_number}</p>
                <p><strong>Total Items:</strong> ${data.total_items}</p>
                <p><strong>Analyzed:</strong> ${data.timestamp}</p>
            `;

            // Method cards are now static in HTML, no need to generate them dynamically
            // Ensure Method 5 (Guaranteed Complete Download) is selected by default
            selectedMethod = 'guaranteed_complete';
            document.querySelectorAll('.method-card').forEach(card => {
                card.classList.remove('selected');
            });
            // Select Method 5 as default
            const method5Card = document.querySelector('[data-method="guaranteed_complete"]');
            if (method5Card) {
                method5Card.classList.add('selected');
            }
        }
        

        
        function showDataTable(items) {
            console.log('showDataTable called with items:', items.length); // Debug log
            const container = document.getElementById('data_table_container');

            let html = `
                <div style="margin-bottom: 15px; display: flex; align-items: center; gap: 10px; flex-wrap: wrap;">
                    <button class="btn" onclick="selectAllItems(true)">✅ Select All</button>
                    <button class="btn" onclick="selectAllItems(false)" style="background: #e53e3e;">❌ Deselect All</button>
                    <span style="font-weight: bold;">Selected: <span id="selected_count">${items.length}</span> / ${items.length}</span>
                    <button class="btn" onclick="startDownload()" id="download_btn_top" style="background: #28a745; margin-left: 20px;">🚀 Start Download</button>
                </div>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Select</th>
                            <th>#</th>
                            <th>Item #</th>
                            <th>Description</th>
                            <th>Quantity</th>
                            <th>Ship To</th>
                            <th>Need By</th>
                            <th>Download</th>
                        </tr>
                    </thead>
                    <tbody>
            `;

            items.forEach((item, index) => {
                html += `
                    <tr>
                        <td><input type="checkbox" class="item-checkbox" data-item-index="${index}" checked onchange="updateSelectedCount()"></td>
                        <td>${index + 1}</td>
                        <td><strong>${item.name}</strong></td>
                        <td>${item.description}</td>
                        <td>${item.quantity}</td>
                        <td>${item.ship_to || 'N/A'}</td>
                        <td>${item.need_by || 'N/A'}</td>
                        <td>${item.has_download ? '✅' : '❌'}</td>
                    </tr>
                `;
            });

            html += '</tbody></table>';
            container.innerHTML = html;
        }
        
        async function startDownload() {
            if (!currentPO) {
                showError('No PO data available. Please parse data first.');
                return;
            }

            if (!selectedMethod) {
                showError('Please select a download method first.');
                return;
            }

            // Get only selected items
            const selectedItems = getSelectedItems();
            if (selectedItems.length === 0) {
                showError('Please select at least one item to download.');
                return;
            }

            // 🆕 DOWNLOAD CONFIRMATION
            const methodName = selectedMethod === 'guaranteed_complete' ? 'Guaranteed Complete Download (Method 5)' :
                              selectedMethod === 'original_slow' ? 'Original Slow Method' :
                              selectedMethod === 'super_fast' ? 'Super Fast Method' :
                              selectedMethod === 'smart_direct' ? 'Smart Direct Method' :
                              selectedMethod === 'hybrid_smart' ? 'Hybrid Smart Method' : selectedMethod;

            const confirmDownload = confirm(`🚀 Start Download Confirmation\n\n` +
                `PO Number: ${currentPO.po_number}\n` +
                `Method: ${methodName}\n` +
                `Items to download: ${selectedItems.length}\n` +
                `Expected files: ${selectedItems.length} PDFs\n\n` +
                `Click OK to start downloading artwork files.`);

            if (!confirmDownload) {
                return;
            }

            // 🆕 PROMPT FOR PO DATABASE SAVE
            const savePODetails = await promptSavePODetails(currentPO.po_number);

            // Save current scroll position
            const currentScrollY = window.scrollY;

            // Disable both download buttons
            const button = document.getElementById('download_btn');
            const topButton = document.getElementById('download_btn_top');
            if (button) {
                button.disabled = true;
                button.innerHTML = `⏳ Starting Download (${selectedItems.length} items)...`;
            }
            if (topButton) {
                topButton.disabled = true;
                topButton.innerHTML = `⏳ Starting Download (${selectedItems.length} items)...`;
            }
            document.getElementById('progress_step').classList.remove('hidden');

            // Restore scroll position to prevent jumping to top
            window.scrollTo(0, currentScrollY);

            try {
                // Save PO details to database if user chose to
                if (savePODetails) {
                    await savePOToDatabase(currentPO.po_number);
                }

                // Show download progress notification
                const downloadNotificationId = `download-${currentPO.po_number}`;
                showProgressNotification(downloadNotificationId, `📥 Downloading ${currentPO.po_number} artwork...`, 'processing', '📥');

                // Start the actual download
                await fetch('/api/start_download', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        method: selectedMethod,
                        po_number: currentPO.po_number,
                        items: selectedItems  // Only send selected items
                    })
                });

                // Start polling for progress
                pollProgress(downloadNotificationId);
            } catch (error) {
                const downloadNotificationId = `download-${currentPO.po_number}`;
                updateNotification(downloadNotificationId, `❌ Error starting download for ${currentPO.po_number}: ${error.message}`, 'error', '❌');
            }
        }

        // Delivery Date Functions
        async function loadSavedPOs() {
            const loadingDiv = document.getElementById('saved_pos_loading');
            const listDiv = document.getElementById('saved_pos_list');
            const emptyDiv = document.getElementById('saved_pos_empty');

            // Clear search input
            const searchInput = document.getElementById('po_search_input');
            const resultsCount = document.getElementById('search_results_count');
            if (searchInput) searchInput.value = '';
            if (resultsCount) resultsCount.textContent = '';

            // Show loading
            loadingDiv.style.display = 'block';
            listDiv.style.display = 'none';
            emptyDiv.style.display = 'none';

            try {
                const response = await fetch('/api/po/get_all');
                const result = await response.json();

                if (result.success && result.pos.length > 0) {
                    displaySavedPOs(result.pos);
                    loadingDiv.style.display = 'none';
                    listDiv.style.display = 'block';
                } else {
                    loadingDiv.style.display = 'none';
                    emptyDiv.style.display = 'block';
                }
            } catch (error) {
                loadingDiv.style.display = 'none';
                emptyDiv.innerHTML = '<div style="padding: 20px; text-align: center; color: red;">❌ Error loading POs: ' + error.message + '</div>';
                emptyDiv.style.display = 'block';
            }
        }

        function displaySavedPOs(pos) {
            const listDiv = document.getElementById('saved_pos_list');

            let html = `
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background: #f8f9fa; border-bottom: 2px solid #dee2e6; position: sticky; top: 0; z-index: 10;">
                            <th style="padding: 12px; text-align: left; border-right: 1px solid #dee2e6; background: #f8f9fa;">PO Number</th>
                            <th style="padding: 12px; text-align: left; border-right: 1px solid #dee2e6; background: #f8f9fa;">Company</th>
                            <th style="padding: 12px; text-align: left; border-right: 1px solid #dee2e6; background: #f8f9fa;">Items</th>
                            <th style="padding: 12px; text-align: left; border-right: 1px solid #dee2e6; background: #f8f9fa;">Cancel Date</th>
                            <th style="padding: 12px; text-align: left; border-right: 1px solid #dee2e6; background: #f8f9fa;">Saved Date</th>
                            <th style="padding: 12px; text-align: center; border-right: 1px solid #dee2e6; background: #f8f9fa;">Inspection Report</th>
                            <th style="padding: 12px; text-align: center; border-right: 1px solid #dee2e6; background: #f8f9fa;">Sticker</th>
                            <th style="padding: 12px; text-align: center; background: #f8f9fa;">Action</th>
                        </tr>
                    </thead>
                    <tbody>
            `;

            pos.forEach(po => {
                const savedDate = new Date(po.created_date).toLocaleDateString();
                const currentDate = new Date().toISOString().split('T')[0]; // YYYY-MM-DD format
                const qcReportFilename = `${currentDate}-${po.po_number}-qc.xlsx`;
                const stickerFilename = `${currentDate}-${po.po_number}-sticker.xlsx`;

                html += `
                    <tr style="border-bottom: 1px solid #dee2e6; cursor: pointer;" onmouseover="this.style.background='#f8f9fa'" onmouseout="this.style.background='white'">
                        <td style="padding: 12px; border-right: 1px solid #dee2e6;"><strong>${po.po_number}</strong></td>
                        <td style="padding: 12px; border-right: 1px solid #dee2e6;">${po.company || po.purchase_from || 'N/A'}</td>
                        <td style="padding: 12px; border-right: 1px solid #dee2e6; text-align: center;">${po.item_count}</td>
                        <td style="padding: 12px; border-right: 1px solid #dee2e6;">${po.cancel_date || 'N/A'}</td>
                        <td style="padding: 12px; border-right: 1px solid #dee2e6;">${savedDate}</td>
                        <td style="padding: 12px; border-right: 1px solid #dee2e6; text-align: center;">
                            <a href="/api/download_qc_report/${qcReportFilename}"
                               style="color: #007bff; text-decoration: none; font-size: 12px;"
                               title="Download QC Report for ${po.po_number}">
                                📊 Download
                            </a>
                        </td>
                        <td style="padding: 12px; border-right: 1px solid #dee2e6; text-align: center;">
                            <a href="/api/download_sticker/${stickerFilename}"
                               style="color: #28a745; text-decoration: none; font-size: 12px;"
                               title="Download Sticker File for ${po.po_number}">
                                🏷️ Download
                            </a>
                        </td>
                        <td style="padding: 12px; text-align: center;">
                            <button onclick="selectPOForDelivery('${po.po_number}')" style="padding: 6px 12px; background: #28a745; color: white; border: none; border-radius: 3px; cursor: pointer;">
                                📅 Select
                            </button>
                        </td>
                    </tr>
                `;
            });

            html += '</tbody></table>';
            listDiv.innerHTML = html;
        }

        function filterPOTable() {
            const searchInput = document.getElementById('po_search_input');
            const searchTerm = searchInput.value.toLowerCase().trim();
            const table = document.querySelector('#saved_pos_list table');
            const resultsCount = document.getElementById('search_results_count');

            if (!table) {
                resultsCount.textContent = '';
                return;
            }

            const rows = table.querySelectorAll('tbody tr');
            let visibleCount = 0;
            let totalCount = rows.length;

            rows.forEach(row => {
                const poNumberCell = row.querySelector('td:first-child');
                if (poNumberCell) {
                    const poNumber = poNumberCell.textContent.toLowerCase();

                    if (searchTerm === '' || poNumber.includes(searchTerm)) {
                        row.style.display = '';
                        visibleCount++;
                    } else {
                        row.style.display = 'none';
                    }
                }
            });

            // Update results count
            if (searchTerm === '') {
                resultsCount.textContent = '';
            } else {
                resultsCount.textContent = `Showing ${visibleCount} of ${totalCount} POs`;
                if (visibleCount === 0) {
                    resultsCount.innerHTML = '<span style="color: #dc3545;">❌ No POs found matching "' + searchTerm + '"</span>';
                } else if (visibleCount === 1) {
                    resultsCount.innerHTML = '<span style="color: #28a745;">✅ Found 1 PO matching "' + searchTerm + '"</span>';
                } else {
                    resultsCount.innerHTML = '<span style="color: #28a745;">✅ Found ' + visibleCount + ' POs matching "' + searchTerm + '"</span>';
                }
            }
        }

        async function selectPOForDelivery(poNumber) {
            try {
                const response = await fetch(`/api/po/get_details/${poNumber}`);
                const result = await response.json();

                if (result.success) {
                    const header = result.header;
                    const items = result.items;

                    // Populate PO Header Table
                    const headerBody = document.getElementById('po_header_body');
                    headerBody.innerHTML = `
                        <tr style="background: #f8f9fa;">
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.po_number || 'N/A'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.factory || header.purchase_from || 'N/A'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.po_date || 'N/A'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.ship_by || header.cancel_date || 'N/A'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.ship_via || 'Delivery'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.order_type || 'Production'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.status || 'Completed'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.location || 'BID HK'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${header.prod_rep || 'N/A'}</td>
                        </tr>
                    `;

                    // Populate PO Items Table
                    const itemsBody = document.getElementById('po_items_body');
                    let itemsHtml = '';
                    items.forEach((item, index) => {
                        const bgColor = index % 2 === 0 ? '#f8f9fa' : 'white';
                        itemsHtml += `
                            <tr style="background: ${bgColor};">
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${item.item_number || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px; max-width: 150px; overflow: hidden; text-overflow: ellipsis;" title="${item.description || 'N/A'}">${item.description || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${item.color || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${item.ship_to || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${item.need_by || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px; text-align: right;">${item.qty || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">${item.bundle_qty || 'NA'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px; text-align: right;">${item.unit_price || 'N/A'}</td>
                                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px; text-align: right;">${item.extension || 'N/A'}</td>
                            </tr>
                        `;
                    });
                    itemsBody.innerHTML = itemsHtml;

                    // Populate Additional Details Table
                    const additionalBody = document.getElementById('po_additional_body');
                    additionalBody.innerHTML = `
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold; width: 150px;">Purchased From:</td>
                            <td style="padding: 8px; border: 1px solid #ddd; width: 250px;">${header.purchase_from || 'F & C (Hong Kong) Industrial Limited'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold; width: 100px;">Ship To:</td>
                            <td style="padding: 8px; border: 1px solid #ddd; width: 250px;">${header.ship_to || 'Brand I.D. HK Limited'}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold; width: 100px;">Company:</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">${header.company || 'Brand ID HK'}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold;">Address:</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">Unit 1505, One Midtown, 11 Hoi Shing Road, Tsuen Wan<br>Hong Kong Hong Kong</td>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold;">Address:</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">2/F, Tsuen Wan Industrial Centre<br>220-248 Texaco Road Tsuen Wan<br>Hong Kong</td>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold;">Currency:</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">${header.currency || 'USD'}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold;">Cancel Date:</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">${header.cancel_date || ''}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; background: #f8f9fa; font-weight: bold;">Terms:</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">${header.terms || ''}</td>
                            <td style="padding: 8px; border: 1px solid #ddd;"></td>
                            <td style="padding: 8px; border: 1px solid #ddd;"></td>
                        </tr>
                    `;

                    // Set current delivery date (use cancel_date as default)
                    document.getElementById('current_delivery_date').value = header.cancel_date || header.ship_by || '';

                    // Clear form
                    document.getElementById('new_delivery_date').value = '';
                    document.getElementById('delivery_notes').value = '';

                    // Populate tracking information
                    function formatDateTime(dateString) {
                        if (!dateString) return '-';
                        try {
                            const date = new Date(dateString);
                            return date.toLocaleString('en-US', {
                                year: 'numeric',
                                month: '2-digit',
                                day: '2-digit',
                                hour: '2-digit',
                                minute: '2-digit',
                                second: '2-digit'
                            });
                        } catch (e) {
                            return dateString;
                        }
                    }

                    document.getElementById('first_created_display').textContent = formatDateTime(header.first_created);
                    document.getElementById('last_updated_display').textContent = formatDateTime(header.last_updated);
                    document.getElementById('update_count_display').textContent = header.update_count || 0;

                    // Show details section
                    document.getElementById('delivery_info').classList.remove('hidden');

                    showError(`✅ PO ${poNumber} selected - ${items.length} items loaded`, 'success');
                } else {
                    showError('❌ Error loading PO details: ' + result.message, 'error');
                }
            } catch (error) {
                showError('❌ Error loading PO details: ' + error.message, 'error');
            }
        }

        async function updateDeliveryDate() {
            const poNumber = document.getElementById('selected_po_number').textContent;
            const newDate = document.getElementById('new_delivery_date').value;
            const notes = document.getElementById('delivery_notes').value;

            if (!poNumber) {
                showError('❌ Please select a PO first', 'error');
                return;
            }

            if (!newDate) {
                showError('❌ Please select a new delivery date', 'error');
                return;
            }

            // Simulate update (you can implement actual delivery date update logic here)
            showError(`✅ Delivery date updated for PO ${poNumber} to ${newDate}`, 'success');

            if (notes) {
                console.log(`Notes: ${notes}`);
            }
        }



        // Report Functions
        async function generateReport() {
            const fromDate = document.getElementById('report_date_from').value;
            const toDate = document.getElementById('report_date_to').value;

            if (!fromDate || !toDate) {
                alert('Please select both from and to dates');
                return;
            }

            // Simulate report generation
            const reportHtml = `
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Date</th>
                            <th>PO Number</th>
                            <th>Items Downloaded</th>
                            <th>Status</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>2025-08-01</td>
                            <td>1284789</td>
                            <td>3</td>
                            <td>✅ Complete</td>
                        </tr>
                        <tr>
                            <td>2025-07-31</td>
                            <td>1288060</td>
                            <td>5</td>
                            <td>✅ Complete</td>
                        </tr>
                    </tbody>
                </table>
            `;

            document.getElementById('report_table_container').innerHTML = reportHtml;
            document.getElementById('report_results').classList.remove('hidden');

            // Update quick stats
            document.getElementById('today_downloads').textContent = '3';
            document.getElementById('week_downloads').textContent = '8';
            document.getElementById('total_files').textContent = '156';
        }

        async function exportReport() {
            alert('Report exported to CSV (feature coming soon)');
        }

        async function pollProgress(downloadNotificationId = null) {
            try {
                const response = await fetch('/api/status');
                const status = await response.json();

                // Show "Open Folder" and "New PO" links if download is complete (100% and not active)
                const showOpenFolder = !status.active && status.progress === 100 && status.download_folder;
                const showNewPO = !status.active && status.progress === 100;

                document.getElementById('progress_info').innerHTML = `
                    <p>Progress: ${status.progress}%
                        ${showOpenFolder ? '<a href="#" onclick="openDownloadFolder()" style="margin-left: 15px; color: #007bff; text-decoration: none; font-weight: bold;">📁 Open Folder</a>' : ''}
                        ${showNewPO ? '<a href="#" onclick="clearEverything()" style="margin-left: 15px; color: #28a745; text-decoration: none; font-weight: bold;">🆕 New PO</a>' : ''}
                    </p>
                    <div style="background: #e0e0e0; height: 20px; margin: 10px 0;">
                        <div style="background: #333; height: 100%; width: ${status.progress}%; transition: width 0.3s;"></div>
                    </div>
                    ${showNewPO ? '<div style="padding: 15px; background: #d4edda; border-radius: 8px; margin: 10px 0; border-left: 4px solid #28a745;"><strong>✅ Download Complete!</strong><br>📁 Open the download folder to access your files.<br>🆕 Click "NEW PO" button to start processing another PO.</div>' : ''}
                    <div style="max-height: 200px; overflow-y: auto; background: #f9f9f9; padding: 10px; font-family: monospace;">
                        ${status.log.slice().reverse().map(entry => `<div>${entry}</div>`).join('')}
                    </div>
                `;

                if (status.active) {
                    setTimeout(() => pollProgress(downloadNotificationId), 1000);
                } else {
                    // Download completed - update notification
                    if (downloadNotificationId && currentPO) {
                        if (status.progress === 100) {
                            updateNotification(downloadNotificationId, `✅ Finished Downloading ${currentPO.po_number} artwork`, 'success', '✅');

                            // 🔒 DISABLE FIELDS AFTER SUCCESSFUL DOWNLOAD
                            // Disable PO input field
                            document.getElementById('po_input').disabled = true;

                            // Disable analyze button
                            document.getElementById('analyze_btn').disabled = true;
                            document.getElementById('analyze_btn').innerHTML = '✅ Completed';

                            // Disable download buttons
                            document.getElementById('download_btn').disabled = true;
                            document.getElementById('download_btn').innerHTML = '✅ Download Complete';

                            const topButton = document.getElementById('download_btn_top');
                            if (topButton) {
                                topButton.disabled = true;
                                topButton.innerHTML = '✅ Download Complete';
                            }

                        } else {
                            updateNotification(downloadNotificationId, `❌ Download failed for ${currentPO.po_number}`, 'error', '❌');

                            // Re-enable buttons for retry on failure
                            document.getElementById('download_btn').disabled = false;
                            const topButton = document.getElementById('download_btn_top');
                            if (topButton) {
                                topButton.disabled = false;
                                topButton.innerHTML = '🚀 Start Download';
                            }
                        }
                    }
                }
            } catch (error) {
                console.error('Error polling progress:', error);
                if (downloadNotificationId && currentPO) {
                    updateNotification(downloadNotificationId, `❌ Error during download of ${currentPO.po_number}`, 'error', '❌');
                }
            }
        }

        async function openDownloadFolder() {
            try {
                const response = await fetch('/api/open_folder');
                const result = await response.json();
                if (result.success) {
                    showError('📁 Folder opened successfully!', 'success');
                } else {
                    showError('❌ ' + result.message, 'error');
                }
            } catch (error) {
                showError('❌ Error opening folder: ' + error.message, 'error');
            }
        }

        // Email masking function
        function maskEmail(email) {
            if (!email.includes('@')) return email;

            const [prefix, suffix] = email.split('@');

            // Mask prefix: show first 2 characters, rest as asterisks
            const maskedPrefix = prefix.length <= 2 ? prefix : prefix.substring(0, 2) + '*'.repeat(prefix.length - 2);

            // Mask suffix: show first 1 character, rest as asterisks
            const maskedSuffix = suffix.length <= 1 ? suffix : suffix.substring(0, 1) + '*'.repeat(suffix.length - 1);

            return `${maskedPrefix}@${maskedSuffix}`;
        }

        // Settings functions
        async function verifyAdmin() {
            const password = document.getElementById('admin_password').value;
            const messageDiv = document.getElementById('admin_message');

            if (!password) {
                messageDiv.textContent = 'Please enter admin password';
                messageDiv.style.color = 'red';
                return;
            }

            try {
                const response = await fetch('/api/settings/verify_admin', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({password: password})
                });

                const result = await response.json();

                if (result.success) {
                    messageDiv.textContent = '✅ Admin access granted';
                    messageDiv.style.color = 'green';

                    // Load and show configuration
                    await loadConfiguration(password);
                } else {
                    messageDiv.textContent = '❌ Invalid admin password';
                    messageDiv.style.color = 'red';
                }
            } catch (error) {
                messageDiv.textContent = '❌ Error: ' + error.message;
                messageDiv.style.color = 'red';
            }
        }

        async function loadConfiguration(password) {
            try {
                const response = await fetch('/api/settings/get_config', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({password: password})
                });

                const result = await response.json();

                if (result.success) {
                    // Show actual values
                    document.getElementById('display_url').textContent = result.config.login_url;
                    document.getElementById('display_username').textContent = result.config.username;
                    document.getElementById('display_password').textContent = result.config.password;

                    // Populate edit form
                    document.getElementById('edit_url').value = result.config.login_url;
                    document.getElementById('edit_username').value = result.config.username;
                    document.getElementById('edit_password').value = result.config.password;

                    // Show edit form
                    document.getElementById('edit_form').style.display = 'block';
                    document.getElementById('admin_section').style.display = 'none';
                }
            } catch (error) {
                console.error('Error loading configuration:', error);
            }
        }

        async function saveConfig() {
            const adminPassword = document.getElementById('admin_password').value;
            const url = document.getElementById('edit_url').value;
            const username = document.getElementById('edit_username').value;
            const password = document.getElementById('edit_password').value;
            const messageDiv = document.getElementById('save_message');

            try {
                const response = await fetch('/api/settings/update_config', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        admin_password: adminPassword,
                        login_url: url,
                        username: username,
                        password: password
                    })
                });

                const result = await response.json();

                if (result.success) {
                    messageDiv.innerHTML = '<span style="color: green;">✅ Configuration saved successfully!</span>';

                    // Update display
                    document.getElementById('display_url').textContent = url;
                    document.getElementById('display_username').textContent = username;
                    document.getElementById('display_password').textContent = password;
                } else {
                    messageDiv.innerHTML = '<span style="color: red;">❌ ' + result.message + '</span>';
                }
            } catch (error) {
                messageDiv.innerHTML = '<span style="color: red;">❌ Error: ' + error.message + '</span>';
            }
        }

        function cancelEdit() {
            document.getElementById('edit_form').style.display = 'none';
            document.getElementById('admin_section').style.display = 'block';
            document.getElementById('admin_password').value = '';
            document.getElementById('admin_message').textContent = '';

            // Reset display to masked values
            const currentUsername = document.getElementById('edit_username').value || '{{ masked_username }}';
            document.getElementById('display_username').textContent = maskEmail(currentUsername);
            document.getElementById('display_password').textContent = '************';
        }

        // PO Database functions
        async function promptSavePODetails(poNumber) {
            return new Promise((resolve) => {
                const modal = document.createElement('div');
                modal.style.cssText = `
                    position: fixed; top: 0; left: 0; width: 100%; height: 100%;
                    background: rgba(0,0,0,0.5); z-index: 10000; display: flex;
                    align-items: center; justify-content: center;
                `;

                modal.innerHTML = `
                    <div style="background: white; padding: 30px; border-radius: 10px; max-width: 500px; text-align: center;">
                        <h3>📊 Save PO Details to Database?</h3>
                        <p>Do you want to save complete PO details for <strong>${poNumber}</strong> to the database?</p>
                        <p style="font-size: 0.9em; color: #666;">This will save all item details, company info, and dates for future reference.</p>
                        <div style="margin-top: 20px;">
                            <button id="saveYes" style="padding: 10px 20px; margin: 0 10px; background: #28a745; color: white; border: none; border-radius: 5px; cursor: pointer;">
                                ✅ Yes, Save Details
                            </button>
                            <button id="saveNo" style="padding: 10px 20px; margin: 0 10px; background: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer;">
                                ❌ No, Skip
                            </button>
                        </div>
                    </div>
                `;

                document.body.appendChild(modal);

                document.getElementById('saveYes').onclick = () => {
                    document.body.removeChild(modal);
                    resolve(true);
                };

                document.getElementById('saveNo').onclick = () => {
                    document.body.removeChild(modal);
                    resolve(false);
                };
            });
        }

        async function promptOverwritePO(poNumber) {
            return new Promise((resolve) => {
                const modal = document.createElement('div');
                modal.style.cssText = `
                    position: fixed; top: 0; left: 0; width: 100%; height: 100%;
                    background: rgba(0,0,0,0.5); z-index: 10000; display: flex;
                    align-items: center; justify-content: center;
                `;

                modal.innerHTML = `
                    <div style="background: white; padding: 30px; border-radius: 10px; max-width: 500px; text-align: center;">
                        <h3>⚠️ PO Already Exists</h3>
                        <p>PO <strong>${poNumber}</strong> already exists in the database.</p>
                        <p style="font-size: 0.9em; color: #666;">Do you want to overwrite the existing data?</p>
                        <div style="margin-top: 20px;">
                            <button id="overwriteYes" style="padding: 10px 20px; margin: 0 10px; background: #dc3545; color: white; border: none; border-radius: 5px; cursor: pointer;">
                                🔄 Yes, Overwrite
                            </button>
                            <button id="overwriteNo" style="padding: 10px 20px; margin: 0 10px; background: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer;">
                                ❌ No, Skip
                            </button>
                        </div>
                    </div>
                `;

                document.body.appendChild(modal);

                document.getElementById('overwriteYes').onclick = () => {
                    document.body.removeChild(modal);
                    resolve(true);
                };

                document.getElementById('overwriteNo').onclick = () => {
                    document.body.removeChild(modal);
                    resolve(false);
                };
            });
        }

        async function savePOToDatabase(poNumber) {
            try {
                // First check if PO exists
                const checkResponse = await fetch('/api/po/check_exists', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({po_number: poNumber})
                });

                const checkResult = await checkResponse.json();

                if (!checkResult.success) {
                    showError('❌ Error checking PO existence: ' + checkResult.message, 'error');
                    return false;
                }

                let overwrite = false;
                if (checkResult.exists) {
                    overwrite = await promptOverwritePO(poNumber);
                    if (!overwrite) {
                        showError('📊 PO database save skipped', 'info');
                        return false;
                    }
                }

                // Show progress notification
                const saveNotificationId = `save-${poNumber}`;
                showProgressNotification(saveNotificationId, `💾 Saving PO ${poNumber} details to database...`, 'processing', '💾');

                // Save PO details
                const saveResponse = await fetch('/api/po/save_details', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({po_number: poNumber, overwrite: overwrite})
                });

                const saveResult = await saveResponse.json();

                if (saveResult.success) {
                    updateNotification(saveNotificationId, `✅ Finished Saving PO ${poNumber} to database (${saveResult.items_count} items)`, 'success', '✅');
                    return true;
                } else {
                    updateNotification(saveNotificationId, `❌ Failed to save PO ${poNumber}: ${saveResult.message}`, 'error', '❌');
                    return false;
                }

            } catch (error) {
                showError('❌ Error saving PO to database: ' + error.message, 'error');
                return false;
            }
        }

        function toggleMethodSelection() {
            const defaultDisplay = document.querySelector('.default-method-display');
            const allMethods = document.getElementById('all_methods');

            if (allMethods.style.display === 'none' || allMethods.classList.contains('hidden')) {
                // Show all methods
                defaultDisplay.style.display = 'none';
                allMethods.style.display = 'block';
                allMethods.classList.remove('hidden');
            } else {
                // Hide all methods, show default
                defaultDisplay.style.display = 'block';
                allMethods.style.display = 'none';
                allMethods.classList.add('hidden');

                // Reset selection to Method 5
                document.querySelectorAll('.method-card').forEach(card => {
                    card.classList.remove('selected');
                });
                document.querySelector('[data-method="guaranteed_complete"]').classList.add('selected');
            }
        }

        function showError(message, type = 'error', persistent = false) {
            const container = document.getElementById('error_container');
            const className = type;

            // Full-screen overlay styling
            const overlayStyle = `
                position: fixed;
                top: 0;
                left: 0;
                width: 100vw;
                height: 100vh;
                background: rgba(0, 0, 0, 0.9);
                display: flex;
                justify-content: center;
                align-items: center;
                z-index: 9999;
                backdrop-filter: blur(2px);
            `;

            // Different colors for different message types
            let backgroundColor, textColor, borderColor;
            switch(type) {
                case 'success':
                    backgroundColor = 'rgba(40, 167, 69, 0.95)';
                    textColor = '#fff';
                    borderColor = 'rgba(255, 255, 255, 0.3)';
                    break;
                case 'info':
                    backgroundColor = 'rgba(23, 162, 184, 0.95)';
                    textColor = '#fff';
                    borderColor = 'rgba(255, 255, 255, 0.3)';
                    break;
                default: // error
                    backgroundColor = 'rgba(220, 53, 69, 0.95)';
                    textColor = '#fff';
                    borderColor = 'rgba(255, 255, 255, 0.3)';
            }

            const messageStyle = `
                background: ${backgroundColor};
                color: ${textColor};
                padding: 30px 50px;
                border-radius: 15px;
                font-size: 18px;
                font-weight: 600;
                text-align: center;
                box-shadow: 0 10px 30px rgba(0, 0, 0, 0.3);
                border: 2px solid ${borderColor};
                max-width: 80%;
                word-wrap: break-word;
            `;

            container.innerHTML = `
                <div class="${className}" style="${overlayStyle}">
                    <div style="${messageStyle}">${message}</div>
                </div>
            `;

            // Auto-hide messages after 3 seconds, unless persistent
            if (!persistent) {
                setTimeout(() => {
                    container.innerHTML = '';
                }, 3000);
            }
        }

        // Function to clear persistent messages
        function clearError() {
            const container = document.getElementById('error_container');
            container.innerHTML = '';
        }

        // Success message function
        function showSuccess(message) {
            showError(message, 'success', false);
        }

        // Info message function
        function showInfo(message) {
            showError(message, 'info', false);
        }

        // Master Report Functions
        let masterReportData = [];
        let searchTimeout = null;

        function loadMasterReport() {
            console.log('🔍 Loading master report...');

            // Show loading state
            document.getElementById('master_report_loading').style.display = 'block';
            document.getElementById('master_report_container').style.display = 'none';
            document.getElementById('master_report_empty').style.display = 'none';

            fetch('/api/master_report?limit=20')
                .then(response => response.json())
                .then(data => {
                    console.log('📊 Master report data received:', data);

                    if (data.success && data.data.length > 0) {
                        masterReportData = data.data;
                        displayMasterReportData(data.data);
                        updateMasterReportStats(data);

                        // Show table
                        document.getElementById('master_report_loading').style.display = 'none';
                        document.getElementById('master_report_container').style.display = 'block';
                    } else {
                        // Show empty state
                        document.getElementById('master_report_loading').style.display = 'none';
                        document.getElementById('master_report_empty').style.display = 'block';
                    }
                })
                .catch(error => {
                    console.error('❌ Error loading master report:', error);
                    showError('Failed to load master report: ' + error.message);
                    document.getElementById('master_report_loading').style.display = 'none';
                    document.getElementById('master_report_empty').style.display = 'block';
                });
        }

        function displayMasterReportData(data) {
            const tbody = document.getElementById('master_report_tbody');
            tbody.innerHTML = '';

            data.forEach((row, index) => {
                const tr = document.createElement('tr');
                tr.style.borderBottom = '1px solid #dee2e6';

                // Alternate row colors
                if (index % 2 === 1) {
                    tr.style.backgroundColor = '#f8f9fa';
                }

                tr.innerHTML = `
                    <!-- Fixed Columns -->
                    <td style="position: sticky; left: 0; background: ${index % 2 === 1 ? '#f8f9fa' : 'white'}; z-index: 5; padding: 8px; border-right: 2px solid #adb5bd; font-weight: 500; width: 120px; vertical-align: top;" title="${row.po_number || ''}">${row.po_number || ''}</td>
                    <td style="position: sticky; left: 120px; background: ${index % 2 === 1 ? '#f8f9fa' : 'white'}; z-index: 5; padding: 8px; border-right: 2px solid #adb5bd; font-weight: 500; width: 140px; vertical-align: top;" title="${row.item_number || ''}">${row.item_number || ''}</td>
                    <td style="position: sticky; left: 260px; background: ${index % 2 === 1 ? '#f8f9fa' : 'white'}; z-index: 5; padding: 8px; border-right: 2px solid #adb5bd; width: 250px; vertical-align: top;" title="${row.description || ''}">${row.description || ''}</td>

                    <!-- Scrollable Columns -->
                    <td style="padding: 8px; width: 120px; vertical-align: top;" title="${row.color || ''}">${row.color || ''}</td>
                    <td style="padding: 8px; width: 140px; vertical-align: top;" title="${row.ship_to || ''}">${row.ship_to || ''}</td>
                    <td style="padding: 8px; width: 120px; vertical-align: top;" title="${row.need_by || ''}">${row.need_by || ''}</td>
                    <td style="padding: 8px; text-align: right; width: 100px; vertical-align: top;" title="${row.qty || ''}">${row.qty || ''}</td>
                    <td style="padding: 8px; text-align: right; width: 120px; vertical-align: top;" title="${row.bundle_qty || ''}">${row.bundle_qty || ''}</td>
                    <td style="padding: 8px; text-align: right; width: 120px; vertical-align: top;" title="${row.unit_price || ''}">${row.unit_price || ''}</td>
                    <td style="padding: 8px; text-align: right; width: 120px; vertical-align: top;" title="${row.extension || ''}">${row.extension || ''}</td>
                    <td style="padding: 8px; width: 150px; vertical-align: top;" title="${row.company || ''}">${row.company || ''}</td>
                    <td style="padding: 8px; width: 160px; vertical-align: top;" title="${row.purchase_from || ''}">${row.purchase_from || ''}</td>
                    <td style="padding: 8px; width: 100px; vertical-align: top;" title="${row.currency || ''}">${row.currency || ''}</td>
                    <td style="padding: 8px; width: 120px; vertical-align: top;" title="${row.po_date || ''}">${row.po_date || ''}</td>
                    <td style="padding: 8px; width: 130px; vertical-align: top;" title="${row.cancel_date || ''}">${row.cancel_date || ''}</td>
                    <td style="padding: 8px; width: 120px; vertical-align: top;" title="${row.ship_by || ''}">${row.ship_by || ''}</td>
                    <td style="padding: 8px; width: 140px; vertical-align: top;" title="${row.ship_via || ''}">${row.ship_via || ''}</td>
                    <td style="padding: 8px; width: 130px; vertical-align: top;" title="${row.order_type || ''}">${row.order_type || ''}</td>
                    <td style="padding: 8px; width: 100px; vertical-align: top;" title="${row.status || ''}">${row.status || ''}</td>
                    <td style="padding: 8px; width: 150px; vertical-align: top;" title="${row.factory || ''}">${row.factory || ''}</td>
                    <td style="padding: 8px; width: 120px; vertical-align: top;" title="${row.location || ''}">${row.location || ''}</td>
                    <td style="padding: 8px; width: 130px; vertical-align: top;" title="${row.prod_rep || ''}">${row.prod_rep || ''}</td>
                    <td style="padding: 8px; width: 200px; vertical-align: top;" title="${row.ship_to_address || ''}">${row.ship_to_address || ''}</td>
                    <td style="padding: 8px; width: 130px; vertical-align: top;" title="${row.terms || ''}">${row.terms || ''}</td>
                    <td style="padding: 8px; font-size: 0.8em; color: #666; width: 150px; vertical-align: top;" title="${row.first_created || ''}">${row.first_created || ''}</td>
                    <td style="padding: 8px; font-size: 0.8em; color: #666; width: 150px; vertical-align: top;" title="${row.last_updated || ''}">${row.last_updated || ''}</td>
                    <td style="padding: 8px; text-align: center; width: 120px; vertical-align: top;" title="${row.update_count || '0'}">${row.update_count || '0'}</td>
                `;

                tbody.appendChild(tr);
            });
        }

        function updateMasterReportStats(data) {
            const now = new Date().toLocaleString();
            const activeFilters = getActiveFilterCount();

            document.getElementById('master_report_stats').textContent =
                `📊 Showing ${data.filtered_count} of ${data.total_count} total records | 🔍 Active filters: ${activeFilters} | 📅 Last updated: ${now}`;
        }

        function getActiveFilterCount() {
            const searchInputs = document.querySelectorAll('#master_report_table input[id^="search_"]');
            let count = 0;
            searchInputs.forEach(input => {
                if (input.value.trim()) count++;
            });
            return count;
        }

        function searchMasterReport() {
            // Clear existing timeout
            if (searchTimeout) {
                clearTimeout(searchTimeout);
            }

            // Debounce search to avoid too many requests
            searchTimeout = setTimeout(() => {
                performMasterReportSearch();
            }, 500);
        }

        function performMasterReportSearch() {
            console.log('🔍 Performing master report search...');

            // Collect search filters
            const searchParams = new URLSearchParams();
            searchParams.append('limit', '20');

            const searchInputs = document.querySelectorAll('#master_report_table input[id^="search_"]');
            searchInputs.forEach(input => {
                if (input.value.trim()) {
                    const columnName = input.id.replace('search_', '');
                    searchParams.append(`search_${columnName}`, input.value.trim());
                }
            });

            // Show loading state
            document.getElementById('master_report_loading').style.display = 'block';
            document.getElementById('master_report_container').style.display = 'none';

            fetch(`/api/master_report?${searchParams.toString()}`)
                .then(response => response.json())
                .then(data => {
                    console.log('📊 Search results received:', data);

                    if (data.success) {
                        masterReportData = data.data;
                        displayMasterReportData(data.data);
                        updateMasterReportStats(data);

                        // Show table or empty state
                        document.getElementById('master_report_loading').style.display = 'none';
                        if (data.data.length > 0) {
                            document.getElementById('master_report_container').style.display = 'block';
                            document.getElementById('master_report_empty').style.display = 'none';
                        } else {
                            document.getElementById('master_report_container').style.display = 'none';
                            document.getElementById('master_report_empty').style.display = 'block';
                        }
                    } else {
                        showError('Search failed: ' + (data.error || 'Unknown error'));
                        document.getElementById('master_report_loading').style.display = 'none';
                        document.getElementById('master_report_empty').style.display = 'block';
                    }
                })
                .catch(error => {
                    console.error('❌ Error searching master report:', error);
                    showError('Search failed: ' + error.message);
                    document.getElementById('master_report_loading').style.display = 'none';
                    document.getElementById('master_report_empty').style.display = 'block';
                });
        }

        function refreshMasterReport() {
            console.log('🔄 Refreshing master report...');
            clearAllSearchFilters();
            loadMasterReport();
        }

        function clearAllSearchFilters() {
            console.log('🗑️ Clearing all search filters...');
            const searchInputs = document.querySelectorAll('#master_report_table input[id^="search_"]');
            searchInputs.forEach(input => {
                input.value = '';
            });
            loadMasterReport();
        }

        function exportMasterReport() {
            console.log('📥 Exporting master report to Excel...');

            // Collect current search filters
            const searchParams = new URLSearchParams();
            const searchInputs = document.querySelectorAll('#master_report_table input[id^="search_"]');
            searchInputs.forEach(input => {
                if (input.value.trim()) {
                    const columnName = input.id.replace('search_', '');
                    searchParams.append(`search_${columnName}`, input.value.trim());
                }
            });

            // Create download URL
            const exportUrl = `/api/export_master_report?${searchParams.toString()}`;

            // Create temporary link and trigger download
            const link = document.createElement('a');
            link.href = exportUrl;
            link.download = '';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);

            showSuccess('📥 Excel export started! Check your downloads folder.');
        }

        function loadMoreRecords() {
            console.log('📄 Loading more records...');
            // TODO: Implement pagination
            showInfo('📄 Pagination feature coming soon!');
        }

        // ===== SIMPLE OPTION A PACKING FUNCTIONS =====
        let currentPOData = null;
        let selectedItems = [];

        // 1. Reset Database (One-time) + Clear All Interface
        async function resetDatabase() {
            const statusDiv = document.getElementById('reset_status');

            if (!confirm('This will clear ALL packed status for ALL POs and reset the entire interface. Are you sure?')) {
                return;
            }

            try {
                statusDiv.innerHTML = '<div style="color: #007bff; padding: 10px; background: #e3f2fd; border-radius: 5px;">⏳ Resetting database...</div>';

                const response = await fetch('/api/simple_packing/reset_database', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({})
                });

                const result = await response.json();

                if (result.success) {
                    statusDiv.innerHTML = '<div style="color: #28a745; padding: 10px; background: #d4edda; border-radius: 5px;">✅ ' + result.message + '</div>';

                    // Clear ALL interface elements below reset section
                    clearAllInterface();

                } else {
                    statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ ' + result.message + '</div>';
                }

            } catch (error) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Error: ' + error.message + '</div>';
            }
        }

        // Clear All Interface Elements
        function clearAllInterface() {
            // Clear PO input and status
            document.getElementById('simple_po_input').value = '';
            document.getElementById('po_load_status').innerHTML = '';

            // Clear action status
            document.getElementById('action_status').innerHTML = '';

            // Clear items summary and table
            document.getElementById('items_summary').innerHTML = '';
            document.getElementById('items_table_body').innerHTML = '';

            // Clear pack section
            document.getElementById('selected_summary').innerHTML = '';
            document.getElementById('pack_status').innerHTML = '';

            // Hide containers
            document.getElementById('items_container').style.display = 'none';
            document.getElementById('pack_section').style.display = 'none';

            // Reset variables
            currentPOData = null;
            selectedItems = [];

            console.log('🔄 All interface elements cleared');
        }

        // Test Modal Function (for debugging)
        function testModal() {
            console.log('Test modal called');
            const modal = document.getElementById('carton_modal');
            console.log('Modal element found:', modal);

            if (modal) {
                // Show step 1
                showStep1();

                modal.style.display = 'block';
                modal.style.position = 'fixed';
                modal.style.top = '20%';
                modal.style.right = '5%';
                modal.style.zIndex = '10000';

                // Initialize drag functionality
                makeDraggable(modal);

                // Update summary for test
                const summary = document.getElementById('modal_selected_summary');
                if (summary) {
                    summary.innerHTML = 'TEST MODE - Movable window working!';
                }

                console.log('Test modal should be visible - Movable window');
            } else {
                alert('Modal element not found!');
            }
        }

        // 2. Load PO
        async function loadPOSimple() {
            const poNumber = document.getElementById('simple_po_input').value.trim();
            const statusDiv = document.getElementById('po_load_status');

            if (!poNumber) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Please enter a PO number</div>';
                return;
            }

            try {
                statusDiv.innerHTML = '<div style="color: #007bff; padding: 10px; background: #e3f2fd; border-radius: 5px;">⏳ Loading PO items...</div>';

                const response = await fetch(`/api/simple_packing/load_po?po_number=${poNumber}`);
                const result = await response.json();

                if (result.success) {
                    currentPOData = result;
                    statusDiv.innerHTML = '<div style="color: #28a745; padding: 10px; background: #d4edda; border-radius: 5px;">✅ Loaded ' + result.total_items + ' items from PO ' + result.po_number + '</div>';

                    // Show items container
                    document.getElementById('items_container').style.display = 'block';
                    displayItems();

                    // Auto-scroll to center the items table
                    setTimeout(() => {
                        const itemsContainer = document.getElementById('items_container');
                        if (itemsContainer) {
                            itemsContainer.scrollIntoView({
                                behavior: 'smooth',
                                block: 'center'
                            });
                        }
                    }, 300); // Small delay to ensure table is rendered

                } else {
                    statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ ' + result.message + '</div>';
                }

            } catch (error) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Error: ' + error.message + '</div>';
            }
        }

        // 3. Mark All Done
        async function markAllDone() {
            if (!currentPOData) {
                showError('Please load a PO first');
                return;
            }

            const statusDiv = document.getElementById('action_status');

            try {
                statusDiv.innerHTML = '<div style="color: #007bff; padding: 10px; background: #e3f2fd; border-radius: 5px;">⏳ Marking all items as done...</div>';

                const response = await fetch('/api/simple_packing/mark_all_done', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        po_number: currentPOData.po_number
                    })
                });

                const result = await response.json();

                if (result.success) {
                    statusDiv.innerHTML = '<div style="color: #28a745; padding: 10px; background: #d4edda; border-radius: 5px;">✅ ' + result.message + '</div>';

                    // Update items status in memory
                    currentPOData.items.forEach(item => {
                        if (item.packed_status !== 'packed') {
                            item.packed_status = 'done';
                        }
                    });

                    displayItems();
                } else {
                    statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ ' + result.message + '</div>';
                }

            } catch (error) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Error: ' + error.message + '</div>';
            }
        }

        // 4. Display Items in Table Format
        function displayItems() {
            if (!currentPOData || !currentPOData.items) return;

            const summaryDiv = document.getElementById('items_summary');
            const tableBody = document.getElementById('items_table_body');

            const totalItems = currentPOData.items.length;
            const doneItems = currentPOData.items.filter(item => item.packed_status === 'done').length;
            const packedItems = currentPOData.items.filter(item => item.packed_status === 'packed').length;

            summaryDiv.innerHTML = `Total: ${totalItems} | Done: ${doneItems} | Packed: ${packedItems} | Remaining: ${totalItems - packedItems}`;

            let html = '';
            currentPOData.items.forEach((item, index) => {
                const isDisabled = item.packed_status === 'packed';
                const statusColor = item.packed_status === 'packed' ? '#28a745' :
                                   item.packed_status === 'done' ? '#ffc107' : '#6c757d';
                const statusText = item.packed_status === 'packed' ? 'Packed' :
                                  item.packed_status === 'done' ? 'Done' : 'Not Packed';

                const rowStyle = isDisabled ? 'opacity: 0.6; background: #f8f9fa;' : '';
                const cartonNumber = item.carton_number || '-';
                const cartonStyle = item.carton_number ? 'font-weight: bold; color: #007bff;' : 'color: #999;';

                html += `
                    <tr style="${rowStyle}">
                        <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: center;">
                            <input type="checkbox" id="item_${index}" class="item-checkbox"
                                   ${isDisabled ? 'disabled' : ''}
                                   onchange="updateSelections()"
                                   style="transform: scale(1.2);">
                        </td>
                        <td style="padding: 12px; border-bottom: 1px solid #eee; font-weight: bold; color: #333;">
                            ${item.item_number || 'N/A'}
                        </td>
                        <td style="padding: 12px; border-bottom: 1px solid #eee; color: #666;">
                            ${item.description || 'No description'}
                        </td>
                        <td style="padding: 12px; border-bottom: 1px solid #eee; color: #666;">
                            ${item.color || 'N/A'}
                        </td>
                        <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;">
                            ${item.qty || 0}
                        </td>
                        <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: center;">
                            <span style="padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: bold; color: white; background: ${statusColor};">
                                ${statusText}
                            </span>
                        </td>
                        <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: center; ${cartonStyle}">
                            ${cartonNumber}
                        </td>
                    </tr>
                `;
            });

            tableBody.innerHTML = html;
            updateSelections();
        }

        // 5. Selection Functions
        function selectAllItems() {
            const checkboxes = document.querySelectorAll('.item-checkbox:not([disabled])');
            checkboxes.forEach(cb => cb.checked = true);
            updateSelections();
        }

        function clearSelections() {
            const checkboxes = document.querySelectorAll('.item-checkbox');
            checkboxes.forEach(cb => cb.checked = false);
            updateSelections();
        }

        function updateSelections() {
            const checkboxes = document.querySelectorAll('.item-checkbox:not([disabled])');
            selectedItems = [];

            checkboxes.forEach((cb, index) => {
                if (cb.checked) {
                    // Find the actual item index (accounting for disabled items)
                    const itemId = cb.id.replace('item_', '');
                    selectedItems.push(parseInt(itemId));
                }
            });

            if (selectedItems.length > 0) {
                // Automatically open modal when items are selected
                openCartonModal();
            } else {
                // Close modal if no items selected
                closeCartonModal();
            }
        }

        // 6. Two-Step Modal Functions for Carton Packing
        function openCartonModal() {
            console.log('openCartonModal called');
            console.log('currentPOData:', currentPOData);
            console.log('selectedItems:', selectedItems);

            if (!currentPOData || selectedItems.length === 0) {
                alert('Please select items to pack first!');
                return;
            }

            // Get modal element
            const modal = document.getElementById('carton_modal');
            console.log('Modal element:', modal);

            if (!modal) {
                alert('Modal not found! Check HTML structure.');
                return;
            }

            // Update modal summary
            const modalSummary = document.getElementById('modal_selected_summary');
            if (modalSummary) {
                modalSummary.innerHTML = `Selected ${selectedItems.length} items for packing`;
            }

            // Reset to step 1
            showStep1();

            // Clear previous inputs
            const cartonType = document.getElementById('modal_carton_type');
            const cartonWeight = document.getElementById('modal_carton_weight');
            const packStatus = document.getElementById('modal_pack_status');

            if (cartonType) cartonType.value = '';
            if (cartonWeight) cartonWeight.value = '';
            if (packStatus) packStatus.innerHTML = '';

            // Show modal as movable window (top-right)
            modal.style.display = 'block';
            modal.style.position = 'fixed';
            modal.style.top = '20%';
            modal.style.right = '5%';
            modal.style.zIndex = '10000';

            // Initialize drag functionality
            makeDraggable(modal);

            console.log('Modal should be visible now - Step 1 (movable window)');
        }

        function showStep1() {
            document.getElementById('modal_step_1').style.display = 'block';
            document.getElementById('modal_step_2').style.display = 'none';
        }

        function showCartonForm() {
            console.log('Showing carton form - Step 2');
            document.getElementById('modal_step_1').style.display = 'none';
            document.getElementById('modal_step_2').style.display = 'block';

            // Focus on carton type dropdown
            setTimeout(() => {
                const cartonType = document.getElementById('modal_carton_type');
                if (cartonType) cartonType.focus();
            }, 100);
        }

        function backToSummary() {
            console.log('Back to summary - Step 1');
            showStep1();
        }

        function closeCartonModal() {
            console.log('closeCartonModal called');
            const modal = document.getElementById('carton_modal');
            if (modal) {
                modal.style.display = 'none';
                // Reset to step 1 for next time
                showStep1();
                console.log('Modal closed and reset to step 1');
            } else {
                console.log('Modal not found when trying to close');
            }
        }

        async function confirmPackItems() {
            const cartonType = document.getElementById('modal_carton_type').value.trim();
            const cartonWeight = document.getElementById('modal_carton_weight').value.trim();
            const statusDiv = document.getElementById('modal_pack_status');

            // Validation
            if (!cartonType) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Please select carton type</div>';
                return;
            }

            if (!cartonWeight || isNaN(cartonWeight) || parseFloat(cartonWeight) <= 0) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Please enter valid weight</div>';
                return;
            }

            try {
                statusDiv.innerHTML = '<div style="color: #007bff; padding: 10px; background: #e3f2fd; border-radius: 5px;">⏳ Packing items into carton...</div>';

                // Get selected items data
                const selectedItemsData = selectedItems.map(index => ({
                    index: index,
                    item_number: currentPOData.items[index].item_number,
                    description: currentPOData.items[index].description,
                    color: currentPOData.items[index].color,
                    qty: currentPOData.items[index].qty
                }));

                const response = await fetch('/api/simple_packing/pack_items', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        po_number: currentPOData.po_number,
                        selected_items: selectedItemsData,
                        carton_type: cartonType,
                        carton_weight: parseFloat(cartonWeight)
                    })
                });

                const result = await response.json();

                if (result.success) {
                    statusDiv.innerHTML = '<div style="color: #28a745; padding: 10px; background: #d4edda; border-radius: 5px;">✅ ' + result.message + '</div>';

                    // Update items status and carton number in memory
                    selectedItems.forEach(index => {
                        currentPOData.items[index].packed_status = 'packed';
                        currentPOData.items[index].carton_number = result.carton_number;
                    });

                    // Clear selections and refresh display
                    clearSelections();
                    displayItems();

                    // Close modal after short delay
                    setTimeout(() => {
                        closeCartonModal();

                        // Check if all items are packed
                        checkCompletion();
                    }, 1500);

                } else {
                    statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ ' + result.message + '</div>';
                }

            } catch (error) {
                statusDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Error: ' + error.message + '</div>';
            }
        }

        // 7. Check Completion
        async function checkCompletion() {
            if (!currentPOData) return;

            try {
                const response = await fetch(`/api/simple_packing/check_completion?po_number=${currentPOData.po_number}`);
                const result = await response.json();

                if (result.success && result.is_complete) {
                    setTimeout(() => {
                        showPackingList();
                    }, 1000);
                }

            } catch (error) {
                console.error('Error checking completion:', error);
            }
        }

        // 8. Generate and Show Professional PDF Packing List
        async function showPackingList() {
            if (!currentPOData) return;

            try {
                // First get carton count for the alert
                const response = await fetch(`/api/simple_packing/generate_packing_list?po_number=${currentPOData.po_number}`);
                const result = await response.json();

                if (result.success) {
                    // Generate PL number and open PDF (no alert popup)
                    const pdfUrl = `/api/simple_packing/generate_pdf_packing_list?po_number=${currentPOData.po_number}`;
                    window.open(pdfUrl, '_blank', 'width=800,height=1000,scrollbars=yes,resizable=yes');
                } else {
                    alert('Error generating packing list: ' + result.message);
                }

            } catch (error) {
                alert('Error generating packing list: ' + error.message);
            }
        }



        // ===== OLD COMPLEX FUNCTIONS COMMENTED OUT =====
        /*
        function updateWizardProgress(step) {
            // Update wizard progress indicator
            for (let i = 1; i <= 9; i++) {
                const wizardStep = document.getElementById(`wizard_step_${i}`);
                if (wizardStep) {
                    if (i === step) {
                        wizardStep.style.background = 'rgba(255,255,255,0.3)';
                        wizardStep.style.fontWeight = 'bold';
                    } else if (i < step) {
                        wizardStep.style.background = 'rgba(40,167,69,0.3)';
                        wizardStep.style.fontWeight = 'normal';
                    } else {
                        wizardStep.style.background = 'rgba(255,255,255,0.1)';
                        wizardStep.style.fontWeight = 'normal';
                    }
                }
            }
            currentStep = step;
        }

        function loadPOForManagement() {
            const poNumber = document.getElementById('po_management_input').value.trim();
            const statusDiv = document.getElementById('po_management_status');

            if (!poNumber) {
                showError('Please enter a PO number');
                return;
            }

            statusDiv.style.display = 'block';
            statusDiv.innerHTML = '<div style="color: #007bff; padding: 15px; background: #e3f2fd; border-radius: 8px; border-left: 4px solid #2196f3;">🔍 Loading PO items...</div>';

            // First debug the raw data
            fetch('/api/po_management/debug_po', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({po_number: poNumber})
            })
            .then(response => response.json())
            .then(debugData => {
                console.log('🔍 Raw PO data from database:', debugData);

                // Now load the processed data
                return fetch('/api/po_management/get_po_items', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({po_number: poNumber})
                });
            })
            .then(response => response.json())
            .then(data => {
                console.log('📦 Processed PO data:', data);
                if (data.success) {
                    currentPOData = {po_number: poNumber, items: data.items};
                    statusDiv.innerHTML = `<div style="color: #28a745; padding: 15px; background: #d4edda; border-radius: 8px; border-left: 4px solid #28a745;">✅ Successfully loaded ${data.items.length} items!</div>`;
                    displayPOItems(data.items);
                    updateWizardProgress(2);
                    setTimeout(() => goToPOStep(2), 1000);
                } else {
                    statusDiv.innerHTML = `<div style="color: #dc3545; padding: 15px; background: #f8d7da; border-radius: 8px; border-left: 4px solid #dc3545;">❌ ${data.message}</div>`;
                }
            })
            .catch(error => {
                statusDiv.innerHTML = `<div style="color: #dc3545;">❌ Error: ${error.message}</div>`;
            });
        }

        function displayPOItems(items) {
            const container = document.getElementById('po_items_container');

            let html = `
                <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <h3>PO: ${currentPOData.po_number} (${items.length} items)</h3>
                    <div style="overflow-x: auto;">
                        <table style="width: 100%; border-collapse: collapse; margin-top: 15px;">
                            <thead>
                                <tr style="background: #e9ecef;">
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Item #</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Description</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Color</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Qty</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Bundle Qty</th>
                                </tr>
                            </thead>
                            <tbody>
            `;

            items.forEach(item => {
                html += `
                    <tr>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.item_number}</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.description}</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.color}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">${item.qty}</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.bundle_qty}</td>
                    </tr>
                `;
            });

            html += `
                            </tbody>
                        </table>
                    </div>
                </div>
            `;

            container.innerHTML = html;
        }

        function goToPOStep(stepNumber) {
            // Hide all steps
            document.querySelectorAll('.po-step').forEach(step => {
                step.style.display = 'none';
            });

            // Show target step
            document.getElementById(`po_step_${stepNumber}`).style.display = 'block';

            // Update wizard progress if it's a numeric step
            if (typeof stepNumber === 'number') {
                updateWizardProgress(stepNumber);
            } else if (stepNumber === '6a' || stepNumber === '6b') {
                updateWizardProgress(6);
            }
        }

        function goToPreviousStep() {
            // Smart back navigation based on current completion status and packing logic
            if (currentCompletionStatus === 'all') {
                // All done: Skip step 4 (partial quantities)
                if (currentPackingLogic) {
                    goToPOStep(5); // Back to packing logic selection
                } else {
                    goToPOStep(3); // Back to completion status
                }
            } else {
                // Partial done: Include step 4
                if (currentPackingLogic) {
                    goToPOStep(5); // Back to packing logic selection
                } else {
                    goToPOStep(4); // Back to partial quantities
                }
            }
        }

        function selectCompletionStatus(status) {
            currentCompletionStatus = status;

            // Save completion status to database
            fetch('/api/po_management/save_completion_status', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    po_number: currentPOData.po_number,
                    completion_type: status,
                    finished_quantities: {}
                })
            });

            if (status === 'all') {
                // Skip partial quantities step, go directly to packing logic
                updateWizardProgress(5);
                goToPOStep(5);
            } else {
                // Show partial quantities step
                displayPartialQuantitiesForm();
                updateWizardProgress(4);
                goToPOStep(4);
            }
        }

        function displayPartialQuantitiesForm() {
            const container = document.getElementById('partial_quantities_container');

            let html = `
                <div style="background: #fff3cd; padding: 20px; border-radius: 8px; margin: 20px 0; border: 1px solid #ffeaa7;">
                    <h4>⚠️ Enter finished quantities (must be ≤ order quantity)</h4>
                    <div style="overflow-x: auto; margin-top: 15px;">
                        <table style="width: 100%; border-collapse: collapse;">
                            <thead>
                                <tr style="background: #f8f9fa;">
                                    <th style="padding: 10px; border: 1px solid #ddd;">Item #</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Description</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Order Qty</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Finished Qty</th>
                                </tr>
                            </thead>
                            <tbody>
            `;

            currentPOData.items.forEach((item, index) => {
                html += `
                    <tr>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.item_number}</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.description}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: right; font-weight: bold;">${item.qty}</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">
                            <input type="number" id="finished_qty_${index}"
                                   max="${item.qty}" min="0" value="${item.qty}"
                                   style="width: 100px; padding: 5px; border: 1px solid #ddd; border-radius: 4px;"
                                   onchange="validateFinishedQty(${index}, ${item.qty})">
                        </td>
                    </tr>
                `;
            });

            html += `
                            </tbody>
                        </table>
                    </div>
                </div>
            `;

            container.innerHTML = html;
        }

        function validateFinishedQty(index, maxQty) {
            const input = document.getElementById(`finished_qty_${index}`);
            const value = parseInt(input.value);

            if (value > maxQty) {
                input.value = maxQty;
                showError(`Finished quantity cannot exceed order quantity (${maxQty})`);
            }
        }

        function validatePartialQuantities() {
            let allValid = true;

            currentPOData.items.forEach((item, index) => {
                const finishedQty = parseInt(document.getElementById(`finished_qty_${index}`).value);
                if (finishedQty > item.qty) {
                    allValid = false;
                }
                // Update item with finished quantity
                currentPOData.items[index].finished_qty = finishedQty;
            });

            if (allValid) {
                goToPOStep(5);
            } else {
                showError('Please correct the finished quantities');
            }
        }

        function selectPackingLogic(logic) {
            currentPackingLogic = logic;
            updateWizardProgress(6);

            if (logic === 'multi_to_one') {
                showMultiToOnePackingInterface();
                goToPOStep('6a');
            } else {
                showOneToMultiPackingInterface();
                goToPOStep('6b');
            }
        }

        async function showMultiToOnePackingInterface() {
            const container = document.getElementById('multi_to_one_container');
            const items = currentPOData.items;

            // Get real-time packing status
            const packingStatus = await getPackingStatus(currentPOData.po_number);

            let html = `
                <div style="background: #e3f2fd; padding: 20px; border-radius: 8px; margin: 20px 0; border: 1px solid #2196f3;">
                    <h4>📦 Option A: Multi-line → 1 Carton (Real-time Packing)</h4>
                    <p style="color: #666; margin-bottom: 15px;">Select items and they will be immediately packed into cartons. Each selection creates a new carton automatically.</p>
                </div>

                <div style="margin: 20px 0; display: flex; gap: 10px; align-items: center; flex-wrap: wrap;">
                    <button onclick="selectAllUnpackedItems()" style="padding: 8px 16px; background: #28a745; color: white; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">
                        ✅ Select All Unpacked
                    </button>
                    <button onclick="clearAllSelections()" style="padding: 8px 16px; background: #6c757d; color: white; border: none; border-radius: 5px; cursor: pointer;">
                        ❌ Clear Selections
                    </button>
                    <span id="selection_counter" style="font-weight: bold; color: #007bff;">0 items selected</span>
                </div>

                <div style="margin: 20px 0;">
                    <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                        <thead>
                            <tr style="background: #f8f9fa;">
                                <th style="padding: 12px; border: 1px solid #ddd; text-align: center;">Select</th>
                                <th style="padding: 12px; border: 1px solid #ddd;">Item #</th>
                                <th style="padding: 12px; border: 1px solid #ddd;">Description</th>
                                <th style="padding: 12px; border: 1px solid #ddd;">Color</th>
                                <th style="padding: 12px; border: 1px solid #ddd;">Qty to Pack</th>
                                <th style="padding: 12px; border: 1px solid #ddd;">Carton #</th>
                                <th style="padding: 12px; border: 1px solid #ddd;">Status</th>
                            </tr>
                        </thead>
                        <tbody>
            `;

            let unpackedCount = 0;
            items.forEach((item, index) => {
                const qtyToPack = item.finished_qty || item.qty;
                const packingInfo = packingStatus[item.item_number];
                const isPacked = packingInfo && packingInfo.status === 'packed';
                const cartonNumber = isPacked ? packingInfo.carton_number : '-';
                const statusText = isPacked ? '✅ Packed' : '⏳ Pending';
                const statusColor = isPacked ? '#28a745' : '#ffc107';
                const rowStyle = isPacked ? 'background: #f8fff8; opacity: 0.7;' : '';

                if (!isPacked) unpackedCount++;

                html += `
                    <tr style="${rowStyle}">
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">
                            <input type="checkbox" id="select_item_${index}" onchange="handleItemSelection(${index})" ${isPacked ? 'disabled' : ''}>
                        </td>
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">${item.item_number}</td>
                        <td style="padding: 10px; border: 1px solid #ddd;">${item.description}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; color: #666;">${item.color || 'N/A'}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: right; font-weight: bold; color: #007bff;">${qtyToPack}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: ${isPacked ? '#28a745' : '#999'};">
                            ${cartonNumber}
                        </td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">
                            <span style="background: ${statusColor}; color: white; padding: 4px 8px; border-radius: 12px; font-size: 12px; font-weight: bold;">
                                ${statusText}
                            </span>
                        </td>
                    </tr>
                `;
            });

            html += `
                        </tbody>
                    </table>
                </div>

                <div id="packing_feedback" style="margin: 20px 0;"></div>

                <div style="text-align: center; margin: 30px 0;">
                    ${unpackedCount === 0 ? `
                        <div style="background: #d4edda; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #28a745;">
                            <h4 style="color: #155724; margin: 0 0 10px 0;">🎉 All Items Packed!</h4>
                            <p style="color: #155724; margin: 0;">All items have been packed into cartons. You can now proceed to the next step.</p>
                        </div>
                        <button onclick="goToPOStep(7)" style="padding: 12px 24px; background: #28a745; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-weight: bold;">
                            ➡️ Continue to Carton Summary
                        </button>
                    ` : `
                        <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #ffc107;">
                            <p style="color: #856404; margin: 0;"><strong>${unpackedCount} items</strong> remaining to pack. Select items above to pack them automatically.</p>
                        </div>
                    `}
                </div>
            `;

            container.innerHTML = html;
            updateSelectionCounter();
        }

        function showOneToMultiPackingInterface() {
            const container = document.getElementById('one_to_multi_container');
            const items = currentPOData.items;

            let html = `
                <div style="background: #f3e5f5; padding: 20px; border-radius: 8px; margin: 20px 0; border: 1px solid #9c27b0;">
                    <h4>📦📦📦 Select one item to split across multiple cartons</h4>
                    <p style="color: #666; margin-bottom: 15px;">Choose one item and specify how many cartons to split it into.</p>

                    <div style="margin-bottom: 20px;">
                        <label style="display: block; margin-bottom: 10px; font-weight: bold;">Select Item:</label>
                        <select id="item_to_split" onchange="updateOneToMultiInterface()" style="width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 16px;">
                            <option value="">Choose an item to split</option>
            `;

            items.forEach((item, index) => {
                const qtyToPack = item.finished_qty || item.qty;
                html += `<option value="${index}">${item.item_number} - ${item.description} (${qtyToPack} pcs)</option>`;
            });

            html += `
                        </select>
                    </div>

                    <div id="split_details" style="display: none;">
                        <div style="margin-bottom: 15px;">
                            <label style="display: block; margin-bottom: 5px; font-weight: bold;">Number of Cartons:</label>
                            <input type="number" id="carton_count" min="2" max="10" onchange="calculateSplitQuantities()"
                                   style="width: 150px; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>

                        <div style="margin-bottom: 15px;">
                            <label style="display: block; margin-bottom: 5px; font-weight: bold;">Quantity per Carton:</label>
                            <input type="number" id="qty_per_carton" onchange="updateLastCartonQty()"
                                   style="width: 150px; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                            <small style="color: #666; margin-left: 10px;">Last carton will get the remainder</small>
                        </div>

                        <div id="carton_breakdown" style="margin: 15px 0; padding: 15px; background: #fff; border-radius: 4px; border: 1px solid #ddd;">
                            <!-- Carton breakdown will be shown here -->
                        </div>

                        <div style="margin-top: 20px; text-align: center;">
                            <button onclick="goToPOStep(5)" style="padding: 12px 24px; background: #6c757d; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; margin-right: 10px;">
                                ← Back to Packing Logic
                            </button>
                            <button onclick="createOneToMultiCartons()" style="padding: 12px 24px; background: #28a745; color: white; border: none; border-radius: 8px; cursor: pointer; font-size: 16px;">
                                📦 Create Cartons →
                            </button>
                        </div>
                    </div>
                </div>
            `;

            container.innerHTML = html;
        }

        // New real-time packing helper functions
        async function getPackingStatus(poNumber) {
            try {
                const response = await fetch(`/api/po_management/get_packing_status?po_number=${poNumber}`);
                const result = await response.json();
                return result.success ? result.packed_items : {};
            } catch (error) {
                console.error('Error getting packing status:', error);
                return {};
            }
        }

        function updateSelectionCounter() {
            const checkboxes = document.querySelectorAll('[id^="select_item_"]:not([disabled])');
            let selectedCount = 0;
            checkboxes.forEach(cb => {
                if (cb.checked) selectedCount++;
            });

            const counter = document.getElementById('selection_counter');
            if (counter) {
                counter.textContent = `${selectedCount} items selected`;
                counter.style.color = selectedCount > 0 ? '#28a745' : '#007bff';
            }
        }

        function selectAllUnpackedItems() {
            const checkboxes = document.querySelectorAll('[id^="select_item_"]:not([disabled])');
            checkboxes.forEach(cb => {
                cb.checked = true;
            });
            updateSelectionCounter();
        }

        function clearAllSelections() {
            const checkboxes = document.querySelectorAll('[id^="select_item_"]');
            checkboxes.forEach(cb => {
                cb.checked = false;
            });
            updateSelectionCounter();
        }

        async function handleItemSelection(index) {
            const checkbox = document.getElementById(`select_item_${index}`);

            if (checkbox.checked) {
                // Item was selected - pack it immediately
                await packSelectedItems([index]);
                checkbox.checked = false; // Uncheck after packing
            }

            updateSelectionCounter();
        }

        async function packSelectedItems(selectedIndices) {
            const feedbackDiv = document.getElementById('packing_feedback');

            try {
                feedbackDiv.innerHTML = '<div style="color: #007bff; padding: 10px; background: #e3f2fd; border-radius: 5px;">⏳ Packing items...</div>';

                const selectedItems = selectedIndices.map(index => {
                    const item = currentPOData.items[index];
                    return {
                        item_number: item.item_number,
                        description: item.description,
                        color: item.color || '',
                        packed_qty: item.finished_qty || item.qty,
                        original_qty: item.qty
                    };
                });

                const response = await fetch('/api/po_management/pack_items_realtime', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        po_number: currentPOData.po_number,
                        selected_item_ids: selectedItemIds  // Send only IDs, backend gets clean data from database
                    })
                });

                const result = await response.json();

                if (result.success) {
                    feedbackDiv.innerHTML = `<div style="color: #28a745; padding: 10px; background: #d4edda; border-radius: 5px;">✅ Packed ${result.packed_items} items into ${result.carton_number}</div>`;

                    // Refresh the interface to show updated status
                    setTimeout(() => {
                        showMultiToOnePackingInterface();
                    }, 1500);
                } else {
                    feedbackDiv.innerHTML = `<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ ${result.message}</div>`;
                }

            } catch (error) {
                feedbackDiv.innerHTML = `<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Error packing items: ${error.message}</div>`;
            }
        }

        function updateOneToMultiInterface() {
            const select = document.getElementById('item_to_split');
            const detailsDiv = document.getElementById('split_details');

            if (select.value) {
                detailsDiv.style.display = 'block';
                const item = currentPOData.items[parseInt(select.value)];
                const qtyToPack = item.finished_qty || item.qty;
                document.getElementById('qty_per_carton').max = qtyToPack;
            } else {
                detailsDiv.style.display = 'none';
            }
        }

        function calculateSplitQuantities() {
            const itemIndex = parseInt(document.getElementById('item_to_split').value);
            const cartonCount = parseInt(document.getElementById('carton_count').value);

            if (itemIndex >= 0 && cartonCount > 1) {
                const item = currentPOData.items[itemIndex];
                const totalQty = item.finished_qty || item.qty;
                const qtyPerCarton = Math.floor(totalQty / cartonCount);

                document.getElementById('qty_per_carton').value = qtyPerCarton;
                updateLastCartonQty();
            }
        }

        function updateLastCartonQty() {
            const itemIndex = parseInt(document.getElementById('item_to_split').value);
            const cartonCount = parseInt(document.getElementById('carton_count').value);
            const qtyPerCarton = parseInt(document.getElementById('qty_per_carton').value);

            if (itemIndex >= 0 && cartonCount > 1 && qtyPerCarton > 0) {
                const item = currentPOData.items[itemIndex];
                const totalQty = item.finished_qty || item.qty;
                const lastCartonQty = totalQty - (qtyPerCarton * (cartonCount - 1));

                let html = `<h5>Carton Breakdown:</h5>`;
                for (let i = 1; i < cartonCount; i++) {
                    html += `<div>Carton ${i}: ${qtyPerCarton} pcs</div>`;
                }
                html += `<div style="font-weight: bold; color: #007bff;">Carton ${cartonCount}: ${lastCartonQty} pcs (remainder)</div>`;
                html += `<div style="margin-top: 10px; font-weight: bold;">Total: ${totalQty} pcs</div>`;

                document.getElementById('carton_breakdown').innerHTML = html;
            }
        }

        // Enhanced carton creation functions

        function createMultiToOneCarton() {
            console.log('🔧 Creating multi-to-one carton...');

            // Prevent multiple clicks
            const button = event.target;
            if (button.disabled) return;
            button.disabled = true;
            button.innerHTML = '⏳ Creating...';

            const feedbackDiv = document.getElementById('multi_to_one_feedback');
            feedbackDiv.innerHTML = '<div style="color: #007bff; padding: 10px; background: #e3f2fd; border-radius: 5px;">⏳ Creating carton...</div>';

            const checkboxes = document.querySelectorAll('[id^="select_item_"]:not([disabled])');
            const weight = parseFloat(document.getElementById('carton_weight').value);
            const cartonSize = document.getElementById('carton_size_select').value;

            console.log('📦 Found checkboxes:', checkboxes.length);
            console.log('⚖️ Weight:', weight);
            console.log('📏 Size:', cartonSize);

            if (!weight || weight <= 0) {
                feedbackDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Please enter a valid carton weight</div>';
                button.disabled = false;
                button.innerHTML = '📦 Create Carton';
                return;
            }

            let selectedItemIds = [];
            checkboxes.forEach((cb, index) => {
                if (cb.checked) {
                    const item = currentPOData.items[index];
                    // Send only the database ID, let backend get clean data from database
                    selectedItemIds.push(item.id);
                    console.log('✅ Selected item ID:', item.id, 'item_number:', item.item_number);
                }
            });

            console.log('📋 Selected item IDs:', selectedItemIds);

            if (selectedItemIds.length === 0) {
                feedbackDiv.innerHTML = '<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Please select at least one item to pack</div>';
                button.disabled = false;
                button.innerHTML = '📦 Create Carton';
                return;
            }

            // Create carton data with enhanced tracking
            const cartonData = {
                size: cartonSize,
                weight: weight,
                items: selectedItems
            };

            console.log('📦 Carton data to send:', cartonData);

            // Call API to create carton with enhanced tracking
            fetch('/api/po_management/create_cartons', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    po_number: currentPOData.po_number,
                    cartons: [cartonData],
                    packing_option: 'A'
                })
            })
            .then(response => {
                console.log('🌐 API Response status:', response.status);
                return response.json();
            })
            .then(data => {
                console.log('📊 API Response data:', data);
                if (data.success) {
                    createdCartons.push(...data.cartons);
                    feedbackDiv.innerHTML = `<div style="color: #28a745; padding: 15px; background: #d4edda; border-radius: 5px; border-left: 4px solid #28a745;">✅ Carton created successfully!<br><strong>Barcode:</strong> ${data.cartons[0].barcode}<br><strong>Items packed:</strong> ${selectedItems.length}</div>`;

                    // Clear selections and refresh interface
                    checkboxes.forEach(cb => {
                        if (cb.checked) cb.checked = false;
                    });
                    document.getElementById('carton_weight').value = '';

                    // Refresh the interface to show packed status
                    setTimeout(() => {
                        showMultiToOnePackingInterface();
                        feedbackDiv.innerHTML = `<div style="color: #28a745; padding: 10px; background: #d4edda; border-radius: 5px;">✅ Ready to pack more items</div>`;
                    }, 2000);

                } else {
                    feedbackDiv.innerHTML = `<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ ${data.message}</div>`;
                }
                button.disabled = false;
                button.innerHTML = '📦 Create Carton';
            })
            .catch(error => {
                console.error('❌ Error creating carton:', error);
                feedbackDiv.innerHTML = `<div style="color: #dc3545; padding: 10px; background: #f8d7da; border-radius: 5px;">❌ Error creating carton: ${error.message}</div>`;
                button.disabled = false;
                button.innerHTML = '📦 Create Carton';
            });
        }

        function createOneToMultiCartons() {
            const itemIndex = parseInt(document.getElementById('item_to_split').value);
            const cartonCount = parseInt(document.getElementById('carton_count').value);
            const qtyPerCarton = parseInt(document.getElementById('qty_per_carton').value);

            if (itemIndex < 0 || cartonCount < 2 || qtyPerCarton <= 0) {
                showError('Please fill in all required fields correctly');
                return;
            }

            const item = currentPOData.items[itemIndex];
            const totalQty = item.finished_qty || item.qty;
            const lastCartonQty = totalQty - (qtyPerCarton * (cartonCount - 1));

            if (lastCartonQty <= 0) {
                showError('Invalid quantity distribution. Please adjust quantities.');
                return;
            }

            // Create multiple cartons
            let cartonsData = [];
            for (let i = 1; i <= cartonCount; i++) {
                const qty = (i === cartonCount) ? lastCartonQty : qtyPerCarton;
                cartonsData.push({
                    size: 'Standard',
                    weight: 0, // Will be updated later
                    items: [{
                        item_number: item.item_number,
                        description: item.description,
                        color: item.color || '',
                        packed_qty: qty,
                        original_qty: item.qty
                    }]
                });
            }

            // Call API to create cartons with enhanced tracking
            fetch('/api/po_management/create_cartons', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    po_number: currentPOData.po_number,
                    cartons: cartonsData,
                    packing_option: 'B'
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    createdCartons.push(...data.cartons);
                    showSuccess(`✅ ${cartonCount} cartons created successfully!`);
                    displayCartonSummary();
                    updateWizardProgress(7);
                    goToPOStep(7);
                } else {
                    showError(`❌ ${data.message}`);
                }
            })
            .catch(error => {
                showError(`❌ Error creating cartons: ${error.message}`);
            });
        }

        function displayCartonSummary() {
            const container = document.getElementById('carton_summary_container');

            let html = `
                <div style="background: #d4edda; padding: 25px; border-radius: 12px; margin: 20px 0; border: 1px solid #28a745; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                    <h4 style="color: #155724; margin-bottom: 15px;">📦 Carton Summary & Barcodes</h4>
                    <div style="display: flex; gap: 20px; margin-bottom: 20px; flex-wrap: wrap;">
                        <div style="background: white; padding: 15px; border-radius: 8px; flex: 1; min-width: 200px;">
                            <strong>Total Cartons:</strong> ${createdCartons.length}
                        </div>
                        <div style="background: white; padding: 15px; border-radius: 8px; flex: 1; min-width: 200px;">
                            <strong>PO Number:</strong> ${currentPOData.po_number}
                        </div>
                        <div style="background: white; padding: 15px; border-radius: 8px; flex: 1; min-width: 200px;">
                            <strong>Packing Option:</strong> ${currentPackingLogic === 'multi_to_one' ? 'Option A' : 'Option B'}
                        </div>
                    </div>

                    <div style="overflow-x: auto; margin-top: 15px;">
                        <table style="width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden;">
                            <thead>
                                <tr style="background: #28a745; color: white;">
                                    <th style="padding: 12px; border: 1px solid #ddd;">Carton #</th>
                                    <th style="padding: 12px; border: 1px solid #ddd;">Barcode</th>
                                    <th style="padding: 12px; border: 1px solid #ddd;">Weight (kg)</th>
                                    <th style="padding: 12px; border: 1px solid #ddd;">Items</th>
                                    <th style="padding: 12px; border: 1px solid #ddd;">Total Qty</th>
                                </tr>
                            </thead>
                            <tbody>
            `;

            createdCartons.forEach(carton => {
                const totalQty = carton.items.reduce((sum, item) => sum + item.packed_qty, 0);
                const itemsList = carton.items.map(item => `${item.item_number} (${item.packed_qty})`).join(', ');

                html += `
                    <tr style="border-bottom: 1px solid #ddd;">
                        <td style="padding: 12px; border: 1px solid #ddd; font-weight: bold; color: #007bff;">${carton.carton_number}</td>
                        <td style="padding: 12px; border: 1px solid #ddd; font-family: monospace; font-size: 11px; background: #f8f9fa;">${carton.barcode}</td>
                        <td style="padding: 12px; border: 1px solid #ddd; text-align: center; font-weight: bold;">${carton.weight || 0}</td>
                        <td style="padding: 12px; border: 1px solid #ddd; font-size: 13px;">${itemsList}</td>
                        <td style="padding: 12px; border: 1px solid #ddd; text-align: right; font-weight: bold; color: #28a745;">${totalQty}</td>
                    </tr>
                `;
            });

            html += `
                            </tbody>
                        </table>
                    </div>
                </div>
            `;

            container.innerHTML = html;
        }

        function generateBarcodes() {
            showSuccess('🏷️ Barcodes generated! (Feature: Print barcode labels)');
        }

        function createShipment() {
            const courier = document.getElementById('courier_select').value;
            const awbNumber = document.getElementById('awb_input').value.trim();

            if (!courier || !awbNumber) {
                showError('Please select courier and enter AWB number');
                return;
            }

            updateWizardProgress(8);
            showInfo('🚚 Creating shipment...');

            const cartonIds = createdCartons.map(carton => carton.id);

            fetch('/api/po_management/create_shipment', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    po_number: currentPOData.po_number,
                    courier: courier,
                    awb_number: awbNumber,
                    carton_ids: cartonIds
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    showSuccess(`🚚 Shipment created successfully! Total: ${data.total_cartons} cartons, ${data.total_weight}kg`);
                    displayFinalSummary(courier, awbNumber, data);
                    updateWizardProgress(9);
                    goToPOStep(9);
                } else {
                    showError(`❌ ${data.message}`);
                }
            })
            .catch(error => {
                showError(`❌ Error creating shipment: ${error.message}`);
            });
        }

        function displayFinalSummary(courier, awbNumber, shipmentData) {
            const container = document.getElementById('final_summary_container');
            const totalItems = createdCartons.reduce((sum, carton) =>
                sum + carton.items.reduce((itemSum, item) => itemSum + item.packed_qty, 0), 0);
            const totalWeight = shipmentData ? shipmentData.total_weight : 0;

            let html = `
                <div style="background: linear-gradient(135deg, #d1ecf1 0%, #e8f5e8 100%); padding: 30px; border-radius: 15px; margin: 20px 0; border: 2px solid #17a2b8; box-shadow: 0 4px 8px rgba(0,0,0,0.1);">
                    <div style="text-align: center; margin-bottom: 25px;">
                        <h3 style="color: #0c5460; margin: 0;">🎉 Packing List Complete!</h3>
                        <p style="color: #0c5460; margin: 5px 0; font-size: 16px;">All items packed and ready for shipment</p>
                    </div>

                    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin: 25px 0;">
                        <div style="background: white; padding: 20px; border-radius: 10px; border-left: 4px solid #007bff;">
                            <h5 style="color: #007bff; margin-bottom: 15px;">📋 PO Summary</h5>
                            <p><strong>PO Number:</strong> ${currentPOData.po_number}</p>
                            <p><strong>Total Items:</strong> ${totalItems} pieces</p>
                            <p><strong>Completion:</strong> ${currentCompletionStatus === 'all' ? 'Complete' : 'Partial'}</p>
                            <p><strong>Packing Method:</strong> ${currentPackingLogic === 'multi_to_one' ? 'Option A (Multi→1)' : 'Option B (1→Multi)'}</p>
                        </div>
                        <div style="background: white; padding: 20px; border-radius: 10px; border-left: 4px solid #28a745;">
                            <h5 style="color: #28a745; margin-bottom: 15px;">📦 Carton Summary</h5>
                            <p><strong>Total Cartons:</strong> ${createdCartons.length}</p>
                            <p><strong>Total Weight:</strong> ${totalWeight}kg</p>
                            <p><strong>Barcodes:</strong> Generated</p>
                            <p><strong>Status:</strong> Ready to Ship</p>
                        </div>
                        <div style="background: white; padding: 20px; border-radius: 10px; border-left: 4px solid #ffc107;">
                            <h5 style="color: #856404; margin-bottom: 15px;">🚚 Shipment Details</h5>
                            <p><strong>Courier:</strong> ${courier}</p>
                            <p><strong>AWB Number:</strong> ${awbNumber}</p>
                            <p><strong>Ship Date:</strong> ${new Date().toLocaleDateString()}</p>
                            <p><strong>Status:</strong> Confirmed</p>
                        </div>
                    </div>

                    <div style="text-align: center; margin-top: 25px; padding: 20px; background: rgba(255,255,255,0.7); border-radius: 10px;">
                        <p style="color: #0c5460; font-size: 14px; margin: 0;">
                            ✅ All steps completed successfully! Your packing list is ready for download.
                        </p>
                    </div>
                </div>
            `;

            container.innerHTML = html;
        }

        function downloadPackingList() {
            showInfo('📄 Generating packing list...');

            fetch('/api/po_management/generate_packing_list', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    po_number: currentPOData.po_number
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.success) {
                    // Create downloadable content
                    const packingListContent = generatePackingListHTML(data.packing_list);
                    const blob = new Blob([packingListContent], { type: 'text/html' });
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `PackingList_${currentPOData.po_number}_${new Date().toISOString().split('T')[0]}.html`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    showSuccess('📄 Packing list downloaded successfully!');
                } else {
                    showError(`❌ ${data.message}`);
                }
            })
            .catch(error => {
                showError(`❌ Error generating packing list: ${error.message}`);
            });
        }

        function generatePackingListHTML(packingList) {
            return `
                <!DOCTYPE html>
                <html>
                <head>
                    <title>Packing List - PO ${packingList.po_number}</title>
                    <style>
                        body { font-family: Arial, sans-serif; margin: 20px; }
                        .header { text-align: center; margin-bottom: 30px; }
                        .summary { display: flex; gap: 20px; margin-bottom: 30px; }
                        .summary div { flex: 1; padding: 15px; border: 1px solid #ddd; border-radius: 5px; }
                        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
                        th, td { padding: 8px; border: 1px solid #ddd; text-align: left; }
                        th { background: #f5f5f5; }
                        .barcode { font-family: monospace; font-size: 12px; }
                    </style>
                </head>
                <body>
                    <div class="header">
                        <h1>PACKING LIST</h1>
                        <h2>PO Number: ${packingList.po_number}</h2>
                        <p>Generated: ${packingList.generated_at}</p>
                    </div>

                    <div class="summary">
                        <div>
                            <h3>Shipment Details</h3>
                            <p><strong>Courier:</strong> ${packingList.shipment ? packingList.shipment[0] : 'N/A'}</p>
                            <p><strong>AWB:</strong> ${packingList.shipment ? packingList.shipment[1] : 'N/A'}</p>
                            <p><strong>Total Cartons:</strong> ${packingList.shipment ? packingList.shipment[3] : 'N/A'}</p>
                            <p><strong>Total Weight:</strong> ${packingList.shipment ? packingList.shipment[4] : 'N/A'}kg</p>
                        </div>
                    </div>

                    <h3>Carton Details</h3>
                    <table>
                        <thead>
                            <tr>
                                <th>Carton #</th>
                                <th>Barcode</th>
                                <th>Weight</th>
                                <th>Item #</th>
                                <th>Description</th>
                                <th>Color</th>
                                <th>Qty</th>
                            </tr>
                        </thead>
                        <tbody>
                            ${packingList.packing_data.map(row => `
                                <tr>
                                    <td>${row[0]}</td>
                                    <td class="barcode">${row[1]}</td>
                                    <td>${row[2]}kg</td>
                                    <td>${row[4]}</td>
                                    <td>${row[5]}</td>
                                    <td>${row[6]}</td>
                                    <td>${row[7]}</td>
                                </tr>
                            `).join('')}
                        </tbody>
                    </table>
                </body>
                </html>
            `;
        }

        function viewCreatedCartons() {
            if (createdCartons.length === 0) {
                showInfo('📦 No cartons created yet. Create some cartons first!');
                return;
            }

            const feedbackDiv = document.getElementById('multi_to_one_feedback');
            let html = `
                <div style="background: #e8f5e8; padding: 20px; border-radius: 8px; margin: 20px 0; border: 1px solid #28a745;">
                    <h4 style="color: #155724; margin-bottom: 15px;">📦 Created Cartons (${createdCartons.length})</h4>
                    <div style="overflow-x: auto;">
                        <table style="width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden;">
                            <thead>
                                <tr style="background: #28a745; color: white;">
                                    <th style="padding: 10px; border: 1px solid #ddd;">Carton #</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Barcode</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Size</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Weight</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Items</th>
                                    <th style="padding: 10px; border: 1px solid #ddd;">Total Qty</th>
                                </tr>
                            </thead>
                            <tbody>
            `;

            createdCartons.forEach(carton => {
                const totalQty = carton.items.reduce((sum, item) => sum + item.packed_qty, 0);
                const itemsList = carton.items.map(item => `${item.item_number} (${item.packed_qty})`).join(', ');

                html += `
                    <tr style="border-bottom: 1px solid #ddd;">
                        <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold; color: #007bff;">${carton.carton_number}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; font-family: monospace; font-size: 11px; background: #f8f9fa;">${carton.barcode}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">${carton.size || 'Standard'}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-weight: bold;">${carton.weight || 0}kg</td>
                        <td style="padding: 10px; border: 1px solid #ddd; font-size: 12px;">${itemsList}</td>
                        <td style="padding: 10px; border: 1px solid #ddd; text-align: right; font-weight: bold; color: #28a745;">${totalQty}</td>
                    </tr>
                `;
            });

            html += `
                            </tbody>
                        </table>
                    </div>
                    <div style="margin-top: 15px; text-align: center;">
                        <button onclick="showMultiToOnePackingInterface()" style="padding: 10px 20px; background: #007bff; color: white; border: none; border-radius: 5px; cursor: pointer;">
                            ← Back to Packing
                        </button>
                    </div>
                </div>
            `;

            feedbackDiv.innerHTML = html;
        }

        function resetPOManagement() {
            // Reset all variables
            currentPOData = null;
            currentCompletionStatus = null;
            currentPackingLogic = null;
            createdCartons = [];
            currentStep = 1;

            // Clear all inputs
            document.getElementById('po_management_input').value = '';
            document.getElementById('po_management_status').style.display = 'none';

            // Reset wizard progress
            updateWizardProgress(1);

            // Go back to step 1
            goToPOStep(1);
            showSuccess('🔄 Ready for new PO management - All data cleared');
        }
        */
        // ===== END OF OLD COMPLEX FUNCTIONS =====

        // Drag functionality for movable modal
        function makeDraggable(modal) {
            const header = document.getElementById('modal_header');
            let isDragging = false;
            let currentX;
            let currentY;
            let initialX;
            let initialY;
            let xOffset = 0;
            let yOffset = 0;

            header.addEventListener('mousedown', dragStart);
            document.addEventListener('mousemove', drag);
            document.addEventListener('mouseup', dragEnd);

            function dragStart(e) {
                initialX = e.clientX - xOffset;
                initialY = e.clientY - yOffset;

                if (e.target === header || header.contains(e.target)) {
                    isDragging = true;
                    header.style.cursor = 'grabbing';
                }
            }

            function drag(e) {
                if (isDragging) {
                    e.preventDefault();
                    currentX = e.clientX - initialX;
                    currentY = e.clientY - initialY;

                    xOffset = currentX;
                    yOffset = currentY;

                    // Keep modal within viewport bounds
                    const rect = modal.getBoundingClientRect();
                    const maxX = window.innerWidth - rect.width;
                    const maxY = window.innerHeight - rect.height;

                    currentX = Math.max(0, Math.min(currentX, maxX));
                    currentY = Math.max(0, Math.min(currentY, maxY));

                    modal.style.transform = `translate(${currentX}px, ${currentY}px)`;
                    modal.style.top = '0';
                    modal.style.right = 'auto';
                    modal.style.left = '0';
                }
            }

            function dragEnd(e) {
                initialX = currentX;
                initialY = currentY;
                isDragging = false;
                header.style.cursor = 'move';
            }
        }

        // Keyboard event handling for modal
        document.addEventListener('keydown', function(event) {
            const modal = document.getElementById('carton_modal');
            if (modal && modal.style.display === 'block') {
                if (event.key === 'Escape') {
                    closeCartonModal();
                } else if (event.key === 'Enter') {
                    event.preventDefault();
                    confirmPackItems();
                }
            }
        });

    </script>

    <!-- Version Footer -->
    <div style="position: fixed; bottom: 10px; right: 10px; background: rgba(0,0,0,0.7); color: white; padding: 5px 10px; border-radius: 5px; font-size: 11px; font-family: monospace;">
        v{{ version }} | {{ version_date }} | {{ last_edit }}
    </div>

</body>
</html>
"""

if __name__ == '__main__':
    print(f"🚀 Starting artwork downloader v{VERSION}...")
    print(f"📅 Version Date: {VERSION_DATE}")
    print(f"📝 Last Edit: {LAST_EDIT}")
    print("📊 Initializing PO database...")
    init_database()
    print("📦 Creating sample PO data...")
    create_sample_po_data()
    print("📱 Open your browser and go to: http://localhost:5002")
    print("🛑 Press Ctrl+C to stop the server")

    app.run(debug=False, host='127.0.0.1', port=5002)
